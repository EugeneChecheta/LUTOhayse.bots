import telebot
from telebot import types
import openpyxl
import os
import random
import string
import traceback
from datetime import datetime

# ===== КОНФИГУРАЦИЯ ПУТЕЙ И ДИРЕКТОРИЙ =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_BASE = os.path.join(BASE_DIR, 'DataBase')
os.makedirs(DATA_BASE, exist_ok=True)

PHOTO_DIR = os.path.join(BASE_DIR, 'Photo', 'Products')
PHOTO_MATERIALS_DIR = os.path.join(BASE_DIR, 'Photo', 'Materials')
PHOTO_CONSTRUCTION_DIR = os.path.join(BASE_DIR, 'Photo', 'Construction')
os.makedirs(PHOTO_DIR, exist_ok=True)
os.makedirs(PHOTO_MATERIALS_DIR, exist_ok=True)
os.makedirs(PHOTO_CONSTRUCTION_DIR, exist_ok=True)

bot = telebot.TeleBot('токен')

# ===== КОНФИГУРАЦИЯ ФАЙЛОВ БАЗЫ ДАННЫХ =====
products_log = os.path.join(DATA_BASE, 'Products_log.xlsx')
materials_log = os.path.join(DATA_BASE, 'Materials_log.xlsx')
modules_log = os.path.join(DATA_BASE, 'Moduls_log.xlsx')
sofas_log = os.path.join(DATA_BASE, 'Sofas_log.xlsx')
orders_log = os.path.join(DATA_BASE, 'Orders_log.xlsx')
cart_log = os.path.join(DATA_BASE, 'Cart_log.xlsx')

site_shop = 'https://lutohouse.ru/'
site_materials = 'https://lutohouse.ru/materials.html'

# ===== КОНФИГУРАЦИЯ МЕНЕДЖЕРА ДЛЯ УВЕДОМЛЕНИЙ =====
MANAGER_CHAT_ID = 671837055  # chat_id пользователя @IIICripsesIII

# ===== ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ И КОНФИГУРАЦИИ =====
user_data = {}

module_type_names = {
    'P': 'Тайль',
    'A': 'Бауэн',
    'R': 'Бальт',
    'K': 'Царт',
    'T': 'Вельтраум',
    'F': 'Зофа',
    'J': 'Фэст',
    'G': 'Унви',
    'V': 'Вэрт',
    'Z': 'Андэр'
}

material_types = ['Букле', 'Эко-кожа', 'Велюр', 'Рогожка']
material_type_mapping = {
    'Букле': 'Букле',
    'Эко-кожа': 'Эко-кожа',
    'Велюр': 'Велюто',
    'Рогожка': 'Ромео'
}

field_descriptions = {
    'fio': 'ФИО',
    'phone': 'телефона',
    'address': 'адреса',
    'comment': 'комментария'
}

field_limits = {
    'fio': 100,
    'phone': 20,
    'address': 200,
    'comment': 500
}

# ===== СПИСОК МЕНЕДЖЕРОВ =====
managers = [
    {'name': 'Кирилл', 'link': 'https://t.me/luytui'}
]


# ===== СОЗДАНИЕ ТАБЛИЦЫ КОРЗИН (ИСПРАВЛЕНО) =====
def create_cart_table():
    """Создает таблицу для корзин пользователей, если она не существует"""
    if not os.path.exists(cart_log):
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = [
            "Мессенджер",          # столбец 1
            "ID записи",           # столбец 2
            "ID чата",             # столбец 3
            "Тип позиции",         # столбец 4
            "Код позиции",         # столбец 5
            "Название позиции",    # столбец 6
            "Код материала",       # столбец 7
            "Тип материала",       # столбец 8
            "Цвет материала",      # столбец 9
            "Цена",                # столбец 10
            "Дата добавления"      # столбец 11
        ]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        wb.save(cart_log)
        print("✅ Таблица корзин создана (с исправленными заголовками)")


create_cart_table()


# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С КОРЗИНОЙ (ИСПРАВЛЕНЫ) =====
def save_to_cart(chat_id, item_type, item_code, item_name, material_code, material_type, material_color, price):
    """Сохраняет позицию в корзину пользователя"""
    try:
        wb = openpyxl.load_workbook(cart_log)
        sheet = wb.active
        row_num = sheet.max_row + 1

        # Столбцы после добавления "Мессенджер" (1):
        # 1: Мессенджер
        # 2: ID записи
        # 3: ID чата
        # 4: Тип позиции
        # 5: Код позиции
        # 6: Название позиции
        # 7: Код материала
        # 8: Тип материала
        # 9: Цвет материала
        # 10: Цена
        # 11: Дата добавления

        sheet.cell(row=row_num, column=1, value="telegram")                     # Мессенджер
        sheet.cell(row=row_num, column=2, value=row_num - 1)                    # ID записи
        sheet.cell(row=row_num, column=3, value=chat_id)                        # ID чата
        sheet.cell(row=row_num, column=4, value=item_type)                      # Тип позиции
        sheet.cell(row=row_num, column=5, value=item_code)                      # Код позиции
        sheet.cell(row=row_num, column=6, value=item_name)                      # Название позиции
        sheet.cell(row=row_num, column=7, value=material_code)                  # Код материала
        sheet.cell(row=row_num, column=8, value=material_type)                  # Тип материала
        sheet.cell(row=row_num, column=9, value=material_color)                 # Цвет материала
        sheet.cell(row=row_num, column=10, value=price)                         # Цена
        sheet.cell(row=row_num, column=11, value=datetime.now().strftime("%d.%m.%Y %H:%M:%S"))  # Дата

        wb.save(cart_log)
        return True
    except Exception as e:
        print(f"Ошибка сохранения в корзину: {e}")
        return False


def get_user_cart(chat_id):
    """Получает корзину пользователя"""
    try:
        wb = openpyxl.load_workbook(cart_log)
        sheet = wb.active
        cart_items = []

        for row in range(2, sheet.max_row + 1):
            # ID чата теперь в столбце 3
            if str(sheet.cell(row=row, column=3).value) == str(chat_id):
                item = {
                    'db_id': sheet.cell(row=row, column=2).value,          # ID записи
                    'item_type': sheet.cell(row=row, column=4).value,      # Тип позиции
                    'item_code': sheet.cell(row=row, column=5).value,      # Код позиции
                    'item_name': sheet.cell(row=row, column=6).value,      # Название позиции
                    'material_code': sheet.cell(row=row, column=7).value,  # Код материала
                    'material_type': sheet.cell(row=row, column=8).value,  # Тип материала
                    'material_color': sheet.cell(row=row, column=9).value, # Цвет материала
                    'price': sheet.cell(row=row, column=10).value,         # Цена
                    'date_added': sheet.cell(row=row, column=11).value     # Дата добавления
                }
                cart_items.append(item)

        return cart_items
    except Exception as e:
        print(f"Ошибка получения корзины: {e}")
        return []


def remove_from_cart(chat_id, item_index):
    """Удаляет позицию из корзины"""
    try:
        wb = openpyxl.load_workbook(cart_log)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            # ID чата в столбце 3, ID записи в столбце 2
            if str(sheet.cell(row=row, column=3).value) == str(chat_id) and sheet.cell(row=row, column=2).value == item_index:
                sheet.delete_rows(row)
                wb.save(cart_log)
                return True

        return False
    except Exception as e:
        print(f"Ошибка удаления из корзины: {e}")
        return False


def clear_user_cart(chat_id):
    """Очищает корзину пользователя"""
    try:
        wb = openpyxl.load_workbook(cart_log)
        sheet = wb.active

        rows_to_delete = []
        for row in range(2, sheet.max_row + 1):
            # ID чата в столбце 3
            if str(sheet.cell(row=row, column=3).value) == str(chat_id):
                rows_to_delete.append(row)

        for row in reversed(rows_to_delete):
            sheet.delete_rows(row)

        wb.save(cart_log)
        return True
    except Exception as e:
        print(f"Ошибка очистки корзины: {e}")
        return False


# ===== КОНФИГУРАЦИИ КЛАВИАТУР =====
keyboard_configs = {
    'main': {
        'buttons': ['Сконструировать диван', 'Мои диваны', 'Ввести код дивана',
                    'Оформить заказ', 'Мои заказы', 'Перейти на сайт',
                    'Связаться с менеджером', 'Перезапустить бота'],
        'rows': [2, 4, 6, 8]
    },
    'sofa_type': {'buttons': ['Назад'], 'rows': [1]},
    'base_module': {'buttons': ['Назад'], 'rows': []},
    'construction': {
        'buttons': ['Добавить модуль', 'Подробнее о модуле', 'Завершить конструирование', 'Удалить последний модуль',
                    'Сохранить диван', 'Добавить в корзину', 'Оформить заказ', 'Назад'],
        'rows': [1, 2, 4, 6, 8]
    },
    'extension_module': {'buttons': ['Назад'], 'rows': []},
    'editing_sofa': {
        'buttons': ['Редактировать диван', 'Сохранить как новый', 'Добавить в заказ', 'Назад'],
        'rows': [1, 2, 4]
    },
    'saved_sofas': {
        'buttons': ['Сконструировать диван', 'Ввести код дивана', 'Назад'],
        'rows': [1, 3]
    },
    'order_menu': {
        'buttons': ['Добавить диван в заказ', 'Добавить позицию', 'Посмотреть корзину', 'Ввести код дивана',
                    'Отправить заказ на обработку', 'Назад', 'Перезапустить бота', 'Перейти на сайте'],
        'rows': [1, 3, 5, 7]
    },
    'sofa_code_input': {'buttons': ['Назад'], 'rows': [1]},
    'material_selection': {'buttons': ['Назад'], 'rows': []},
    'material_navigation': {
        'buttons': ['Предыдущий', 'Следующий', 'Назад', 'Добавить в корзину'],
        'rows': [2, 4]
    },
    'material_code_input': {'buttons': ['Назад'], 'rows': [1]},
    'material_confirmation': {
        'buttons': ['Выбрать этот материал', 'Ввести код заново', 'Назад'],
        'rows': [2, 3]
    },
    'cart_management': {
        'buttons': ['Добавить позицию', 'Ввести код дивана', 'Отправить заказ на обработку',
                    'Назад', 'Перезапустить бота', 'Перейти на сайте'],
        'rows': [1, 3, 5, 6]
    },
    'collecting_data': {'buttons': ['Назад'], 'rows': [1]},
    'order_confirmation': {
        'buttons': ['Подтвердить заказ', 'Отменить заказ'],
        'rows': [1]
    },
    'my_orders': {
        'buttons': ['Оформить заказ', 'Ввести код дивана', 'Назад', 'Перезапустить бота', 'Перейти на сайте'],
        'rows': [1, 3, 4, 5]
    },
    'product_type': {
        'buttons': ['Ввести код позиции', 'Назад'],
        'rows': [2]
    },
    'product_navigation': {
        'buttons': ['Предыдущий', 'Следующий', 'Назад', 'Перейти на сайте', 'Добавить в корзину'],
        'rows': [2, 4]
    },
    'product_code_input': {
        'buttons': ['Назад', 'Перейти на сайте'],
        'rows': [1]
    },
    'product_confirmation': {
        'buttons': ['Выбрать эту позицию', 'Ввести код заново', 'Назад'],
        'rows': [2, 3]
    },
    'material_method': {
        'buttons': ['Ввести код материала', 'Выбрать через бот', 'Назад', 'Перейти на сайте'],
        'rows': [2, 3, 4]
    },
    'material_method_sofa': {
        'buttons': ['Ввести код материала', 'Выбрать через бот', 'Назад'],
        'rows': [2, 3]
    },
    'material_method_product': {
        'buttons': ['Ввести код материала', 'Выбрать через бот', 'Назад'],
        'rows': [2, 3]
    },
    'material_code_input_product': {'buttons': ['Назад'], 'rows': [1]},
    'material_confirmation_product': {
        'buttons': ['Выбрать этот материал', 'Ввести код заново', 'Назад'],
        'rows': [2, 3]
    },
    'product_view_confirmation': {
        'buttons': ['Подтвердить', 'Отмена'],
        'rows': [1]
    },
    'view_cart_item': {'buttons': ['Назад в корзину'], 'rows': [1]},
    'view_saved_sofa': {'buttons': ['Назад к списку'], 'rows': [1]},
    'module_details_selection': {'buttons': ['Назад'], 'rows': []},
    'module_details_view': {
        'buttons': ['Предыдущий модуль', 'Следующий модуль', 'Назад к списку модулей'],
        'rows': [2, 3]
    },
    'order_details_navigation': {
        'buttons': ['Предыдущая позиция', 'Следующая позиция', 'Назад к заказу'],
        'rows': [1, 2, 3]
    }
}


def create_keyboard(keyboard_type, chat_id=None, extra_buttons=None, row_breaks_override=None, exclude_buttons=None,
                    available_material_types=None, available_module_types=None):
    """Создает reply-клавиатуры"""
    config = keyboard_configs.get(keyboard_type, {})
    buttons = config.get('buttons', [])[:]
    row_breaks = row_breaks_override or config.get('rows', [])

    if exclude_buttons:
        buttons = [btn for btn in buttons if btn not in exclude_buttons]

    if extra_buttons:
        buttons.extend(extra_buttons)

    if keyboard_type == 'sofa_type':
        available_types = get_available_sofa_types()
        buttons.extend(available_types)
        row_breaks = []

    elif keyboard_type == 'base_module':
        if chat_id is not None:
            sofa_type = user_data.get(chat_id, {}).get('current_sofa_type')
            if sofa_type:
                base_modules = get_base_modules(sofa_type)
                buttons.extend([module['name'] for module in base_modules])
                row_breaks = []

    elif keyboard_type == 'extension_module':
        if chat_id is not None:
            available_extensions = user_data.get(chat_id, {}).get('available_extensions', [])
            buttons.extend([module['name'] for module in available_extensions])
            row_breaks = []

    elif keyboard_type == 'material_selection':
        if available_material_types:
            buttons.extend(available_material_types)
        else:
            material_types_list = get_material_types()
            buttons.extend(material_types_list)
        row_breaks = []

    elif keyboard_type == 'product_type':
        products, unique_types = load_products()
        buttons.extend(unique_types)
        row_breaks = []

    elif keyboard_type == 'module_details_selection':
        if available_module_types:
            buttons.extend(available_module_types)
        row_breaks = []

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    if not row_breaks:
        for button in buttons:
            markup.add(types.KeyboardButton(button))
    else:
        start_idx = 0
        for end_idx in row_breaks:
            if end_idx <= len(buttons):
                row_buttons = buttons[start_idx:end_idx]
                markup.row(*[types.KeyboardButton(btn) for btn in row_buttons])
                start_idx = end_idx

        if start_idx < len(buttons):
            row_buttons = buttons[start_idx:]
            markup.row(*[types.KeyboardButton(btn) for btn in row_buttons])

    return markup


# ===== ШАБЛОНЫ СООБЩЕНИЙ =====
message_templates = {
    'welcome':
        "🛋️ <b>Добро пожаловать в конструктор диванов и магазин!</b>\n\n"
        "Здесь вы можете:\n"
        "• Сконструировать диван из модулей\n"
        "• Сохранить свои диваны\n"
        "• Загрузить диван по коду\n"
        "• Выбрать готовые продукты\n"
        "• Оформить заказ на конструкторские диваны и готовые продукты\n"
        "• Отслеживать статус заказов\n"
        "• Связаться с менеджером",
    'sofa_info':
        "🛋️ <b>Диван</b>\n"
        "📋 <b>Код:</b> <code>{code}</code>\n\n"
        "<b>Состав:</b>\n{modules}\n\n"
        "<b>Стоимость для разных материалов:</b>\n{prices}",
    'cart_item':
        "<b>Позиция #{idx}</b>\n🛋️ <b>Диван:</b> {sofa_code}\n"
        "🧵 <b>Материал:</b> {material_info}\n"
        "💰 <b>Цена:</b> {price} руб\n"
        "<b>Состав:</b>\n{modules}",
    'order_confirmation_header': "✅ <b>Подтвердите данные заказа:</b>",
    'order_field': {
        'fio': "👤 <b>ФИО:</b> {value}",
        'phone': "📱 <b>Телефон:</b> {value}",
        'address': "🏠 <b>Адрес:</b> {value}",
        'comment': "💬 <b>Комментарий:</b> {value}"
    },
    'order_summary': "💳 <b>Итого:</b> {total} руб",
    'material_info':
        "<b>Материал:</b> {type} {color}\nКод: {code}\n{link}",
    'material_selection_help':
        "<b>Тестовый режим</b>\n🧵 Выберите материал. Его можно выбрать через бот или ввести код материала.",
    'product_info':
        "<b>{name}</b>\nКод: {code}\nТип: {type}\n{link_info}\n{prices}",
    'product_cart_item':
        "<b>Позиция #{idx}</b>\nТовар: {name}\nКод: {code}\nМатериал: {material_info}\nЦена: {price} руб",
    'material_for_product':
        "🧵 <b>Выбор материала для товара:</b> {product_name}\nКод товара: {product_code}",
    'product_view_confirmation_text':
        "✅ <b>Продукт найден!</b>\n\nВы хотите добавить этот продукт в корзину или вернуться в главное меню?",
    'no_available_materials':
        "❌ <b>Нет доступных материалов для этого изделия.</b>\n\n"
        "Для данного дивана/продукта не найдено ни одного материала, у которого есть цена. "
        "Пожалуйста, выберите другой диван или продукт.",
    'material_not_available':
        "❌ <b>Материал недоступен для этого изделия.</b>\n\n"
        "Для выбранного материала нет цены у данного дивана/продукта. "
        "Пожалуйста, выберите другой материал.",
    'cart_item_details':
        "<b>Детали позиции #{idx}</b>\n\n"
        "<b>Тип:</b> {item_type}\n"
        "<b>Код:</b> {item_code}\n"
        "<b>Название:</b> {item_name}\n"
        "<b>Материал:</b> {material_type}: {material_color}\n"
        "<b>Код материала:</b> {material_code}\n"
        "<b>Цена:</b> {price} руб\n"
        "<b>Дата добавления:</b> {date_added}",
    'module_info':
        "<b>{name}</b>\nКод: {code}\nТип: {type}\n{link_info}\n{size_info}\n{prices}",
    'module_selection':
        "📋 <b>Выберите модуль для просмотра деталей:</b>\n\n"
        "В текущем диване есть следующие модули:\n{module_list}",
    'manager_contact':
        "👨‍💼 <b>Связь с менеджером</b>\n\n"
        "Вы можете связаться с нашим менеджером для консультации:\n"
        "<b>{name}</b>\n"
        "Ссылка для связи: {link}",
    'order_details_header':
        "📦 <b>Детали заказа:</b> <code>{order_code}</code>\n"
        "📅 <b>Дата создания:</b> {datetime}\n"
        "🔄 <b>Статус:</b> {status}\n\n"
        "<b>Позиция {current}/{total}:</b>",
    'order_item_sofa':
        "🛋️ <b>Диван</b> (код: <code>{item_code}</code>)\n"
        "🧵 <b>Материал:</b> {material_info}\n"
        "💰 <b>Цена:</b> {price} руб\n\n"
        "<b>Состав:</b>\n{modules}",
    'order_item_product':
        "📦 <b>Продукт:</b> {item_name}\n"
        "📋 <b>Код:</b> <code>{item_code}</code>\n"
        "🧵 <b>Материал:</b> {material_info}\n"
        "💰 <b>Цена:</b> {price} руб"
}


# ===== НОВЫЕ ФУНКЦИИ ДЛЯ ПРОСМОТРА ДЕТАЛЕЙ ЗАКАЗА =====
def show_order_details(chat_id, order_code, item_index=0):
    """Показывает детали заказа с навигацией по позициям"""
    orders = get_user_orders(chat_id)
    if order_code not in orders:
        bot.send_message(chat_id, "❌ Заказ не найден")
        return

    order = orders[order_code]
    items = order['items']

    if not items:
        bot.send_message(chat_id, "❌ В заказе нет позиций")
        return

    if item_index < 0 or item_index >= len(items):
        item_index = 0

    # Сохраняем данные для навигации
    user_data[chat_id]['viewing_order_code'] = order_code
    user_data[chat_id]['viewing_order_index'] = item_index
    user_data[chat_id]['viewing_order_items'] = items
    user_data[chat_id]['state'] = 'viewing_order_details'

    item = items[item_index]

    # Отправляем заголовок заказа
    header = message_templates['order_details_header'].format(
        order_code=order_code,
        datetime=order['datetime'],
        status=order['status'],
        current=item_index + 1,
        total=len(items)
    )

    # Определяем тип позиции
    if item['code'].startswith('C'):  # Сконструированный диван
        modules = parse_sofa_code(item['code'])
        if modules:
            # Показываем схему дивана
            send_sofa_schema(chat_id, item['code'])

            # Формируем информацию о модулях
            modules_text = ""
            for i, module in enumerate(modules, 1):
                modules_text += f"{i}. {module['name']} ({module['code']})\n"

            # Отправляем информацию о диване
            item_info = message_templates['order_item_sofa'].format(
                item_code=item['code'],
                material_info=item['material_info'],
                price=item['price'],
                modules=modules_text
            )
            bot.send_message(chat_id, header + "\n" + item_info, parse_mode='HTML')
        else:
            item_info = message_templates['order_item_sofa'].format(
                item_code=item['code'],
                material_info=item['material_info'],
                price=item['price'],
                modules="Не удалось загрузить состав"
            )
            bot.send_message(chat_id, header + "\n" + item_info, parse_mode='HTML')

    elif item['code'].startswith('S'):  # Готовый продукт
        product = find_product_by_code(item['code'])
        if product:
            # Показываем фотографии продукта
            send_media_group(chat_id, 'product', product['code'])

            item_info = message_templates['order_item_product'].format(
                item_name=product['name'],
                item_code=item['code'],
                material_info=item['material_info'],
                price=item['price']
            )
            bot.send_message(chat_id, header + "\n" + item_info, parse_mode='HTML')
        else:
            item_info = message_templates['order_item_product'].format(
                item_name="Неизвестный продукт",
                item_code=item['code'],
                material_info=item['material_info'],
                price=item['price']
            )
            bot.send_message(chat_id, header + "\n" + item_info, parse_mode='HTML')

    else:
        # Неизвестный тип позиции
        item_info = f"<b>Позиция:</b> {item['code']}\n<b>Материал:</b> {item['material_info']}\n<b>Цена:</b> {item['price']} руб"
        bot.send_message(chat_id, header + "\n" + item_info, parse_mode='HTML')

    # Показываем материал
    if item.get('material_code'):
        material = find_material_by_code(item['material_code'])
        if material:
            # Отправляем фотографии материала
            caption = message_templates['material_info'].format(
                type=material['type'],
                color=material['color'],
                code=material['code'],
                link=f"Ссылка: {material['link']}" if material['link'] else ""
            )
            send_material_photos(chat_id, material['code'], caption)
        else:
            bot.send_message(chat_id, f"🧵 <b>Материал:</b> {item['material_info']}", parse_mode='HTML')
    else:
        bot.send_message(chat_id, f"🧵 <b>Материал:</b> {item['material_info']}", parse_mode='HTML')

    # Добавляем кнопки навигации
    markup = types.InlineKeyboardMarkup()

    if item_index > 0:
        markup.add(types.InlineKeyboardButton("⬅️ Предыдущая позиция",
                                              callback_data=f"order_prev_{order_code}_{item_index}"))

    if item_index < len(items) - 1:
        markup.add(types.InlineKeyboardButton("Следующая позиция ➡️",
                                              callback_data=f"order_next_{order_code}_{item_index}"))

    markup.add(types.InlineKeyboardButton("📋 Назад к заказу",
                                          callback_data=f"order_back_{order_code}"))

    bot.send_message(chat_id, f"Позиция {item_index + 1} из {len(items)}", reply_markup=markup)


# ===== ОБРАБОТЧИКИ CALLBACK ДЛЯ ДЕТАЛЕЙ ЗАКАЗА =====
@bot.callback_query_handler(func=lambda call: call.data.startswith('order_details_'))
def order_details_callback(call):
    """Обработчик кнопки 'Подробнее' в заказе"""
    chat_id = call.message.chat.id
    order_code = call.data.split('_')[2]

    delete_previous_messages(chat_id, call.message.message_id)
    show_order_details(chat_id, order_code, 0)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('order_prev_'))
def order_prev_callback(call):
    """Обработчик кнопки 'Предыдущая позиция'"""
    chat_id = call.message.chat.id
    parts = call.data.split('_')
    order_code = parts[2]
    current_index = int(parts[3])

    delete_previous_messages(chat_id, call.message.message_id)
    show_order_details(chat_id, order_code, current_index - 1)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('order_next_'))
def order_next_callback(call):
    """Обработчик кнопки 'Следующая позиция'"""
    chat_id = call.message.chat.id
    parts = call.data.split('_')
    order_code = parts[2]
    current_index = int(parts[3])

    delete_previous_messages(chat_id, call.message.message_id)
    show_order_details(chat_id, order_code, current_index + 1)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('order_back_'))
def order_back_callback(call):
    """Обработчик кнопки 'Назад к заказу'"""
    chat_id = call.message.chat.id
    order_code = call.data.split('_')[2]

    # Возвращаемся к просмотру заказа
    user_data[chat_id]['state'] = 'viewing_my_orders'
    delete_previous_messages(chat_id, call.message.message_id)
    show_user_orders(chat_id)
    bot.answer_callback_query(call.id)


# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С МОДУЛЯМИ (ДОПОЛНЕННЫЕ) =====
def get_module_details(code):
    """Получает полную информацию о модуле из таблицы"""
    try:
        wb = openpyxl.load_workbook(modules_log)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            module_code = sheet.cell(row=row, column=2).value
            if module_code and str(module_code).strip() == code:
                return {
                    'code': str(module_code).strip(),
                    'type': sheet.cell(row=row, column=3).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'prices': {
                        'Велюто': sheet.cell(row=row, column=5).value,
                        'Ромео': sheet.cell(row=row, column=6).value,
                        'Букле': sheet.cell(row=row, column=7).value,
                        'Эко-кожа': sheet.cell(row=row, column=8).value
                    },
                    'link': sheet.cell(row=row, column=9).value,
                    'height': sheet.cell(row=row, column=10).value,
                    'depth': sheet.cell(row=row, column=11).value,
                    'length': sheet.cell(row=row, column=12).value
                }
        return None
    except Exception as e:
        print(f"Ошибка загрузки деталей модуля: {e}")
        return None


def get_module_photos(module_code):
    """Получает фотографии модуля"""
    # Для модулей фото хранятся в Photo/Products с префиксом S
    photo_code = f"S{module_code}"
    base_dir = PHOTO_DIR
    photos_dir = os.path.join(base_dir, photo_code)

    photos = []
    if os.path.exists(photos_dir) and os.path.isdir(photos_dir):
        for file in os.listdir(photos_dir):
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                photos.append(os.path.join(photos_dir, file))
        photos.sort(key=lambda x: int(os.path.splitext(os.path.basename(x))[0]))
    return photos


def show_module_details(chat_id, module_code, module_index=None, total_modules=None):
    """Показывает детальную информацию о модуле"""
    module_details = get_module_details(module_code)
    if not module_details:
        bot.send_message(chat_id, f"❌ Не удалось загрузить информацию о модуле {module_code}")
        return

    # Получаем фотографии модуля
    photos = get_module_photos(module_code)

    # Формируем информацию о размерах
    size_info = ""
    if module_details.get('height') or module_details.get('depth') or module_details.get('length'):
        size_info = "📏 <b>Размеры:</b>\n"
        if module_details.get('height'):
            size_info += f"  • Высота: {module_details['height']} см\n"
        if module_details.get('depth'):
            size_info += f"  • Глубина: {module_details['depth']} см\n"
        if module_details.get('length'):
            size_info += f"  • Длина: {module_details['length']} см\n"

    # Формируем информацию о ссылке
    link_info = ""
    if module_details.get('link'):
        link_info = f"🔗 <b>Ссылка:</b> {module_details['link']}\n"

    # Формируем информацию о ценах
    prices_text = ""
    for material_type in material_types:
        price_key = material_type_mapping.get(material_type)
        if price_key and price_key in module_details['prices']:
            price = module_details['prices'][price_key]
            if price and price != '' and price != 0:
                prices_text += f"  {material_type}: {price} руб\n"

    if prices_text:
        prices_text = "💰 <b>Цены:</b>\n" + prices_text

    # Формируем полное сообщение
    module_info = message_templates['module_info'].format(
        name=module_details['name'],
        code=module_details['code'],
        type=module_details['type'],
        link_info=link_info,
        size_info=size_info,
        prices=prices_text
    )

    # Если есть фотографии, отправляем их
    if photos:
        try:
            media_group = []
            for i, photo_path in enumerate(photos):
                if i == 0:
                    media_group.append(
                        types.InputMediaPhoto(media=open(photo_path, 'rb'), caption=module_info, parse_mode='HTML'))
                else:
                    media_group.append(types.InputMediaPhoto(media=open(photo_path, 'rb')))
            bot.send_media_group(chat_id, media_group)
        except Exception as e:
            print(f"Ошибка отправки фотографий модуля: {e}")
            bot.send_message(chat_id, module_info, parse_mode='HTML')
    else:
        bot.send_message(chat_id, "📷 Фотографии модуля отсутствуют\n\n" + module_info, parse_mode='HTML')

    # Добавляем навигацию, если указаны индекс и общее количество
    if module_index is not None and total_modules is not None:
        user_data[chat_id]['current_module_index'] = module_index
        user_data[chat_id]['viewing_module_codes'] = [m['code'] for m in user_data[chat_id].get('current_modules', [])]
        bot.send_message(chat_id, f"📋 Модуль {module_index + 1} из {total_modules}",
                         reply_markup=create_keyboard('module_details_view', chat_id))
    else:
        bot.send_message(chat_id, "Выберите действие:",
                         reply_markup=create_keyboard('module_details_view', chat_id))


# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====
def is_material_available_for_sofa(modules, material_type):
    """Проверяет, доступен ли материал для дивана"""
    price_key = material_type_mapping.get(material_type)
    if not price_key:
        return False

    for module in modules:
        price = module['prices'].get(price_key)
        if price is None or price == '' or price == 0:
            return False

    return True


def is_material_available_for_product(product, material_type):
    """Проверяет, доступен ли материал для продукта"""
    price_key = material_type_mapping.get(material_type)
    if not price_key:
        return False

    price = product['prices'].get(price_key)
    return price is not None and price != '' and price != 0


def get_available_materials_for_sofa(modules):
    """Возвращает список доступных материалов для дивана"""
    available_materials = []
    for material_type in material_types:
        if is_material_available_for_sofa(modules, material_type):
            available_materials.append(material_type)
    return available_materials


def get_available_materials_for_product(product):
    """Возвращает список доступных материалов для продукта"""
    available_materials = []
    for material_type in material_types:
        if is_material_available_for_product(product, material_type):
            available_materials.append(material_type)
    return available_materials


def get_materials_by_type_with_availability(material_type, modules=None, product=None):
    """Возвращает материалы определенного типа с проверкой доступности"""
    all_materials = get_materials_by_type(material_type)

    if modules:
        available_materials = []
        for material in all_materials:
            if is_material_available_for_sofa(modules, material['type']):
                available_materials.append(material)
        return available_materials
    elif product:
        available_materials = []
        for material in all_materials:
            if is_material_available_for_product(product, material['type']):
                available_materials.append(material)
        return available_materials

    return all_materials


# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С EXCEL =====
def load_products():
    """Загружает каталог продуктов"""
    try:
        wb = openpyxl.load_workbook(products_log)
        sheet = wb.active
        products = []
        unique_types = set()

        for row in range(2, sheet.max_row + 1):
            product = {
                'code': sheet.cell(row=row, column=2).value,
                'type': sheet.cell(row=row, column=3).value,
                'name': sheet.cell(row=row, column=4).value,
                'link': sheet.cell(row=row, column=5).value,
                'prices': {
                    'Велюто': sheet.cell(row=row, column=6).value,
                    'Ромео': sheet.cell(row=row, column=7).value,
                    'Букле': sheet.cell(row=row, column=8).value,
                    'Эко-кожа': sheet.cell(row=row, column=9).value
                }
            }
            if product['code']:
                products.append(product)
                unique_types.add(product['type'])

        return products, list(unique_types)
    except Exception as e:
        print(f"Ошибка загрузки продуктов: {e}")
        return [], []


def find_product_by_code(code):
    """Ищет продукт по коду"""
    try:
        wb = openpyxl.load_workbook(products_log)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=2).value) == code:
                return {
                    'code': code,
                    'type': sheet.cell(row=row, column=3).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'link': sheet.cell(row=row, column=5).value,
                    'prices': {
                        'Велюто': sheet.cell(row=row, column=6).value,
                        'Ромео': sheet.cell(row=row, column=7).value,
                        'Букле': sheet.cell(row=row, column=8).value,
                        'Эко-кожа': sheet.cell(row=row, column=9).value
                    }
                }
        return None
    except Exception as e:
        print(f"Ошибка поиска продукта: {e}")
        return None


def load_modules():
    """Загружает каталог модулей"""
    try:
        wb = openpyxl.load_workbook(modules_log)
        sheet = wb.active
        modules = []

        for row in range(2, sheet.max_row + 1):
            module_code = sheet.cell(row=row, column=2).value
            if module_code:
                modules.append({
                    'code': str(module_code).strip(),
                    'type': sheet.cell(row=row, column=3).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'prices': {
                        'Велюто': sheet.cell(row=row, column=5).value,
                        'Ромео': sheet.cell(row=row, column=6).value,
                        'Букле': sheet.cell(row=row, column=7).value,
                        'Эко-кожа': sheet.cell(row=row, column=8).value
                    }
                })
        return modules
    except Exception as e:
        print(f"Ошибка загрузки модулей: {e}")
        return []


def find_module_by_code(code):
    """Ищет модуль по коду"""
    modules = load_modules()
    for module in modules:
        if module['code'] == code:
            return module
    return None


def get_available_sofa_types():
    """Возвращает список доступных типов диванов"""
    available_types = set()
    if os.path.exists(PHOTO_CONSTRUCTION_DIR):
        for folder in os.listdir(PHOTO_CONSTRUCTION_DIR):
            if folder.startswith('C') and len(folder) >= 3:
                sofa_type = folder[1]
                if sofa_type in module_type_names:
                    available_types.add(module_type_names[sofa_type])
    return sorted(available_types)


def get_base_modules(sofa_type):
    """Возвращает базовые модули для типа дивана"""
    type_letter = None
    for letter, name in module_type_names.items():
        if name == sofa_type:
            type_letter = letter
            break

    if not type_letter:
        return []

    modules = load_modules()
    base_modules = []

    if os.path.exists(PHOTO_CONSTRUCTION_DIR):
        for folder in os.listdir(PHOTO_CONSTRUCTION_DIR):
            if folder.startswith(f'C{type_letter}') and len(folder) == 4:
                module_code = folder[2:]
                module = find_module_by_code(f"{type_letter}{module_code}")
                if module:
                    base_modules.append(module)

    return base_modules


def get_available_extensions(current_sofa_code):
    """Возвращает доступные модули для расширения"""
    if len(current_sofa_code) < 4:
        return []

    sofa_type = current_sofa_code[1]
    modules = load_modules()
    available_modules = []

    if os.path.exists(PHOTO_CONSTRUCTION_DIR):
        for folder in os.listdir(PHOTO_CONSTRUCTION_DIR):
            if (folder.startswith(current_sofa_code) and
                    len(folder) == len(current_sofa_code) + 2 and
                    folder[len(current_sofa_code):].isalnum()):
                extension_code = folder[len(current_sofa_code):]
                module = find_module_by_code(f"{sofa_type}{extension_code}")
                if module:
                    available_modules.append(module)

    return available_modules


def calculate_sofa_prices(modules):
    """Рассчитывает общую стоимость дивана"""
    prices = {}
    for material_type in material_types:
        price_key = material_type_mapping.get(material_type, 'Велюто')
        total = 0
        for module in modules:
            price = module['prices'].get(price_key, 0)
            if price:
                total += price
        prices[material_type] = total
    return prices


def generate_sofa_code(modules):
    """Генерирует код дивана из модулей"""
    if not modules:
        return ""

    sofa_type = modules[0]['code'][0]
    code = f"C{sofa_type}"
    for module in modules:
        code += module['code'][1:]
    return code


def parse_sofa_code(sofa_code):
    """Разбирает код дивана на модули"""
    if not sofa_code.startswith('C') or len(sofa_code) < 4:
        return None

    modules = []
    sofa_type = sofa_code[1]
    module_codes = []

    for i in range(2, len(sofa_code), 2):
        if i + 2 <= len(sofa_code):
            module_code = f"{sofa_type}{sofa_code[i:i + 2]}"
            module_codes.append(module_code)

    for code in module_codes:
        module = find_module_by_code(code)
        if module:
            modules.append(module)
        else:
            return None

    return modules


def save_sofa(sofa_data):
    """Сохраняет диван в Excel"""
    try:
        if not os.path.exists(sofas_log):
            wb = openpyxl.Workbook()
            sheet = wb.active
            headers = [
                "Код дивана", "Модули", "Стоимость Букле", "Стоимость Эко-кожа",
                "Стоимость Велюр", "Стоимость Рогожка", "ID чата", "Дата создания"
            ]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            wb.save(sofas_log)

        wb = openpyxl.load_workbook(sofas_log)
        sheet = wb.active
        row_num = sheet.max_row + 1

        sheet.cell(row=row_num, column=1, value=sofa_data['code'])
        sheet.cell(row=row_num, column=2, value=', '.join([m['code'] for m in sofa_data['modules']]))
        sheet.cell(row=row_num, column=3, value=sofa_data['prices']['Букле'])
        sheet.cell(row=row_num, column=4, value=sofa_data['prices']['Эко-кожа'])
        sheet.cell(row=row_num, column=5, value=sofa_data['prices']['Велюр'])
        sheet.cell(row=row_num, column=6, value=sofa_data['prices']['Рогожка'])
        sheet.cell(row=row_num, column=7, value=sofa_data['chat_id'])
        sheet.cell(row=row_num, column=8, value=datetime.now().strftime("%d.%m.%Y %H:%M:%S"))

        wb.save(sofas_log)
        return True
    except Exception as e:
        print(f"Ошибка сохранения дивана: {e}")
        return False


def get_user_sofas(chat_id):
    """Получает все диваны пользователя"""
    try:
        if not os.path.exists(sofas_log):
            return {}

        wb = openpyxl.load_workbook(sofas_log)
        sheet = wb.active
        sofas = {}

        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=7).value) == str(chat_id):
                sofa_code = sheet.cell(row=row, column=1).value
                sofas[sofa_code] = {
                    'code': sofa_code,
                    'modules': parse_sofa_code(sofa_code) or [],
                    'prices': {
                        'Букле': sheet.cell(row=row, column=3).value,
                        'Эко-кожа': sheet.cell(row=row, column=4).value,
                        'Велюр': sheet.cell(row=row, column=5).value,
                        'Рогожка': sheet.cell(row=row, column=6).value
                    },
                    'datetime': sheet.cell(row=row, column=8).value
                }
        return sofas
    except Exception as e:
        print(f"Ошибка получения диванов: {e}")
        return {}


def load_materials():
    """Загружает каталог материалов"""
    try:
        wb = openpyxl.load_workbook(materials_log)
        sheet = wb.active
        materials = []
        for row in range(2, sheet.max_row + 1):
            code = sheet.cell(row=row, column=2).value
            if code is None:
                continue
            materials.append({
                'code': str(code).strip(),
                'type': sheet.cell(row=row, column=3).value,
                'color': sheet.cell(row=row, column=4).value,
                'link': sheet.cell(row=row, column=5).value
            })
        return materials
    except Exception as e:
        print(f"Ошибка загрузки материалов: {e}")
        return []


def get_material_types():
    """Возвращает список типов материалов"""
    materials = load_materials()
    return sorted(set(m['type'] for m in materials))


def find_material_by_code(code):
    """Ищет материал по коду"""
    materials = load_materials()
    for material in materials:
        if material['code'] == code:
            return material
    return None


def get_materials_by_type(material_type):
    """Возвращает материалы определенного типа"""
    materials = load_materials()
    return [m for m in materials if m['type'] == material_type]


def save_order(order_data):
    """Сохраняет заказ в базу данных"""
    try:
        if not os.path.exists(orders_log):
            wb = openpyxl.Workbook()
            sheet = wb.active
            headers = [
                "Порядковый номер позиции", "Номер задачи", "Код заказа", "ФИО Заказчика",
                "Телефон", "Адрес", "Код чата", "Телеграмм", "Код позиции",
                "Материал", "Цвет материала", "Стоимость", "Дата и время", "Статус заказа", "Комментарий",
                "", "Код Швеи", "ФИО Швеи", "Готовность пошива", "Код Столяра", "ФИО Столяра", "Готовность Каркаса",
                "Код Обтяжчика", "ФИО Обтяжчика", "Готовность продукта", "Код упаковщика", "ФИО Упаковщика",
                "Готовность позиции"
            ]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            wb.save(orders_log)

        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        row_num = sheet.max_row + 1
        order_datetime = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        for idx, item in enumerate(order_data['cart'], 1):
            sheet.cell(row=row_num, column=1, value=idx)
            sheet.cell(row=row_num, column=2, value=order_data['task_number'])
            sheet.cell(row=row_num, column=3, value=order_data['order_code'])
            sheet.cell(row=row_num, column=4, value=order_data['fio'])
            sheet.cell(row=row_num, column=5, value=order_data['phone'])
            sheet.cell(row=row_num, column=6, value=order_data['address'])
            sheet.cell(row=row_num, column=7, value=order_data['chat_id'])
            sheet.cell(row=row_num, column=8, value=order_data['telegram_contact'])
            sheet.cell(row=row_num, column=9, value=item.get('item_code', ''))
            sheet.cell(row=row_num, column=10, value=item.get('material_code', ''))
            sheet.cell(row=row_num, column=11, value=f"{item['material_type']}: {item['material_color']}")
            sheet.cell(row=row_num, column=12, value=item['price'])
            sheet.cell(row=row_num, column=13, value=order_datetime)
            sheet.cell(row=row_num, column=14, value="Ожидает подтверждения")
            sheet.cell(row=row_num, column=15, value=order_data['comment'])

            for col in range(17, 29):
                sheet.cell(row=row_num, column=col, value="нет")

            row_num += 1

        wb.save(orders_log)
        return True
    except Exception as e:
        print(f"Ошибка сохранения заказа: {e}")
        traceback.print_exc()
        return False


def get_user_orders(chat_id):
    """Получает все заказы пользователя"""
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        orders = {}
        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=7).value) == str(chat_id):
                order_code = sheet.cell(row=row, column=3).value
                if order_code not in orders:
                    orders[order_code] = {
                        'order_code': order_code,
                        'datetime': sheet.cell(row=row, column=13).value,
                        'status': sheet.cell(row=row, column=14).value,
                        'fio': sheet.cell(row=row, column=4).value,
                        'phone': sheet.cell(row=row, column=5).value,
                        'address': sheet.cell(row=row, column=6).value,
                        'comment': sheet.cell(row=row, column=15).value,
                        'items': []
                    }
                orders[order_code]['items'].append({
                    'code': sheet.cell(row=row, column=9).value,
                    'material_code': sheet.cell(row=row, column=10).value,
                    'material_info': sheet.cell(row=row, column=11).value,
                    'price': sheet.cell(row=row, column=12).value
                })
        return orders
    except Exception as e:
        print(f"Ошибка получения заказов: {e}")
        return {}


# ===== ФУНКЦИИ ОТОБРАЖЕНИЯ =====
def get_photos(media_type, code):
    """Получает фотографии продуктов/материалов"""
    if media_type == 'product':
        base_dir = PHOTO_DIR
    elif media_type == 'material':
        base_dir = PHOTO_MATERIALS_DIR
    else:
        return []

    media_dir = os.path.join(base_dir, code)
    photos = []
    if os.path.exists(media_dir) and os.path.isdir(media_dir):
        for file in os.listdir(media_dir):
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                photos.append(os.path.join(media_dir, file))
        photos.sort(key=lambda x: int(os.path.splitext(os.path.basename(x))[0]))
    return photos


def send_media_group(chat_id, media_type, code, caption=None):
    """Отправляет медиа-группу"""
    photos = get_photos(media_type, code)
    if not photos:
        return False

    try:
        media_group = []
        for i, photo_path in enumerate(photos):
            if i == 0 and caption:
                media_group.append(
                    types.InputMediaPhoto(media=open(photo_path, 'rb'), caption=caption, parse_mode='HTML'))
            else:
                media_group.append(types.InputMediaPhoto(media=open(photo_path, 'rb')))
        bot.send_media_group(chat_id, media_group)
        return True
    except Exception as e:
        print(f"Ошибка отправки фотографий ({media_type}): {e}")
        return False


def get_sofa_schema(sofa_code):
    """Возвращает путь к файлу схемы дивана"""
    schema_path = os.path.join(PHOTO_CONSTRUCTION_DIR, sofa_code, 'cells.png')
    return schema_path if os.path.exists(schema_path) else None


def send_sofa_schema(chat_id, sofa_code):
    """Отправляет схему дивана"""
    schema_path = get_sofa_schema(sofa_code)
    if schema_path:
        try:
            with open(schema_path, 'rb') as photo:
                bot.send_photo(chat_id, photo)
            return True
        except Exception as e:
            print(f"Ошибка отправки схемы: {e}")
            return False
    return False


def send_material_photos(chat_id, material_code, caption=None):
    """Отправляет фотографии материала"""
    return send_media_group(chat_id, 'material', material_code, caption)


def go_to_site(message):
    """Отправляет ссылку на сайт"""
    chat_id = message.chat.id
    bot.send_message(chat_id, f"🌐 Перейдите на наш сайт: {site_shop}", parse_mode='HTML')


def show_sofa_card(chat_id, modules, sofa_code, show_schema=True):
    """Показывает карточку дивана"""
    if not modules:
        bot.send_message(chat_id, "❌ Диван не содержит модулей")
        return

    prices = calculate_sofa_prices(modules)

    if show_schema:
        send_sofa_schema(chat_id, sofa_code)

    modules_text = ""
    for i, module in enumerate(modules, 1):
        modules_text += f"{i}. {module['name']} ({module['code']})\n"

    prices_text = ""
    for material_type in material_types:
        if is_material_available_for_sofa(modules, material_type):
            prices_text += f"• {material_type}: {prices[material_type]} руб\n"
        else:
            prices_text += f"• {material_type}: <b>недоступно</b>\n"

    sofa_info = message_templates['sofa_info'].format(
        code=sofa_code,
        modules=modules_text,
        prices=prices_text
    )

    bot.send_message(chat_id, sofa_info, parse_mode='HTML')


def show_user_saved_sofas(chat_id):
    """Показывает сохраненные диваны пользователя"""
    sofas = get_user_sofas(chat_id)

    if not sofas:
        bot.send_message(chat_id, "📭 У вас нет сохраненных диванов.")
        return

    for sofa_code, sofa_data in sofas.items():
        markup = types.InlineKeyboardMarkup()
        btn_edit = types.InlineKeyboardButton('Редактировать', callback_data=f'edit_sofa_{sofa_code}')
        btn_delete = types.InlineKeyboardButton('Удалить', callback_data=f'delete_sofa_{sofa_code}')
        btn_add_to_cart = types.InlineKeyboardButton('Добавить в заказ', callback_data=f'add_to_cart_{sofa_code}')
        btn_view = types.InlineKeyboardButton('Просмотр', callback_data=f'view_sofa_{sofa_code}')
        markup.row(btn_edit, btn_delete)
        markup.row(btn_add_to_cart, btn_view)

        bot.send_message(chat_id, f"💾 Сохраненный диван: {sofa_code}", reply_markup=markup)


def show_material(chat_id, material):
    """Показывает информацию о материале"""
    caption = message_templates['material_info'].format(
        type=material['type'],
        color=material['color'],
        code=material['code'],
        link=f"Ссылка: {material['link']}" if material['link'] else ""
    )

    bot.send_message(chat_id, caption, parse_mode='HTML')

    photos_sent = send_material_photos(chat_id, material['code'])
    if not photos_sent:
        bot.send_message(chat_id, "📷 Фотографии отсутствуют")

    bot.send_message(chat_id, "Выберите действие:", reply_markup=create_keyboard('material_navigation', chat_id))


def show_product(chat_id, product, show_navigation=True):
    """Показывает информацию о продукте"""
    photos_sent = send_media_group(chat_id, 'product', product['code'])
    link_info = f"Ссылка: {product['link']}" if product['link'] else "Ссылка: отсутствует"

    prices_text = "Цены:\n"
    for material_type in material_types:
        if is_material_available_for_product(product, material_type):
            price_key = material_type_mapping.get(material_type)
            price = product['prices'].get(price_key)
            prices_text += f"  {material_type}: {price} руб\n"
        else:
            prices_text += f"  {material_type}: <b>недоступно</b>\n"

    response = message_templates['product_info'].format(
        name=product['name'],
        code=product['code'],
        type=product['type'],
        link_info=link_info,
        prices=prices_text
    )

    if not photos_sent:
        response = "📷 Фотографии отсутствуют\n\n" + response

    if show_navigation:
        bot.send_message(chat_id, response, parse_mode='HTML', reply_markup=create_keyboard('product_navigation'))
    else:
        bot.send_message(chat_id, response, parse_mode='HTML')


def view_cart(chat_id):
    """Показывает содержимое корзины"""
    cart_items = get_user_cart(chat_id)

    if not cart_items:
        bot.send_message(chat_id, "🛒 Ваша корзина пуста!", reply_markup=create_keyboard('cart_management', chat_id))
        return

    user_data[chat_id]['state'] = 'viewing_cart'

    for idx, item in enumerate(cart_items, 1):
        markup = types.InlineKeyboardMarkup()
        btn_view = types.InlineKeyboardButton('👁️ Просмотр', callback_data=f'view_cart_item_{item["db_id"]}')

        if item['item_type'] == 'sofa':
            btn_edit = types.InlineKeyboardButton('Изменить материал', callback_data=f'edit_cart_item_{item["db_id"]}')
            btn_delete = types.InlineKeyboardButton('Удалить', callback_data=f'delete_cart_item_{item["db_id"]}')
            markup.row(btn_edit, btn_delete, btn_view)

            modules = parse_sofa_code(item['item_code'])
            if modules:
                modules_text = ""
                for i, module in enumerate(modules, 1):
                    modules_text += f"{i}. {module['name']} ({module['code']})\n"
            else:
                modules_text = "Не удалось загрузить состав\n"

            cart_item_text = message_templates['cart_item'].format(
                idx=idx,
                sofa_code=item['item_code'],
                material_info=f"{item['material_type']}: {item['material_color']}",
                price=item['price'],
                modules=modules_text
            )
        else:
            btn_edit = types.InlineKeyboardButton('Изменить материал', callback_data=f'edit_cart_item_{item["db_id"]}')
            btn_delete = types.InlineKeyboardButton('Удалить', callback_data=f'delete_cart_item_{item["db_id"]}')
            markup.row(btn_edit, btn_delete, btn_view)

            material_info = f"{item['material_type']}: {item['material_color']}"
            cart_item_text = message_templates['product_cart_item'].format(
                idx=idx,
                name=item['item_name'],
                code=item['item_code'],
                material_info=material_info,
                price=item['price']
            )

        bot.send_message(
            chat_id,
            cart_item_text,
            reply_markup=markup,
            parse_mode='HTML'
        )

    bot.send_message(
        chat_id,
        "Используйте кнопки под каждой позицией для управления",
        reply_markup=create_keyboard('cart_management', chat_id),
        parse_mode='HTML'
    )


def show_cart_item_details(chat_id, db_id):
    """Показывает детали позиции из корзины"""
    cart_items = get_user_cart(chat_id)
    item = None

    for cart_item in cart_items:
        if cart_item['db_id'] == db_id:
            item = cart_item
            break

    if not item:
        bot.send_message(chat_id, "❌ Позиция не найдена")
        return

    item_details = message_templates['cart_item_details'].format(
        idx=db_id,
        item_type='Диван' if item['item_type'] == 'sofa' else 'Продукт',
        item_code=item['item_code'],
        item_name=item['item_name'],
        material_type=item['material_type'],
        material_color=item['material_color'],
        material_code=item['material_code'],
        price=item['price'],
        date_added=item['date_added']
    )

    if item['item_type'] == 'sofa':
        send_sofa_schema(chat_id, item['item_code'])
    else:
        send_media_group(chat_id, 'product', item['item_code'])

    bot.send_message(
        chat_id,
        item_details,
        parse_mode='HTML',
        reply_markup=create_keyboard('view_cart_item', chat_id)
    )


def confirm_order(chat_id):
    """Показывает сводную информацию о заказе"""
    order_data = user_data[chat_id]['order_data']
    cart_items = get_user_cart(chat_id)
    total = sum(item['price'] for item in cart_items)

    bot.send_message(chat_id, message_templates['order_confirmation_header'], parse_mode='HTML')

    for field in ['fio', 'phone', 'address', 'comment']:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton('Изменить', callback_data=f'editorder_{field}'))
        value = order_data[field] if field != 'comment' else (order_data[field] or "Отсутствует")
        bot.send_message(
            chat_id,
            message_templates['order_field'][field].format(value=value),
            reply_markup=markup,
            parse_mode='HTML'
        )

    cart_text = "\n".join([
        f"• {item['item_name']} ({item['material_type']}: {item['material_color']}) - {item['price']} руб"
        for item in cart_items
    ])
    bot.send_message(chat_id, f"🛒 <b>Состав заказа:</b>\n{cart_text}", parse_mode='HTML')
    bot.send_message(chat_id, message_templates['order_summary'].format(total=total), parse_mode='HTML')

    bot.send_message(
        chat_id,
        "Подтвердите или отмените заказ:",
        reply_markup=create_keyboard('order_confirmation', chat_id),
        parse_mode='HTML'
    )


def show_user_orders(chat_id):
    """Показывает пользователю все его заказы"""
    orders = get_user_orders(chat_id)
    if not orders:
        bot.send_message(chat_id, "📭 У вас пока нет заказов.")
        bot.send_message(chat_id, "Выберите действие:", reply_markup=create_keyboard('my_orders', chat_id))
        return

    for order_code, order_data in orders.items():
        order_message = (
            f"📦 <b>Заказ:</b> <code>{order_code}</code>\n"
            f"📅 <b>Дата создания:</b> {order_data['datetime']}\n"
            f"🔄 <b>Статус:</b> {order_data['status']}\n\n"
            f"<b>Состав заказа:</b>\n"
        )

        total_price = 0
        for idx, item in enumerate(order_data['items'], 1):
            order_message += (
                f"{idx}. <b>Код позиции:</b> {item['code']}\n"
                f"   <b>Материал:</b> {item['material_info']}\n"
                f"   <b>Цена:</b> {item['price']} руб\n\n"
            )
            total_price += item['price']

        order_message += f"💰 <b>Общая стоимость:</b> {total_price} руб"

        # Добавляем inline кнопку "Подробнее"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔍 Подробнее", callback_data=f"order_details_{order_code}"))

        bot.send_message(chat_id, order_message, parse_mode='HTML', reply_markup=markup)

        contact_info = (
            f"👤 <b>ФИО:</b> {order_data['fio']}\n"
            f"📱 <b>Телефон:</b> {order_data['phone']}\n"
            f"🏠 <b>Адрес:</b> {order_data['address']}\n"
            f"💬 <b>Комментарий:</b> {order_data['comment'] or 'Отсутствует'}"
        )
        bot.send_message(chat_id, contact_info, parse_mode='HTML')

    bot.send_message(chat_id, "Выберите действие:", reply_markup=create_keyboard('my_orders', chat_id))


def delete_previous_messages(chat_id, message_id):
    """Очистка чата"""
    if chat_id not in user_data:
        user_data[chat_id] = {}

    temp_msg = bot.send_message(chat_id, "🔄", parse_mode='HTML')
    last_anchor = user_data[chat_id].get('last_anchor_message_id', message_id)

    try:
        for msg_id in range(last_anchor + 1, temp_msg.message_id):
            try:
                bot.delete_message(chat_id, msg_id)
            except Exception:
                continue
    except Exception as e:
        print(f"Ошибка при удалении сообщений: {e}")

    user_data[chat_id]['last_anchor_message_id'] = temp_msg.message_id

    try:
        bot.delete_message(chat_id, temp_msg.message_id)
    except Exception as e:
        print(f"Ошибка при удалении временного сообщения: {e}")


def generate_task_number():
    """Генерирует 7-значный номер задачи"""
    return random.randint(1000000, 9999999)


def generate_order_code():
    """Генерирует 10-символьный код заказа"""
    chars = string.ascii_uppercase.replace('O', '') + string.digits.replace('0', '')
    return ''.join(random.choice(chars) for _ in range(10))


# ===== НОВЫЕ ФУНКЦИИ ДЛЯ РАБОТЫ С МОДУЛЯМИ =====
def handle_module_details_selection(message):
    """Обработчик выбора модуля для просмотра деталей"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        user_data[chat_id]['state'] = 'constructing_sofa'
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)

        available_extensions = get_available_extensions(sofa_code)
        exclude_buttons = []
        if not available_extensions:
            exclude_buttons.append('Добавить модуль')

        bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                         reply_markup=create_keyboard('construction', chat_id, exclude_buttons=exclude_buttons))
        return

    # Пытаемся найти модуль по названию или номеру
    current_modules = user_data[chat_id]['current_modules']
    selected_module = None
    module_index = None

    # Проверяем, является ли ввод числом (индексом модуля)
    try:
        module_num = int(text)
        if 1 <= module_num <= len(current_modules):
            selected_module = current_modules[module_num - 1]
            module_index = module_num - 1
    except ValueError:
        # Ищем модуль по названию
        for idx, module in enumerate(current_modules):
            if module['name'] == text:
                selected_module = module
                module_index = idx
                break

    if selected_module:
        user_data[chat_id]['state'] = 'viewing_module_details'
        user_data[chat_id]['current_module_index'] = module_index
        delete_previous_messages(chat_id, message.message_id)
        show_module_details(chat_id, selected_module['code'], module_index, len(current_modules))
    else:
        bot.send_message(chat_id, "❌ Модуль не найден. Пожалуйста, выберите модуль из списка.")


def handle_module_details_view(message):
    """Обработчик просмотра деталей модуля"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад к списку модулей':
        user_data[chat_id]['state'] = 'selecting_module_for_details'
        delete_previous_messages(chat_id, message.message_id)
        show_module_list_for_selection(chat_id)
        return

    current_modules = user_data[chat_id]['current_modules']
    current_index = user_data[chat_id].get('current_module_index', 0)

    if text == 'Предыдущий модуль':
        new_index = (current_index - 1) % len(current_modules)
        user_data[chat_id]['current_module_index'] = new_index
        selected_module = current_modules[new_index]
        delete_previous_messages(chat_id, message.message_id)
        show_module_details(chat_id, selected_module['code'], new_index, len(current_modules))

    elif text == 'Следующий модуль':
        new_index = (current_index + 1) % len(current_modules)
        user_data[chat_id]['current_module_index'] = new_index
        selected_module = current_modules[new_index]
        delete_previous_messages(chat_id, message.message_id)
        show_module_details(chat_id, selected_module['code'], new_index, len(current_modules))


def show_module_list_for_selection(chat_id):
    """Показывает список модулей для выбора"""
    current_modules = user_data[chat_id]['current_modules']

    if not current_modules:
        bot.send_message(chat_id, "❌ В диване нет модулей для просмотра.")
        return

    module_list = ""
    for i, module in enumerate(current_modules, 1):
        module_list += f"{i}. {module['name']} ({module['code']})\n"

    message_text = message_templates['module_selection'].format(module_list=module_list)

    # Создаем клавиатуру с названиями модулей
    module_names = [module['name'] for module in current_modules]

    bot.send_message(chat_id, message_text, parse_mode='HTML',
                     reply_markup=create_keyboard('module_details_selection', chat_id,
                                                  available_module_types=module_names))


# ===== ФУНКЦИЯ ДЛЯ ОТПРАВКИ КОНТАКТА МЕНЕДЖЕРА =====
def send_manager_contact(chat_id):
    """Отправляет контакт случайного менеджера"""
    if not managers:
        bot.send_message(chat_id, "❌ На данный момент нет доступных менеджеров. Попробуйте позже.")
        return

    manager = random.choice(managers)
    message = message_templates['manager_contact'].format(
        name=manager['name'],
        link=manager['link']
    )
    bot.send_message(chat_id, message, parse_mode='HTML')


# ===== ОБРАБОТЧИКИ СОСТОЯНИЙ =====
@bot.message_handler(commands=['start'])
def start(message):
    """Обработчик команды /start"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'state': 'main_menu',
        'current_sofa_type': None,
        'current_modules': [],
        'editing_sofa_code': None,
        'cart': [],
        'order_data': {},
        'current_product_type': None,
        'current_product_index': 0,
        'current_material_type': None,
        'current_material_index': 0,
        'last_anchor_message_id': message.message_id,
        'current_module_index': 0,
        'viewing_order_code': None,
        'viewing_order_index': 0,
        'viewing_order_items': []
    }
    delete_previous_messages(chat_id, message.message_id)
    bot.send_message(
        chat_id,
        message_templates['welcome'],
        reply_markup=create_keyboard('main', chat_id),
        parse_mode='HTML'
    )


@bot.message_handler(func=lambda message: message.text == 'Перезапустить бота')
def restart_bot(message):
    """Обработчик перезапуска бота"""
    chat_id = message.chat.id
    if chat_id in user_data:
        del user_data[chat_id]
    delete_previous_messages(chat_id, message.message_id)
    start(message)


@bot.message_handler(func=lambda message: message.text == 'Перейти на сайт')
def handle_go_to_site(message):
    """Обработчик кнопки перехода на сайт"""
    go_to_site(message)


# ===== УНИВЕРСАЛЬНЫЙ ОБРАБОТЧИК =====
@bot.message_handler(func=lambda message: True)
def universal_state_handler(message):
    """Универсальный обработчик сообщений"""
    chat_id = message.chat.id
    text = message.text

    # Инициализация данных пользователя, если их нет
    if chat_id not in user_data:
        user_data[chat_id] = {'state': 'main_menu'}

    current_state = user_data[chat_id].get('state', 'main_menu')

    # Обработка навигации по материалам
    if current_state == 'viewing_materials' and text in ['Предыдущий', 'Следующий', 'Назад', 'Добавить в корзину']:
        handle_material_navigation(message)
        return

    if current_state == 'viewing_materials_for_product' and text in ['Предыдущий', 'Следующий', 'Назад',
                                                                     'Добавить в корзину']:
        handle_material_navigation_for_product(message)
        return

    if current_state == 'viewing_products' and text in ['Предыдущий', 'Следующий', 'Назад', 'Перейти на сайт',
                                                        'Добавить в корзину']:
        handle_product_navigation(message)
        return

    # Обработка навигации по модулям
    if current_state == 'viewing_module_details' and text in ['Предыдущий модуль', 'Следующий модуль',
                                                              'Назад к списку модулей']:
        handle_module_details_view(message)
        return

    # Обработка навигации по деталям заказа
    if current_state == 'viewing_order_details' and text in ['Предыдущая позиция', 'Следующая позиция',
                                                             'Назад к заказу']:
        handle_order_details_navigation(message)
        return

    # Обработка основных состояний
    if current_state == 'main_menu':
        handle_main_menu(message)
    elif current_state == 'selecting_sofa_type':
        handle_sofa_type_selection(message)
    elif current_state == 'selecting_base_module':
        handle_base_module_selection(message)
    elif current_state == 'constructing_sofa':
        handle_sofa_construction(message)
    elif current_state == 'selecting_extension_module':
        handle_extension_module_selection(message)
    elif current_state == 'entering_sofa_code':
        handle_sofa_code_input(message)
    elif current_state == 'product_view_confirmation':
        handle_product_view_confirmation(message)
    elif current_state == 'editing_sofa':
        handle_sofa_editing(message)
    elif current_state == 'selecting_material_method':
        handle_sofa_material_method(message)
    elif current_state == 'selecting_product_material_method':
        handle_product_material_method(message)
    elif current_state == 'entering_material_code_for_product':
        handle_material_code_input_for_product(message)
    elif current_state == 'confirming_material_for_product':
        handle_material_confirmation_for_product(message)
    elif current_state == 'selecting_material_type':
        handle_material_type_selection(message)
    elif current_state == 'selecting_material_type_for_product':
        handle_material_type_selection_for_product(message)
    elif current_state == 'entering_material_code':
        handle_material_code_input(message)
    elif current_state == 'confirming_material':
        handle_material_confirmation(message)
    elif current_state == 'order_menu':
        handle_order_menu(message)
    elif current_state == 'entering_sofa_code_for_order':
        handle_sofa_code_for_order_input(message)
    elif current_state == 'collecting_customer_data':
        handle_collecting_customer_data(message)
    elif current_state == 'confirming_order':
        handle_order_confirmation(message)
    elif current_state == 'viewing_my_orders':
        handle_viewing_my_orders(message)
    elif current_state == 'selecting_product_type':
        handle_product_type_selection(message)
    elif current_state == 'entering_product_code':
        handle_product_code_input(message)
    elif current_state == 'viewing_cart':
        handle_viewing_cart(message)
    elif current_state == 'view_cart_item':
        handle_view_cart_item(message)
    elif current_state == 'viewing_saved_sofa':
        handle_viewing_saved_sofa(message)
    elif current_state == 'editing_order_field':
        handle_order_field_edit(message)
    elif current_state == 'viewing_saved_sofas':
        handle_saved_sofas(message)
    elif current_state == 'selecting_module_for_details':
        handle_module_details_selection(message)
    else:
        # Если состояние не распознано, показываем главное меню
        bot.send_message(chat_id, "ℹ️ Пожалуйста, используйте кнопки для взаимодействия с ботом.")
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:", reply_markup=create_keyboard('main', chat_id))


# ===== ОСНОВНЫЕ ОБРАБОТЧИКИ =====
def handle_main_menu(message):
    """Обработчик главного меню"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Сконструировать диван':
        user_data[chat_id]['state'] = 'selecting_sofa_type'
        available_types = get_available_sofa_types()
        if not available_types:
            bot.send_message(chat_id, "❌ Нет доступных типов диванов")
            return
        bot.send_message(chat_id, "📋 Выберите тип дивана:",
                         reply_markup=create_keyboard('sofa_type', chat_id))

    elif text == 'Мои диваны':
        user_data[chat_id]['state'] = 'viewing_saved_sofas'
        show_user_saved_sofas(chat_id)
        bot.send_message(chat_id, "Выберите действие:",
                         reply_markup=create_keyboard('saved_sofas', chat_id))

    elif text == 'Ввести код дивана':
        user_data[chat_id]['state'] = 'entering_sofa_code'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id))

    elif text == 'Оформить заказ':
        user_data[chat_id]['state'] = 'order_menu'
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')

    elif text == 'Мои заказы':
        user_data[chat_id]['state'] = 'viewing_my_orders'
        show_user_orders(chat_id)

    elif text == 'Связаться с менеджером':
        delete_previous_messages(chat_id, message.message_id)
        send_manager_contact(chat_id)
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))


def handle_saved_sofas(message):
    """Обработчик сохраненных диванов"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Сконструировать диван':
        user_data[chat_id]['state'] = 'selecting_sofa_type'
        bot.send_message(chat_id, "📋 Выберите тип дивана:",
                         reply_markup=create_keyboard('sofa_type', chat_id))

    elif text == 'Ввести код дивана':
        user_data[chat_id]['state'] = 'entering_sofa_code'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id))

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))


def handle_sofa_type_selection(message):
    """Обработчик выбора типа дивана"""
    chat_id = message.chat.id
    text = message.text
    available_types = get_available_sofa_types()

    if text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))
        return

    if text not in available_types:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите тип из списка.")
        return

    user_data[chat_id]['current_sofa_type'] = text
    user_data[chat_id]['state'] = 'selecting_base_module'

    base_modules = get_base_modules(text)
    if not base_modules:
        bot.send_message(chat_id, "❌ Нет доступных базовых модулей для этого типа")
        user_data[chat_id]['state'] = 'selecting_sofa_type'
        return

    bot.send_message(chat_id, "🔧 Выберите базовый модуль:",
                     reply_markup=create_keyboard('base_module', chat_id))


def handle_base_module_selection(message):
    """Обработчик выбора базового модуля"""
    chat_id = message.chat.id
    text = message.text
    sofa_type = user_data[chat_id]['current_sofa_type']

    if text == 'Назад':
        user_data[chat_id]['state'] = 'selecting_sofa_type'
        bot.send_message(chat_id, "📋 Выберите тип дивана:",
                         reply_markup=create_keyboard('sofa_type', chat_id))
        return

    base_modules = get_base_modules(sofa_type)
    selected_module = None

    for module in base_modules:
        if module['name'] == text:
            selected_module = module
            break

    if not selected_module:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите модуль из списка.")
        return

    user_data[chat_id]['current_modules'] = [selected_module]
    user_data[chat_id]['state'] = 'constructing_sofa'

    sofa_code = generate_sofa_code(user_data[chat_id]['current_modules'])
    delete_previous_messages(chat_id, message.message_id)
    show_sofa_card(chat_id, user_data[chat_id]['current_modules'], sofa_code)

    available_extensions = get_available_extensions(sofa_code)
    exclude_buttons = []
    if not available_extensions:
        exclude_buttons.append('Добавить модуль')

    bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                     reply_markup=create_keyboard('construction', chat_id, exclude_buttons=exclude_buttons))


def handle_sofa_construction(message):
    """Обработчик конструирования дивана"""
    chat_id = message.chat.id
    text = message.text
    current_modules = user_data[chat_id]['current_modules']
    current_sofa_code = generate_sofa_code(current_modules)

    available_extensions = get_available_extensions(current_sofa_code)
    exclude_buttons = []
    if not available_extensions:
        exclude_buttons.append('Добавить модуль')

    if text == 'Добавить модуль':
        if not available_extensions:
            bot.send_message(chat_id, "❌ Нет доступных модулей для расширения этого дивана")
            return

        user_data[chat_id]['available_extensions'] = available_extensions
        user_data[chat_id]['state'] = 'selecting_extension_module'

        bot.send_message(chat_id, "🔧 Выберите модуль для добавления:",
                         reply_markup=create_keyboard('extension_module', chat_id))

    elif text == 'Подробнее о модуле':
        if not current_modules:
            bot.send_message(chat_id, "❌ В диване нет модулей для просмотра.")
            return

        user_data[chat_id]['state'] = 'selecting_module_for_details'
        delete_previous_messages(chat_id, message.message_id)
        show_module_list_for_selection(chat_id)

    elif text == 'Удалить последний модуль':
        if len(current_modules) > 1:
            user_data[chat_id]['current_modules'].pop()
            new_sofa_code = generate_sofa_code(user_data[chat_id]['current_modules'])

            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id, "✅ Последний модуль удален")
            show_sofa_card(chat_id, user_data[chat_id]['current_modules'], new_sofa_code)

            new_available_extensions = get_available_extensions(new_sofa_code)
            new_exclude_buttons = []
            if not new_available_extensions:
                new_exclude_buttons.append('Добавить модуль')

            bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                             reply_markup=create_keyboard('construction', chat_id, exclude_buttons=new_exclude_buttons))
        else:
            user_data[chat_id]['state'] = 'selecting_base_module'
            bot.send_message(chat_id, "❌ Базовый модуль удален. Выберите новый базовый модуль:",
                             reply_markup=create_keyboard('base_module', chat_id))

    elif text == 'Завершить конструирование':
        sofa_code = generate_sofa_code(current_modules)
        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)

        bot.send_message(chat_id,
                         f"✅ Конструирование завершено!\n"
                         f"📋 Код вашего дивана: <code>{sofa_code}</code>\n"
                         f"Вы можете сохранить его или начать новый.",
                         parse_mode='HTML',
                         reply_markup=create_keyboard('main', chat_id))

        user_data[chat_id]['state'] = 'main_menu'

    elif text == 'Сохранить диван':
        sofa_code = generate_sofa_code(current_modules)
        prices = calculate_sofa_prices(current_modules)

        sofa_data = {
            'code': sofa_code,
            'modules': current_modules,
            'prices': prices,
            'chat_id': chat_id
        }

        if save_sofa(sofa_data):
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id, f"✅ Диван сохранен!",
                             reply_markup=create_keyboard('main', chat_id))
            user_data[chat_id]['state'] = 'main_menu'
        else:
            bot.send_message(chat_id, "❌ Ошибка при сохранении дивана")

    elif text == 'Добавить в корзину':
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        available_materials = get_available_materials_for_sofa(current_modules)
        if not available_materials:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['selected_sofa'] = {
            'sofa_code': sofa_code,
            'modules': current_modules
        }
        user_data[chat_id]['state'] = 'selecting_material_method'

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))

    elif text == 'Оформить заказ':
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        available_materials = get_available_materials_for_sofa(current_modules)
        if not available_materials:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['selected_sofa'] = {
            'sofa_code': sofa_code,
            'modules': current_modules
        }
        user_data[chat_id]['state'] = 'selecting_material_method'
        user_data[chat_id]['after_material_selection'] = 'proceed_to_order'

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'selecting_base_module'
        sofa_type = user_data[chat_id]['current_sofa_type']
        bot.send_message(chat_id, "🔧 Выберите базовый модуль:",
                         reply_markup=create_keyboard('base_module', chat_id))


def handle_extension_module_selection(message):
    """Обработчик выбора модуля расширения"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        user_data[chat_id]['state'] = 'constructing_sofa'
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)

        available_extensions = get_available_extensions(sofa_code)
        exclude_buttons = []
        if not available_extensions:
            exclude_buttons.append('Добавить модуль')

        bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                         reply_markup=create_keyboard('construction', chat_id, exclude_buttons=exclude_buttons))
        return

    available_extensions = user_data[chat_id].get('available_extensions', [])
    selected_module = None

    for module in available_extensions:
        if module['name'] == text:
            selected_module = module
            break

    if not selected_module:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите модуль из списка.")
        return

    user_data[chat_id]['current_modules'].append(selected_module)
    user_data[chat_id]['state'] = 'constructing_sofa'

    current_modules = user_data[chat_id]['current_modules']
    new_sofa_code = generate_sofa_code(current_modules)

    delete_previous_messages(chat_id, message.message_id)
    bot.send_message(chat_id, f"✅ Модуль '{selected_module['name']}' добавлен")
    show_sofa_card(chat_id, current_modules, new_sofa_code)

    new_available_extensions = get_available_extensions(new_sofa_code)
    new_exclude_buttons = []
    if not new_available_extensions:
        new_exclude_buttons.append('Добавить модуль')

    bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                     reply_markup=create_keyboard('construction', chat_id, exclude_buttons=new_exclude_buttons))


def process_code_input(chat_id, code, context='view'):
    """Обрабатывает введенный код"""
    product = find_product_by_code(code)
    if product:
        if context == 'view':
            user_data[chat_id]['selected_product'] = product
            user_data[chat_id]['state'] = 'product_view_confirmation'
            delete_previous_messages(chat_id, user_data[chat_id].get('last_anchor_message_id', 0))
            show_product(chat_id, product, show_navigation=False)
            bot.send_message(chat_id, message_templates['product_view_confirmation_text'],
                             reply_markup=create_keyboard('product_view_confirmation', chat_id),
                             parse_mode='HTML')
        else:
            user_data[chat_id]['selected_product'] = product
            user_data[chat_id]['state'] = 'selecting_product_material_method'
            delete_previous_messages(chat_id, user_data[chat_id].get('last_anchor_message_id', 0))
            show_product(chat_id, product, show_navigation=True)
            bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                             reply_markup=create_keyboard('material_method_product', chat_id))
        return True

    modules = parse_sofa_code(code)
    if modules and get_sofa_schema(code):
        if context == 'view':
            user_data[chat_id]['current_modules'] = modules
            user_data[chat_id]['editing_sofa_code'] = code
            user_data[chat_id]['state'] = 'editing_sofa'
            delete_previous_messages(chat_id, user_data[chat_id].get('last_anchor_message_id', 0))
            show_sofa_card(chat_id, modules, code)
            bot.send_message(chat_id, "🛠️ Что вы хотите сделать с этим диваном?",
                             reply_markup=create_keyboard('editing_sofa', chat_id))
        else:
            user_data[chat_id]['selected_sofa'] = {
                'sofa_code': code,
                'modules': modules
            }
            user_data[chat_id]['state'] = 'selecting_material_method'
            delete_previous_messages(chat_id, user_data[chat_id].get('last_anchor_message_id', 0))
            show_sofa_card(chat_id, modules, code)
            bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
            bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                             reply_markup=create_keyboard('material_method_sofa', chat_id))
        return True

    return False


def handle_sofa_code_input(message):
    """Обработчик ввода кода дивана"""
    chat_id = message.chat.id
    text = message.text.strip().upper()

    if text == 'НАЗАД':
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))
        return

    if not process_code_input(chat_id, text, context='view'):
        bot.send_message(chat_id, "❌ Неверный код. Попробуйте еще раз:")


def handle_product_view_confirmation(message):
    """Обработчик подтверждения просмотра продукта"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Подтвердить':
        product = user_data[chat_id].get('selected_product')
        if product:
            available_materials = get_available_materials_for_product(product)
            if not available_materials:
                bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
                user_data[chat_id]['state'] = 'main_menu'
                bot.send_message(chat_id, "Главное меню:",
                                 reply_markup=create_keyboard('main', chat_id))
                return

        user_data[chat_id]['state'] = 'selecting_product_material_method'
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id))
    elif text == 'Отмена':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))


def handle_sofa_editing(message):
    """Обработчик редактирования дивана"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Редактировать диван':
        user_data[chat_id]['state'] = 'constructing_sofa'
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)

        available_extensions = get_available_extensions(sofa_code)
        exclude_buttons = []
        if not available_extensions:
            exclude_buttons.append('Добавить модуль')

        bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                         reply_markup=create_keyboard('construction', chat_id, exclude_buttons=exclude_buttons))

    elif text == 'Сохранить как новый':
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)
        prices = calculate_sofa_prices(current_modules)

        sofa_data = {
            'code': sofa_code,
            'modules': current_modules,
            'prices': prices,
            'chat_id': chat_id
        }

        if save_sofa(sofa_data):
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id, f"✅ Диван сохранен как новый!",
                             reply_markup=create_keyboard('main', chat_id))
            user_data[chat_id]['state'] = 'main_menu'
        else:
            bot.send_message(chat_id, "❌ Ошибка при сохранении дивана")

    elif text == 'Добавить в заказ':
        current_modules = user_data[chat_id]['current_modules']
        sofa_code = generate_sofa_code(current_modules)

        available_materials = get_available_materials_for_sofa(current_modules)
        if not available_materials:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['selected_sofa'] = {
            'sofa_code': sofa_code,
            'modules': current_modules
        }
        user_data[chat_id]['state'] = 'selecting_material_method'

        delete_previous_messages(chat_id, message.message_id)
        show_sofa_card(chat_id, current_modules, sofa_code)
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))


# ===== ОБРАБОТЧИКИ ВЫБОРА МАТЕРИАЛА =====
def handle_sofa_material_method(message):
    """Обработчик выбора метода выбора материала для дивана"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Ввести код материала':
        user_data[chat_id]['state'] = 'entering_material_code'
        bot.send_message(chat_id, "⌨️ Введите код материала:",
                         reply_markup=create_keyboard('material_code_input', chat_id))

    elif text == 'Выбрать через бот':
        selected_sofa = user_data[chat_id].get('selected_sofa', {})
        modules = selected_sofa.get('modules', [])

        available_material_types = get_available_materials_for_sofa(modules)

        if not available_material_types:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['state'] = 'selecting_material_type'
        user_data[chat_id]['available_material_types'] = available_material_types

        bot.send_message(chat_id, "🧵 Выберите тип материала:",
                         reply_markup=create_keyboard('material_selection', chat_id,
                                                      available_material_types=available_material_types))

    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        if (user_data[chat_id].get('current_sofa_type') is not None and
                user_data[chat_id].get('current_modules') is not None):
            user_data[chat_id]['state'] = 'constructing_sofa'
            current_modules = user_data[chat_id]['current_modules']
            sofa_code = generate_sofa_code(current_modules)
            show_sofa_card(chat_id, current_modules, sofa_code)

            available_extensions = get_available_extensions(sofa_code)
            exclude_buttons = []
            if not available_extensions:
                exclude_buttons.append('Добавить модуль')

            bot.send_message(chat_id, "🛠️ Вы можете добавить модули или завершить конструирование:",
                             reply_markup=create_keyboard('construction', chat_id, exclude_buttons=exclude_buttons))
        else:
            user_data[chat_id]['state'] = 'main_menu'
            bot.send_message(chat_id, "Главное меню:",
                             reply_markup=create_keyboard('main', chat_id))


def handle_product_material_method(message):
    """Обработчик выбора метода выбора материала для продукта"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Ввести код материала':
        user_data[chat_id]['state'] = 'entering_material_code_for_product'
        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "⌨️ Введите код материала:",
                         reply_markup=create_keyboard('material_code_input_product', chat_id))

    elif text == 'Выбрать через бот':
        product = user_data[chat_id].get('selected_product', {})

        available_material_types = get_available_materials_for_product(product)

        if not available_material_types:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['state'] = 'selecting_material_type_for_product'
        user_data[chat_id]['available_material_types'] = available_material_types

        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите тип материала:",
                         reply_markup=create_keyboard('material_selection', chat_id,
                                                      available_material_types=available_material_types))

    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'order_menu'
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')


def handle_material_code_input_for_product(message):
    """Обработчик ввода кода материала для продукта"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_product_material_method'
        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id))
        return

    code = text.strip()
    material = find_material_by_code(code)
    if not material:
        bot.send_message(chat_id, "❌ Материал с таким кодом не найден. Попробуйте еще раз:")
        return

    product = user_data[chat_id].get('selected_product', {})
    if not is_material_available_for_product(product, material['type']):
        bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
        return

    caption = message_templates['material_info'].format(
        type=material['type'],
        color=material['color'],
        code=material['code'],
        link=f"Ссылка: {material['link']}" if material['link'] else ""
    )

    bot.send_message(chat_id, caption, parse_mode='HTML')

    photos_sent = send_material_photos(chat_id, material['code'])
    if not photos_sent:
        bot.send_message(chat_id, "📷 Фотографии отсутствуют")

    user_data[chat_id]['selected_material'] = material
    user_data[chat_id]['state'] = 'confirming_material_for_product'
    bot.send_message(chat_id, "✅ Подтвердите выбор материала:",
                     reply_markup=create_keyboard('material_confirmation_product', chat_id))


def handle_material_confirmation_for_product(message):
    """Обработчик подтверждения выбора материала для продукта"""
    chat_id = message.chat.id
    text = message.text

    if 'selected_product' not in user_data[chat_id] or 'selected_material' not in user_data[chat_id]:
        bot.send_message(chat_id, "❌ Ошибка: отсутствуют данные о продукте или материале. Начните заново.")
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:", reply_markup=create_keyboard('main', chat_id))
        return

    if text == 'Выбрать этот материал':
        material = user_data[chat_id]['selected_material']
        product = user_data[chat_id]['selected_product']

        if not is_material_available_for_product(product, material['type']):
            bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
            return

        price_key = material_type_mapping.get(material['type'])
        if price_key and price_key in product['prices']:
            price = product['prices'][price_key]
        else:
            price = 0

        save_to_cart(
            chat_id=chat_id,
            item_type='product',
            item_code=product['code'],
            item_name=product['name'],
            material_code=material['code'],
            material_type=material['type'],
            material_color=material['color'],
            price=price
        )

        user_data[chat_id]['state'] = 'order_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            f"✅ <b>{product['name']}</b> добавлен в корзину!\n"
            f"🧵 Материал: {material['type']}: {material['color']}\n"
            f"💰 Цена: {price} руб",
            reply_markup=create_keyboard('order_menu', chat_id),
            parse_mode='HTML'
        )

        if 'selected_product' in user_data[chat_id]:
            del user_data[chat_id]['selected_product']
        if 'selected_material' in user_data[chat_id]:
            del user_data[chat_id]['selected_material']

    elif text == 'Ввести код заново':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'entering_material_code_for_product'
        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "⌨️ Введите код материала:",
                         reply_markup=create_keyboard('material_code_input_product', chat_id), parse_mode='HTML')

    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_product_material_method'
        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id))


def handle_material_type_selection(message):
    """Обработчик выбора типа материала"""
    chat_id = message.chat.id
    text = message.text
    available_material_types = user_data[chat_id].get('available_material_types', [])

    if text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_material_method'
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))
        return

    if text not in available_material_types:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите тип материала из списка.")
        return

    selected_sofa = user_data[chat_id].get('selected_sofa', {})
    modules = selected_sofa.get('modules', [])

    materials = get_materials_by_type_with_availability(text, modules=modules)

    if not materials:
        bot.send_message(chat_id, "⚠️ Нет доступных материалов этого типа для данного дивана")
        return

    user_data[chat_id]['current_material_type'] = text
    user_data[chat_id]['current_material_index'] = 0
    user_data[chat_id]['state'] = 'viewing_materials'
    delete_previous_messages(chat_id, message.message_id)
    show_material(chat_id, materials[0])


def handle_material_type_selection_for_product(message):
    """Обработчик выбора типа материала для продукта"""
    chat_id = message.chat.id
    text = message.text
    available_material_types = user_data[chat_id].get('available_material_types', [])

    if text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_product_material_method'
        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id))
        return

    if text not in available_material_types:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите тип материала из списка.")
        return

    product = user_data[chat_id].get('selected_product', {})

    materials = get_materials_by_type_with_availability(text, product=product)

    if not materials:
        bot.send_message(chat_id, "⚠️ Нет доступных материалов этого типа для данного продукта")
        return

    user_data[chat_id]['current_material_type'] = text
    user_data[chat_id]['current_material_index'] = 0
    user_data[chat_id]['state'] = 'viewing_materials_for_product'
    delete_previous_messages(chat_id, message.message_id)
    show_material(chat_id, materials[0])


def handle_material_navigation(message):
    """Обработчик навигации по материалам для дивана"""
    chat_id = message.chat.id
    text = message.text
    material_type = user_data[chat_id]['current_material_type']
    current_index = user_data[chat_id]['current_material_index']

    selected_sofa = user_data[chat_id].get('selected_sofa', {})
    modules = selected_sofa.get('modules', [])

    materials = get_materials_by_type_with_availability(material_type, modules=modules)

    if text == 'Следующий':
        new_index = (current_index + 1) % len(materials)
        user_data[chat_id]['current_material_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_material(chat_id, materials[new_index])
    elif text == 'Предыдущий':
        new_index = (current_index - 1) % len(materials)
        user_data[chat_id]['current_material_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_material(chat_id, materials[new_index])
    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_material_type'
        available_material_types = user_data[chat_id].get('available_material_types', [])
        bot.send_message(chat_id, "🧵 Выберите тип материала:",
                         reply_markup=create_keyboard('material_selection', chat_id,
                                                      available_material_types=available_material_types))
    elif text == 'Добавить в корзину':
        if 'selected_sofa' not in user_data[chat_id]:
            bot.send_message(chat_id, "❌ Ошибка: отсутствуют данные о диване. Начните заново.")
            user_data[chat_id]['state'] = 'main_menu'
            bot.send_message(chat_id, "Главное меню:", reply_markup=create_keyboard('main', chat_id))
            return

        material = materials[current_index]
        selected_sofa = user_data[chat_id]['selected_sofa']

        if not is_material_available_for_sofa(selected_sofa['modules'], material['type']):
            bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
            return

        prices = calculate_sofa_prices(selected_sofa['modules'])
        price = prices[material['type']]

        save_to_cart(
            chat_id=chat_id,
            item_type='sofa',
            item_code=selected_sofa['sofa_code'],
            item_name=f"Диван {selected_sofa['sofa_code']}",
            material_code=material['code'],
            material_type=material['type'],
            material_color=material['color'],
            price=price
        )

        if user_data[chat_id].get('after_material_selection') == 'proceed_to_order':
            user_data[chat_id]['state'] = 'collecting_customer_data'
            user_data[chat_id]['order_data'] = {'state': 'fio'}
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id, "📝 Введите ваше ФИО (Полностью):",
                             reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')
            if 'after_material_selection' in user_data[chat_id]:
                del user_data[chat_id]['after_material_selection']
        else:
            user_data[chat_id]['state'] = 'order_menu'
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(
                chat_id,
                f"✅ <b>Диван {selected_sofa['sofa_code']}</b> добавлен в корзину!\n"
                f"🧵 Материал: {material['type']}: {material['color']}\n"
                f"💰 Цена: {price} руб",
                reply_markup=create_keyboard('order_menu', chat_id),
                parse_mode='HTML'
            )


def handle_material_navigation_for_product(message):
    """Обработчик навигации по материалам для продукта"""
    chat_id = message.chat.id
    text = message.text
    material_type = user_data[chat_id]['current_material_type']
    current_index = user_data[chat_id]['current_material_index']

    product = user_data[chat_id].get('selected_product', {})

    materials = get_materials_by_type_with_availability(material_type, product=product)

    if text == 'Следующий':
        new_index = (current_index + 1) % len(materials)
        user_data[chat_id]['current_material_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_material(chat_id, materials[new_index])
    elif text == 'Предыдущий':
        new_index = (current_index - 1) % len(materials)
        user_data[chat_id]['current_material_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_material(chat_id, materials[new_index])
    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_material_type_for_product'
        available_material_types = user_data[chat_id].get('available_material_types', [])

        product = user_data[chat_id].get('selected_product', {})
        product_name = product.get('name', 'Продукт')
        product_code = product.get('code', '')

        bot.send_message(chat_id,
                         message_templates['material_for_product'].format(
                             product_name=product_name,
                             product_code=product_code
                         ),
                         parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите тип материала:",
                         reply_markup=create_keyboard('material_selection', chat_id,
                                                      available_material_types=available_material_types))
    elif text == 'Добавить в корзину':
        if 'selected_product' not in user_data[chat_id]:
            bot.send_message(chat_id, "❌ Ошибка: отсутствуют данные о продукте. Начните заново.")
            user_data[chat_id]['state'] = 'main_menu'
            bot.send_message(chat_id, "Главное меню:", reply_markup=create_keyboard('main', chat_id))
            return

        material = materials[current_index]
        product = user_data[chat_id]['selected_product']

        if not is_material_available_for_product(product, material['type']):
            bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
            return

        price_key = material_type_mapping.get(material['type'])
        if price_key and price_key in product['prices']:
            price = product['prices'][price_key]
        else:
            price = 0

        save_to_cart(
            chat_id=chat_id,
            item_type='product',
            item_code=product['code'],
            item_name=product['name'],
            material_code=material['code'],
            material_type=material['type'],
            material_color=material['color'],
            price=price
        )

        user_data[chat_id]['state'] = 'order_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            f"✅ <b>{product['name']}</b> добавлен в корзину!\n"
            f"🧵 Материал: {material['type']}: {material['color']}\n"
            f"💰 Цена: {price} руб",
            reply_markup=create_keyboard('order_menu', chat_id),
            parse_mode='HTML'
        )

        if 'selected_product' in user_data[chat_id]:
            del user_data[chat_id]['selected_product']
        if 'selected_material' in user_data[chat_id]:
            del user_data[chat_id]['selected_material']


def handle_material_code_input(message):
    """Обработчик ввода кода материала"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_material_method'
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))
        return

    code = text.strip()
    material = find_material_by_code(code)
    if not material:
        bot.send_message(chat_id, "❌ Материал с таким кодом не найден. Попробуйте еще раз:")
        return

    selected_sofa = user_data[chat_id].get('selected_sofa', {})
    modules = selected_sofa.get('modules', [])
    if not is_material_available_for_sofa(modules, material['type']):
        bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
        return

    caption = message_templates['material_info'].format(
        type=material['type'],
        color=material['color'],
        code=material['code'],
        link=f"Ссылка: {material['link']}" if material['link'] else ""
    )

    bot.send_message(chat_id, caption, parse_mode='HTML')

    photos_sent = send_material_photos(chat_id, material['code'])
    if not photos_sent:
        bot.send_message(chat_id, "📷 Фотографии отсутствуют")

    user_data[chat_id]['selected_material'] = material
    user_data[chat_id]['state'] = 'confirming_material'
    bot.send_message(chat_id, "✅ Подтвердите выбор материала:",
                     reply_markup=create_keyboard('material_confirmation', chat_id))


def handle_material_confirmation(message):
    """Обработчик подтверждения выбора материала"""
    chat_id = message.chat.id
    text = message.text

    if 'selected_sofa' not in user_data[chat_id] or 'selected_material' not in user_data[chat_id]:
        bot.send_message(chat_id, "❌ Ошибка: отсутствуют данные о диване или материале. Начните заново.")
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:", reply_markup=create_keyboard('main', chat_id))
        return

    if text == 'Выбрать этот материал':
        material = user_data[chat_id]['selected_material']
        selected_sofa = user_data[chat_id]['selected_sofa']

        if not is_material_available_for_sofa(selected_sofa['modules'], material['type']):
            bot.send_message(chat_id, message_templates['material_not_available'], parse_mode='HTML')
            return

        prices = calculate_sofa_prices(selected_sofa['modules'])
        price = prices[material['type']]

        save_to_cart(
            chat_id=chat_id,
            item_type='sofa',
            item_code=selected_sofa['sofa_code'],
            item_name=f"Диван {selected_sofa['sofa_code']}",
            material_code=material['code'],
            material_type=material['type'],
            material_color=material['color'],
            price=price
        )

        if user_data[chat_id].get('after_material_selection') == 'proceed_to_order':
            user_data[chat_id]['state'] = 'collecting_customer_data'
            user_data[chat_id]['order_data'] = {'state': 'fio'}
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id, "📝 Введите ваше ФИО (Полностью):",
                             reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')
            if 'after_material_selection' in user_data[chat_id]:
                del user_data[chat_id]['after_material_selection']
        else:
            user_data[chat_id]['state'] = 'order_menu'
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(
                chat_id,
                f"✅ <b>Диван {selected_sofa['sofa_code']}</b> добавлен в корзину!\n"
                f"🧵 Материал: {material['type']}: {material['color']}\n"
                f"💰 Цена: {price} руб",
                reply_markup=create_keyboard('order_menu', chat_id),
                parse_mode='HTML'
            )

    elif text == 'Ввести код заново':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'entering_material_code'
        bot.send_message(chat_id, "⌨️ Введите код материала:",
                         reply_markup=create_keyboard('material_code_input', chat_id), parse_mode='HTML')

    elif text == 'Назад':
        delete_previous_messages(chat_id, message.message_id)
        user_data[chat_id]['state'] = 'selecting_material_method'
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_sofa', chat_id))


# ===== ОБРАБОТЧИКИ ЗАКАЗОВ =====
def handle_order_menu(message):
    """Обработчик меню оформления заказа"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Добавить диван в заказ':
        user_data[chat_id]['state'] = 'entering_sofa_code_for_order'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта для добавления в заказ:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id), parse_mode='HTML')

    elif text == 'Добавить позицию':
        user_data[chat_id]['state'] = 'selecting_product_type'
        bot.send_message(chat_id, "📦 Выберите тип продукта:",
                         reply_markup=create_keyboard('product_type'), parse_mode='HTML')

    elif text == 'Посмотреть корзину':
        view_cart(chat_id)

    elif text == 'Ввести код дивана':
        user_data[chat_id]['state'] = 'entering_sofa_code_for_order'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id), parse_mode='HTML')

    elif text == 'Отправить заказ на обработку':
        cart_items = get_user_cart(chat_id)
        if not cart_items:
            bot.send_message(chat_id, "🛒 Ваша корзина пуста!",
                             reply_markup=create_keyboard('cart_management', chat_id))
            return

        user_data[chat_id]['state'] = 'collecting_customer_data'
        user_data[chat_id]['order_data'] = {'state': 'fio'}
        bot.send_message(chat_id, "📝 Введите ваше ФИО (Полностью):",
                         reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id), parse_mode='HTML')


def handle_sofa_code_for_order_input(message):
    """Обработчик ввода кода для заказа"""
    chat_id = message.chat.id
    text = message.text.strip().upper()

    if text == 'НАЗАД':
        user_data[chat_id]['state'] = 'order_menu'
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')
        return

    if not process_code_input(chat_id, text, context='order'):
        bot.send_message(chat_id, "❌ Неверный код. Попробуйте еще раз:")


# ===== ОБРАБОТКА ДАННЫХ ЗАКАЗА =====
def handle_collecting_customer_data(message):
    """Обработчик сбора данных заказчика"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        user_data[chat_id]['state'] = 'order_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')
        return

    if 'order_data' not in user_data[chat_id]:
        user_data[chat_id]['order_data'] = {}

    order_data = user_data[chat_id]['order_data']

    if 'state' not in order_data:
        order_data['state'] = 'fio'

    current_state = order_data['state']

    if len(text) > field_limits.get(current_state, 1000):
        bot.send_message(chat_id,
                         f"❌ Превышено ограничение длины ({field_limits[current_state]} символов). Пожалуйста, введите значение заново:",
                         reply_markup=create_keyboard('collecting_data', chat_id))
        return

    order_data[current_state] = text

    if current_state == 'fio':
        order_data['state'] = 'phone'
        bot.send_message(chat_id, "📱 Введите ваш номер телефона:",
                         reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')
    elif current_state == 'phone':
        order_data['state'] = 'address'
        bot.send_message(chat_id, "🏠 Введите ваш адрес доставки:",
                         reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')
    elif current_state == 'address':
        order_data['state'] = 'comment'
        bot.send_message(chat_id, "💬 Введите комментарий к заказу (необязательно):",
                         reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')
    elif current_state == 'comment':
        order_data[
            'telegram_contact'] = f"https://t.me/{message.from_user.username}" if message.from_user.username else f"ID: {message.from_user.id}"
        order_data['task_number'] = generate_task_number()
        order_data['order_code'] = generate_order_code()
        order_data['chat_id'] = chat_id

        user_data[chat_id]['state'] = 'confirming_order'
        delete_previous_messages(chat_id, message.message_id)
        confirm_order(chat_id)


def handle_order_confirmation(message):
    """Обработчик подтверждения заказа"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Подтвердить заказ':
        order_data = user_data[chat_id]['order_data']
        order_data['cart'] = get_user_cart(chat_id)

        if save_order(order_data):
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(chat_id,
                             f"✅ <b>Заказ успешно оформлен!</b>\n"
                             f"📋 Номер заказа: <code>{order_data['order_code']}</code>\n"
                             f"📞 С вами свяжутся для уточнения деталей.",
                             parse_mode='HTML',
                             reply_markup=create_keyboard('main', chat_id))

            # Отправляем уведомление менеджеру
            try:
                bot.send_message(MANAGER_CHAT_ID, "Появился новый заказ")
            except Exception as e:
                print(f"Ошибка отправки уведомления менеджеру: {e}")

            clear_user_cart(chat_id)
            user_data[chat_id]['order_data'] = {}
            user_data[chat_id]['state'] = 'main_menu'
        else:
            bot.send_message(chat_id, "❌ Ошибка при сохранении заказа. Попробуйте еще раз.")

    elif text == 'Отменить заказ':
        user_data[chat_id]['state'] = 'order_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')


# ===== ОБРАБОТЧИК ПРОСМОТРА ЗАКАЗОВ =====
def handle_viewing_my_orders(message):
    """Обработчик просмотра заказов"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Оформить заказ':
        user_data[chat_id]['state'] = 'order_menu'
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')

    elif text == 'Ввести код дивана':
        user_data[chat_id]['state'] = 'entering_sofa_code'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id))

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id))

    elif text == 'Перезапустить бота':
        restart_bot(message)

    elif text == 'Перейти на сайт':
        go_to_site(message)


# ===== НОВЫЙ ОБРАБОТЧИК ДЛЯ ДЕТАЛЕЙ ЗАКАЗА =====
def handle_order_details_navigation(message):
    """Обработчик навигации по деталям заказа"""
    chat_id = message.chat.id
    text = message.text

    if 'viewing_order_code' not in user_data[chat_id]:
        user_data[chat_id]['state'] = 'viewing_my_orders'
        show_user_orders(chat_id)
        return

    order_code = user_data[chat_id]['viewing_order_code']
    current_index = user_data[chat_id]['viewing_order_index']
    items = user_data[chat_id]['viewing_order_items']

    if text == 'Предыдущая позиция':
        if current_index > 0:
            delete_previous_messages(chat_id, message.message_id)
            show_order_details(chat_id, order_code, current_index - 1)
        else:
            bot.send_message(chat_id, "⚠️ Это первая позиция в заказе")

    elif text == 'Следующая позиция':
        if current_index < len(items) - 1:
            delete_previous_messages(chat_id, message.message_id)
            show_order_details(chat_id, order_code, current_index + 1)
        else:
            bot.send_message(chat_id, "⚠️ Это последняя позиция в заказе")

    elif text == 'Назад к заказу':
        user_data[chat_id]['state'] = 'viewing_my_orders'
        delete_previous_messages(chat_id, message.message_id)
        show_user_orders(chat_id)


# ===== ОБРАБОТЧИКИ ПРОДУКТОВ =====
def handle_product_type_selection(message):
    """Обработчик выбора типа продукта"""
    chat_id = message.chat.id
    text = message.text
    products, unique_types = load_products()

    if text == 'Назад':
        user_data[chat_id]['state'] = 'order_menu'
        bot.send_message(chat_id, "🛒 Меню оформления заказа:",
                         reply_markup=create_keyboard('order_menu', chat_id), parse_mode='HTML')
        return

    if text == 'Ввести код позиции':
        user_data[chat_id]['state'] = 'entering_product_code'
        bot.send_message(chat_id, "⌨️ Введите код продукта:",
                         reply_markup=create_keyboard('product_code_input'), parse_mode='HTML')
        return

    if text not in unique_types:
        bot.send_message(chat_id, "ℹ️ Пожалуйста, выберите тип из списка.")
        return

    user_data[chat_id]['current_product_type'] = text
    user_data[chat_id]['current_product_index'] = 0
    user_data[chat_id]['state'] = 'viewing_products'

    products_of_type = [p for p in products if p['type'] == text]
    if products_of_type:
        delete_previous_messages(chat_id, message.message_id)
        show_product(chat_id, products_of_type[0], show_navigation=True)
    else:
        bot.send_message(chat_id, "❌ Нет продуктов этого типа")


def handle_product_navigation(message):
    """Обработчик навигации по продуктам"""
    chat_id = message.chat.id
    text = message.text
    current_type = user_data[chat_id]['current_product_type']
    current_index = user_data[chat_id]['current_product_index']
    products, _ = load_products()
    products_of_type = [p for p in products if p['type'] == current_type]

    if text == 'Следующий':
        new_index = (current_index + 1) % len(products_of_type)
        user_data[chat_id]['current_product_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_product(chat_id, products_of_type[new_index], show_navigation=True)
    elif text == 'Предыдущий':
        new_index = (current_index - 1) % len(products_of_type)
        user_data[chat_id]['current_product_index'] = new_index
        delete_previous_messages(chat_id, message.message_id)
        show_product(chat_id, products_of_type[new_index], show_navigation=True)
    elif text == 'Назад':
        user_data[chat_id]['state'] = 'selecting_product_type'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, "📦 Выберите тип продукта:",
                         reply_markup=create_keyboard('product_type'), parse_mode='HTML')
    elif text == 'Перейти на сайт':
        go_to_site(message)
    elif text == 'Добавить в корзину':
        product = products_of_type[current_index]

        available_materials = get_available_materials_for_product(product)
        if not available_materials:
            bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
            return

        user_data[chat_id]['selected_product'] = product
        user_data[chat_id]['state'] = 'selecting_product_material_method'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id))


def handle_product_code_input(message):
    """Обработчик ввода кода продукта"""
    chat_id = message.chat.id
    text = message.text.strip().upper()

    if text == 'НАЗАД':
        user_data[chat_id]['state'] = 'selecting_product_type'
        bot.send_message(chat_id, "📦 Выберите тип продукта:",
                         reply_markup=create_keyboard('product_type'), parse_mode='HTML')
        return

    product = find_product_by_code(text)
    if not product:
        bot.send_message(chat_id, "❌ Продукт с таким кодом не найден. Попробуйте еще раз:")
        return

    available_materials = get_available_materials_for_product(product)
    if not available_materials:
        bot.send_message(chat_id, message_templates['no_available_materials'], parse_mode='HTML')
        return

    user_data[chat_id]['selected_product'] = product
    user_data[chat_id]['state'] = 'selecting_product_material_method'
    delete_previous_messages(chat_id, message.message_id)
    show_product(chat_id, product, show_navigation=True)
    bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                     reply_markup=create_keyboard('material_method_product', chat_id))


# ===== ОБРАБОТЧИК КОРЗИНЫ =====
def handle_viewing_cart(message):
    """Обработчик просмотра корзины"""
    chat_id = message.chat.id
    text = message.text

    delete_previous_messages(chat_id, message.message_id)

    if text == 'Добавить позицию':
        user_data[chat_id]['state'] = 'selecting_product_type'
        bot.send_message(chat_id, "📦 Выберите тип продукта:",
                         reply_markup=create_keyboard('product_type'), parse_mode='HTML')

    elif text == 'Ввести код дивана':
        user_data[chat_id]['state'] = 'entering_sofa_code_for_order'
        bot.send_message(chat_id, "⌨️ Введите код дивана или продукта для добавления в заказ:",
                         reply_markup=create_keyboard('sofa_code_input', chat_id), parse_mode='HTML')

    elif text == 'Отправить заказ на обработку':
        cart_items = get_user_cart(chat_id)
        if not cart_items:
            bot.send_message(chat_id, "🛒 Ваша корзина пуста!",
                             reply_markup=create_keyboard('cart_management', chat_id))
            return

        user_data[chat_id]['state'] = 'collecting_customer_data'
        user_data[chat_id]['order_data'] = {'state': 'fio'}
        bot.send_message(chat_id, "📝 Введите ваше ФИО (Полностью):",
                         reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')

    elif text == 'Назад':
        user_data[chat_id]['state'] = 'main_menu'
        bot.send_message(chat_id, "Главное меню:",
                         reply_markup=create_keyboard('main', chat_id), parse_mode='HTML')

    elif text == 'Перезапустить бота':
        restart_bot(message)

    elif text == 'Перейти на сайте':
        go_to_site(message)


def handle_view_cart_item(message):
    """Обработчик просмотра деталей позиции"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад в корзину':
        user_data[chat_id]['state'] = 'viewing_cart'
        delete_previous_messages(chat_id, message.message_id)
        view_cart(chat_id)


def handle_viewing_saved_sofa(message):
    """Обработчик просмотра сохраненного дивана"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад к списку':
        user_data[chat_id]['state'] = 'viewing_saved_sofas'
        delete_previous_messages(chat_id, message.message_id)
        show_user_saved_sofas(chat_id)
        bot.send_message(chat_id, "Выберите действие:",
                         reply_markup=create_keyboard('saved_sofas', chat_id))


def handle_order_field_edit(message):
    """Обработчик редактирования поля заказа"""
    chat_id = message.chat.id
    text = message.text

    if text == 'Назад':
        user_data[chat_id]['state'] = 'confirming_order'
        delete_previous_messages(chat_id, message.message_id)
        confirm_order(chat_id)
        return

    field = user_data[chat_id]['editing_field']

    if len(text) > field_limits.get(field, 1000):
        bot.send_message(chat_id,
                         f"❌ Превышено ограничение длины ({field_limits[field]} символов). Пожалуйста, введите значение заново:",
                         reply_markup=create_keyboard('collecting_data', chat_id))
        return

    user_data[chat_id]['order_data'][field] = text
    user_data[chat_id]['state'] = 'confirming_order'
    delete_previous_messages(chat_id, message.message_id)
    confirm_order(chat_id)


# ===== CALLBACK ОБРАБОТЧИКИ =====
@bot.callback_query_handler(func=lambda call: call.data.startswith('edit_sofa_'))
def edit_saved_sofa(call):
    """Редактирование сохраненного дивана"""
    chat_id = call.message.chat.id
    sofa_code = call.data.split('_')[2]

    user_sofas = get_user_sofas(chat_id)
    if sofa_code not in user_sofas:
        bot.answer_callback_query(call.id, "❌ Диван не найден")
        return

    sofa_data = user_sofas[sofa_code]
    user_data[chat_id]['current_modules'] = sofa_data['modules']
    user_data[chat_id]['editing_sofa_code'] = sofa_code
    user_data[chat_id]['state'] = 'editing_sofa'

    delete_previous_messages(chat_id, call.message.message_id)
    show_sofa_card(chat_id, sofa_data['modules'], sofa_code)

    bot.send_message(chat_id, "🛠️ Что вы хотите сделать с этим диваном?",
                     reply_markup=create_keyboard('editing_sofa', chat_id))


@bot.callback_query_handler(func=lambda call: call.data.startswith('delete_sofa_'))
def delete_saved_sofa(call):
    """Удаление сохраненного дивана"""
    chat_id = call.message.chat.id
    sofa_code = call.data.split('_')[2]

    try:
        if not os.path.exists(sofas_log):
            bot.answer_callback_query(call.id, "❌ Файл с диванами не найден")
            return

        wb = openpyxl.load_workbook(sofas_log)
        sheet = wb.active
        row_to_delete = None

        for row in range(2, sheet.max_row + 1):
            if (str(sheet.cell(row=row, column=1).value) == sofa_code and
                    str(sheet.cell(row=row, column=7).value) == str(chat_id)):
                row_to_delete = row
                break

        if row_to_delete:
            sheet.delete_rows(row_to_delete)
            wb.save(sofas_log)
            bot.answer_callback_query(call.id, "✅ Диван удален")
            delete_previous_messages(chat_id, call.message.message_id)
            show_user_saved_sofas(chat_id)
            bot.send_message(chat_id, "Выберите действие:",
                             reply_markup=create_keyboard('saved_sofas', chat_id))
        else:
            bot.answer_callback_query(call.id, "❌ Диван не найден")

    except Exception as e:
        print(f"Ошибка удаления дивана: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка при удалении")


@bot.callback_query_handler(func=lambda call: call.data.startswith('add_to_cart_'))
def add_saved_sofa_to_cart(call):
    """Добавление сохраненного дивана в корзину"""
    chat_id = call.message.chat.id
    sofa_code = call.data.split('_')[3]

    user_sofas = get_user_sofas(chat_id)
    if sofa_code not in user_sofas:
        bot.answer_callback_query(call.id, "❌ Диван не найден")
        return

    sofa_data = user_sofas[sofa_code]

    available_materials = get_available_materials_for_sofa(sofa_data['modules'])
    if not available_materials:
        bot.answer_callback_query(call.id, "❌ Нет доступных материалов для этого дивана")
        return

    user_data[chat_id]['selected_sofa'] = {
        'sofa_code': sofa_code,
        'modules': sofa_data['modules']
    }
    user_data[chat_id]['state'] = 'selecting_material_method'

    delete_previous_messages(chat_id, call.message.message_id)
    show_sofa_card(chat_id, sofa_data['modules'], sofa_code)
    bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
    bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                     reply_markup=create_keyboard('material_method_sofa', chat_id))


@bot.callback_query_handler(func=lambda call: call.data.startswith('view_sofa_'))
def view_saved_sofa(call):
    """Просмотр сохраненного дивана"""
    chat_id = call.message.chat.id
    sofa_code = call.data.split('_')[2]

    user_sofas = get_user_sofas(chat_id)
    if sofa_code not in user_sofas:
        bot.answer_callback_query(call.id, "❌ Диван не найден")
        return

    sofa_data = user_sofas[sofa_code]
    user_data[chat_id]['state'] = 'viewing_saved_sofa'
    user_data[chat_id]['viewing_sofa_code'] = sofa_code

    delete_previous_messages(chat_id, call.message.message_id)
    show_sofa_card(chat_id, sofa_data['modules'], sofa_code)
    bot.send_message(chat_id, "Вы можете вернуться к списку диванов:",
                     reply_markup=create_keyboard('view_saved_sofa', chat_id))


@bot.callback_query_handler(func=lambda call: call.data.startswith('edit_cart_item_'))
def edit_cart_item(call):
    """Редактирование позиции в корзине"""
    chat_id = call.message.chat.id
    db_id = int(call.data.split('_')[3])
    delete_previous_messages(chat_id, call.message.message_id)

    cart_items = get_user_cart(chat_id)
    cart_item = None

    for item in cart_items:
        if item['db_id'] == db_id:
            cart_item = item
            break

    if not cart_item:
        bot.answer_callback_query(call.id, "❌ Позиция не найден")
        return

    if cart_item['item_type'] == 'sofa':
        modules = parse_sofa_code(cart_item['item_code'])
        if modules:
            user_data[chat_id]['editing_cart_index'] = db_id
            user_data[chat_id]['selected_sofa'] = {
                'sofa_code': cart_item['item_code'],
                'modules': modules
            }
            user_data[chat_id]['state'] = 'selecting_material_method'

            bot.send_message(chat_id, message_templates['material_selection_help'], parse_mode='HTML')
            bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                             reply_markup=create_keyboard('material_method_sofa', chat_id))
        else:
            bot.answer_callback_query(call.id, "❌ Не удалось загрузить диван")
    else:
        product = find_product_by_code(cart_item['item_code'])
        if not product:
            bot.answer_callback_query(call.id, "❌ Продукт не найден")
            return

        user_data[chat_id]['editing_index'] = db_id
        user_data[chat_id]['selected_product'] = product
        user_data[chat_id]['state'] = 'selecting_product_material_method'
        markup_inline = types.InlineKeyboardMarkup()
        markup_inline.add(types.InlineKeyboardButton("Перейти на сайт", url=site_shop))
        bot.send_message(chat_id, message_templates['material_selection_help'], reply_markup=markup_inline)
        bot.send_message(chat_id, "🧵 Выберите метод выбора материала:",
                         reply_markup=create_keyboard('material_method_product', chat_id), parse_mode='HTML')


@bot.callback_query_handler(func=lambda call: call.data.startswith('delete_cart_item_'))
def delete_cart_item(call):
    """Удаление позиции из корзины"""
    chat_id = call.message.chat.id
    db_id = int(call.data.split('_')[3])
    delete_previous_messages(chat_id, call.message.message_id)

    if remove_from_cart(chat_id, db_id):
        bot.answer_callback_query(call.id, "✅ Позиция удалена из корзины")
        view_cart(chat_id)
    else:
        bot.answer_callback_query(call.id, "❌ Ошибка при удалении позиции")


@bot.callback_query_handler(func=lambda call: call.data.startswith('view_cart_item_'))
def view_cart_item_details(call):
    """Просмотр деталей позиции в корзине"""
    chat_id = call.message.chat.id
    db_id = int(call.data.split('_')[3])

    delete_previous_messages(chat_id, call.message.message_id)
    user_data[chat_id]['state'] = 'view_cart_item'
    show_cart_item_details(chat_id, db_id)
    bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('editorder_'))
def edit_order_field_callback(call):
    """Редактирование полей заказа"""
    chat_id = call.message.chat.id
    field = call.data.split('_')[1]
    delete_previous_messages(chat_id, call.message.message_id)
    description = field_descriptions.get(field, field)
    user_data[chat_id]['editing_field'] = field
    user_data[chat_id]['state'] = 'editing_order_field'
    bot.send_message(chat_id, f"✏️ Введите новое значение для {description}:",
                     reply_markup=create_keyboard('collecting_data', chat_id), parse_mode='HTML')


# ===== ЗАПУСК БОТА =====
if __name__ == "__main__":
    print("Объединенный бот для конструирования диванов и заказов запущен...")
    print("✅ Таблица корзин создана в директории DataBase (с исправленными заголовками)")
    bot.polling(none_stop=True)