import telebot
from telebot import types
import openpyxl
import os
import traceback

# Определение абсолютных путей
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_BASE = os.path.join(BASE_DIR, 'DataBase')
os.makedirs(DATA_BASE, exist_ok=True)  # Создание папки при отсутствии

# Пути к файлам аутентификации
PASSWORD_FILE = os.path.join(DATA_BASE, 'password.txt')
ADMIN_FILE = os.path.join(DATA_BASE, 'id_workers.txt')

# Пути к папкам с фотографиями (как в клиентском боте)
PHOTO_DIR = os.path.join(BASE_DIR, 'Photo', 'Products')
PHOTO_MATERIALS_DIR = os.path.join(BASE_DIR, 'Photo', 'Materials')
os.makedirs(PHOTO_DIR, exist_ok=True)
os.makedirs(PHOTO_MATERIALS_DIR, exist_ok=True)

# Загрузка пароля и списка администраторов
PASSWORD = None
ADMIN_IDS = set()
ADMIN_DATA = {}


def load_auth_data():
    global PASSWORD, ADMIN_IDS, ADMIN_DATA
    # Загрузка пароля
    if os.path.exists(PASSWORD_FILE):
        try:
            with open(PASSWORD_FILE, 'r', encoding='utf-8') as f:
                PASSWORD = f.read().strip()
        except UnicodeDecodeError:
            # Если UTF-8 не работает, пробуем другие кодировки
            try:
                with open(PASSWORD_FILE, 'r', encoding='cp1251') as f:
                    PASSWORD = f.read().strip()
            except:
                with open(PASSWORD_FILE, 'r', encoding='latin-1') as f:
                    PASSWORD = f.read().strip()

    # Загрузка администраторов
    if os.path.exists(ADMIN_FILE):
        try:
            with open(ADMIN_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split(' - ')
                    if len(parts) >= 5:
                        try:
                            chat_id = int(parts[0])
                            ADMIN_IDS.add(chat_id)
                            ADMIN_DATA[chat_id] = {
                                'username': parts[1],
                                'link': parts[2],
                                'fio': parts[3],
                                'role': parts[4]
                            }
                        except ValueError:
                            continue
        except UnicodeDecodeError:
            # Если UTF-8 не работает, пробуем другие кодировки
            try:
                with open(ADMIN_FILE, 'r', encoding='cp1251') as f:
                    for line in f:
                        parts = line.strip().split(' - ')
                        if len(parts) >= 5:
                            try:
                                chat_id = int(parts[0])
                                ADMIN_IDS.add(chat_id)
                                ADMIN_DATA[chat_id] = {
                                    'username': parts[1],
                                    'link': parts[2],
                                    'fio': parts[3],
                                    'role': parts[4]
                                }
                            except ValueError:
                                continue
            except:
                with open(ADMIN_FILE, 'r', encoding='latin-1') as f:
                    for line in f:
                        parts = line.strip().split(' - ')
                        if len(parts) >= 5:
                            try:
                                chat_id = int(parts[0])
                                ADMIN_IDS.add(chat_id)
                                ADMIN_DATA[chat_id] = {
                                    'username': parts[1],
                                    'link': parts[2],
                                    'fio': parts[3],
                                    'role': parts[4]
                                }
                            except ValueError:
                                continue


# Инициализация данных аутентификации при запуске
load_auth_data()

bot = telebot.TeleBot('токен')

# Конфигурация путей с использованием абсолютных путей
orders_log = os.path.join(DATA_BASE, 'Orders_log.xlsx')
stuff_log = os.path.join(DATA_BASE, 'Stuff_log.xlsx')
products_log = os.path.join(DATA_BASE, 'Products_log.xlsx')
materials_log = os.path.join(DATA_BASE, 'Materials_log.xlsx')

# Фиксированный список статусов
FIXED_STATUSES = [
    "Ожидает подтверждения",
    "Ожидает оплаты",
    "В производстве",
    "Ожидает упаковки",
    "Ожидает отправки",
    "Отправлен",
    "Доставлен",
    "Отменён",
    "В архиве"
]

# Статусы готовности для задач
READINESS_STATUSES = ["нет", "в разработке", "готово"]

# Глобальные переменные для состояний
user_data = {}

# Маппинг типов материалов для корректного определения цены (как в клиентском боте)
material_type_mapping = {
    'Букле': 'Букле',
    'Эко-кожа': 'Эко-кожа',
    'Велюр': 'Велюто',
    'Рогожка': 'Ромео'
}


# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С ФОТОГРАФИЯМИ (ИЗ КЛИЕНТСКОГО БОТА) =====

def get_product_photos(product_code):
    """
    Возвращает отсортированный список путей к фотографиям продукта.
    Ищет фотографии в папке PHOTO_DIR/product_code/
    """
    product_photo_dir = os.path.join(PHOTO_DIR, product_code)
    photos = []
    if os.path.exists(product_photo_dir) and os.path.isdir(product_photo_dir):
        # Собираем все файлы изображений
        for file in os.listdir(product_photo_dir):
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                photos.append(os.path.join(product_photo_dir, file))
        # Сортируем по числовому значению имени файла (без расширения)
        photos.sort(key=lambda x: int(os.path.splitext(os.path.basename(x))[0]))
    return photos


def send_product_photos(chat_id, product_code):
    """
    Отправляет группу фотографий продукта в чат как медиа-группу.
    Возвращает True если отправка успешна, False если фотографий нет или ошибка.
    """
    photos = get_product_photos(product_code)
    if not photos:
        return False
    try:
        media_group = []
        for photo_path in photos:
            media_group.append(types.InputMediaPhoto(open(photo_path, 'rb')))
        bot.send_media_group(chat_id, media_group)
        return True
    except Exception as e:
        print(f"Ошибка отправки фотографий: {e}")
        return False


def get_material_photos(material_code):
    """
    Возвращает отсортированный список путей к фотографиям материала.
    Аналогично get_product_photos, но для материалов.
    """
    material_photo_dir = os.path.join(PHOTO_MATERIALS_DIR, material_code)
    photos = []
    if os.path.exists(material_photo_dir) and os.path.isdir(material_photo_dir):
        for file in os.listdir(material_photo_dir):
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                photos.append(os.path.join(material_photo_dir, file))
        photos.sort(key=lambda x: int(os.path.splitext(os.path.basename(x))[0]))
    return photos


def send_material_photos(chat_id, material_code, caption=None):
    """
    Отправляет фотографии материала с опциональной подписью к первой фотографии.
    Использует HTML-разметку для подписи.
    """
    photos = get_material_photos(material_code)
    if not photos:
        return False
    try:
        media_group = []
        for i, photo_path in enumerate(photos):
            if i == 0 and caption:
                # Только к первой фотографии добавляем подпись
                media_group.append(
                    types.InputMediaPhoto(media=open(photo_path, 'rb'), caption=caption, parse_mode='HTML'))
            else:
                media_group.append(types.InputMediaPhoto(media=open(photo_path, 'rb')))
        bot.send_media_group(chat_id, media_group)
        return True
    except Exception as e:
        print(f"Ошибка отправки фотографий материала: {e}")
        return False


# ===== ФУНКЦИИ ДЛЯ РАБОТЫ С БАЗОЙ ДАННЫХ (ИЗ КЛИЕНТСКОГО БОТА) =====

def load_products():
    """
    Загружает каталог продуктов из Excel файла.
    Возвращает список продуктов и список уникальных типов продуктов.
    """
    try:
        wb = openpyxl.load_workbook(products_log)
        sheet = wb.active
        products = []
        unique_types = set()

        # Читаем данные начиная со второй строки (первая - заголовки)
        for row in range(2, sheet.max_row + 1):
            product = {
                'code': sheet.cell(row=row, column=2).value,  # Код продукта
                'type': sheet.cell(row=row, column=3).value,  # Тип продукта
                'name': sheet.cell(row=row, column=4).value,  # Название продукта
                'link': sheet.cell(row=row, column=5).value,  # Ссылка на сайте
                'prices': {
                    'Велюто': sheet.cell(row=row, column=6).value,  # Цена для Велюто
                    'Ромео': sheet.cell(row=row, column=7).value,  # Цена для Ромео
                    'Букле': sheet.cell(row=row, column=8).value,  # Цена для Букле
                    'Эко-кожа': sheet.cell(row=row, column=9).value  # Цена для Эко-кожи
                }
            }
            products.append(product)
            unique_types.add(product['type'])

        return products, list(unique_types)
    except Exception as e:
        print(f"Ошибка загрузки продуктов: {e}")
        return [], []


def find_product_by_code(code):
    """
    Ищет продукт по коду в Excel файле.
    Возвращает словарь с данными продукта или None если не найден.
    """
    try:
        wb = openpyxl.load_workbook(products_log)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=2).value) == code:
                return {
                    'code': code,
                    'type': sheet.cell(row=row, column=3).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'link': sheet.cell(row=row, column=5).value,
                    'prices': {
                        'Велюто': sheet.cell(row=row, column=6).value,
                        'Ромео': sheet.cell(row=row, column=7).value,
                        'Букле': sheet.cell(row=row, column=8).value,
                        'Эко-кожа': sheet.cell(row=row, column=9).value
                    }
                }
        return None
    except Exception as e:
        print(f"Ошибка поиска продукта: {e}")
        return None


def load_materials():
    """
    Загружает каталог материалов из Excel файла.
    Пропускает пустые строки и нормализует код материала.
    """
    try:
        wb = openpyxl.load_workbook(materials_log)
        sheet = wb.active
        materials = []
        for row in range(2, sheet.max_row + 1):
            # Пропускаем пустые строки
            code = sheet.cell(row=row, column=2).value
            if code is None:
                continue
            materials.append({
                'code': str(code).strip(),  # Код материала
                'type': sheet.cell(row=row, column=3).value,  # Тип материала
                'color': sheet.cell(row=row, column=4).value,  # Цвет материала
                'link': sheet.cell(row=row, column=5).value  # Ссылка на сайте
            })
        return materials
    except Exception as e:
        print(f"Ошибка загрузки материалов: {e}")
        return []


def find_material_by_code(code):
    """Ищет материал по коду в загруженном списке материалов"""
    materials = load_materials()
    for material in materials:
        if material['code'] == code:
            return material
    return None


# ===== ФУНКЦИИ ДЛЯ ПОКАЗА КАРТОЧЕК ТОВАРОВ И МАТЕРИАЛОВ =====

def show_product_card(chat_id, product_code):
    """Показывает карточку товара с фотографиями"""
    product = find_product_by_code(product_code)
    if not product:
        bot.send_message(chat_id, "❌ Товар не найден в базе данных")
        return

    photos_sent = send_product_photos(chat_id, product_code)
    link_info = f"Ссылка: {product['link']}" if product['link'] else "Ссылка: отсутствует"

    # Формируем строку с ценами только для доступных материалов
    prices_text = "Цены:\n"
    for material_type, price_key in material_type_mapping.items():
        price = product['prices'].get(price_key)
        if price is not None and price != '':
            prices_text += f"  {material_type}: {price} руб\n"

    response = (
        f"<b>Карточка товара</b>\n\n"
        f"<b>{product['name']}</b>\n"
        f"Код: {product['code']}\n"
        f"Тип: {product['type']}\n"
        f"{link_info}\n"
        f"{prices_text}"
    )
    if not photos_sent:
        response = "📷 Фотографии отсутствуют\n\n" + response

    bot.send_message(chat_id, response, parse_mode='HTML')


def show_material_card(chat_id, material_code):
    """Показывает карточку материала с фотографиями"""
    material = find_material_by_code(material_code)
    if not material:
        bot.send_message(chat_id, "❌ Материал не найден в базе данных")
        return

    caption = (
        f"<b>Карточка материала</b>\n\n"
        f"<b>Материал:</b> {material['type']} {material['color']}\n"
        f"Код: {material['code']}\n"
    )
    if material['link']:
        caption += f"Ссылка: {material['link']}\n"

    # Сначала отправляем текстовую информацию
    bot.send_message(chat_id, caption, parse_mode='HTML')

    # Затем отправляем фотографии
    photos_sent = send_material_photos(chat_id, material['code'])
    if not photos_sent:
        bot.send_message(chat_id, "📷 Фотографии отсутствуют")


# Сохранение пользователя
def save_user(chat_id, username, link, fio="Без имени", role="пусто"):
    ADMIN_IDS.add(chat_id)
    ADMIN_DATA[chat_id] = {'username': username, 'link': link, 'fio': fio, 'role': role}

    # Если файла не существует, создаем его
    if not os.path.exists(ADMIN_FILE):
        open(ADMIN_FILE, 'a', encoding='utf-8').close()

    # Проверяем, не записан ли уже этот пользователь
    user_exists = False
    if os.path.exists(ADMIN_FILE):
        try:
            with open(ADMIN_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    if str(chat_id) in line:
                        user_exists = True
                        break
        except UnicodeDecodeError:
            try:
                with open(ADMIN_FILE, 'r', encoding='cp1251') as f:
                    for line in f:
                        if str(chat_id) in line:
                            user_exists = True
                            break
            except:
                with open(ADMIN_FILE, 'r', encoding='latin-1') as f:
                    for line in f:
                        if str(chat_id) in line:
                            user_exists = True
                            break

    # Если пользователь новый, добавляем его
    if not user_exists:
        with open(ADMIN_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{chat_id} - {username} - {link} - {fio} - {role}\n")

    # Перезагружаем данные аутентификации для обновления в памяти
    load_auth_data()


# Проверка авторизации
def is_authenticated(chat_id):
    return chat_id in ADMIN_IDS


# Проверка роли администратора
def is_admin(chat_id):
    return chat_id in ADMIN_DATA and ADMIN_DATA[chat_id].get('role') == 'Администратор'


# Проверка роли работника (Швея, Обтяжчик, Столяр, Упаковщик)
def is_worker(chat_id):
    worker_roles = ['Швея', 'Обтяжчик', 'Столяр', 'Упаковщик']
    return chat_id in ADMIN_DATA and ADMIN_DATA[chat_id].get('role') in worker_roles


# Обработчик аутентификации
def handle_authentication(message):
    chat_id = message.chat.id
    username = message.from_user.username or "N/A"
    link = f"https://t.me/{username}" if username != "N/A" else "N/A"

    if PASSWORD is None:
        bot.send_message(chat_id, "❌ Система аутентификации отключена")
        return False

    if message.text.strip() == PASSWORD:
        # Если пользователь новый, сохраняем с ФИО "Без имени"
        if chat_id not in ADMIN_DATA:
            save_user(chat_id, username, link, "Без имени", "пусто")
            bot.send_message(chat_id, "✅ Аутентификация успешна!")
            return True
        else:
            # Если пользователь уже существует, просто авторизуем
            bot.send_message(chat_id, "✅ Аутентификация успешна!")
            return True

    bot.send_message(chat_id, "❌ Неверный пароль")
    return False


# Функция для гарантированной инициализации данных пользователя
def ensure_user_data(chat_id):
    """Создает запись в user_data, если отсутствует"""
    if chat_id not in user_data:
        user_data[chat_id] = {
            'state': 'main_menu',
            'current_status': None,
            'current_page': 0,
            'orders': [],
            'current_order': None,
            'current_role': None,
            'delete_order_code': None,
            'workers_role': None,
            'workers_page': 0,
            'workers_list': [],
            'current_worker_id': None,
            'delete_worker_id': None,
            'current_item_code': None,
            'current_item_index': None,
            'current_workers_for_role': [],
            'current_workers_page': 0,
            'my_tasks_page': 0,
            'my_tasks': [],
            'current_task': None,
            'current_task_role': None
        }


# Функция для удаления предыдущих сообщений
def delete_previous_messages(chat_id, message_id):
    """Удаляет все сообщения из истории последних сообщений"""
    try:
        bot.send_message(chat_id, "🔄", parse_mode='HTML')
        for i in range(20):
            try:
                bot.delete_message(chat_id, message_id - i)
            except:
                pass
    except Exception as e:
        print(f"Ошибка при удалении сообщений: {e}")


# Общая клавиатура с кнопками Назад/На главную
def common_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_back = types.KeyboardButton('Назад')
    btn_main = types.KeyboardButton('На главную')
    markup.row(btn_back, btn_main)
    return markup


# Клавиатура главного меню
def main_keyboard(chat_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    # Кнопка "Мой профиль" для всех пользователей
    btn_profile = types.KeyboardButton('Мой профиль')
    markup.add(btn_profile)

    # Кнопка "Мои задачи" для всех пользователей
    btn_my_tasks = types.KeyboardButton('Мои задачи')
    markup.add(btn_my_tasks)

    # Кнопка "Заказы" только для администраторов
    if is_admin(chat_id):
        btn_orders = types.KeyboardButton('Заказы')
        btn_workers = types.KeyboardButton('Работники')
        markup.add(btn_orders, btn_workers)

    return markup


# Клавиатура меню управления паролем
def password_management_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_change = types.KeyboardButton('Изменить пароль')
    btn_delete = types.KeyboardButton('Удалить пароль')
    btn_back = types.KeyboardButton('Назад')
    btn_main = types.KeyboardButton('На главную')
    markup.row(btn_change, btn_delete)
    markup.row(btn_back, btn_main)
    return markup


# Клавиатура меню профиля
def profile_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_edit_fio = types.KeyboardButton('✏️ Изменить ФИО')
    btn_back = types.KeyboardButton('Назад')
    btn_main = types.KeyboardButton('На главную')
    markup.row(btn_edit_fio)
    markup.row(btn_back, btn_main)
    return markup


# Клавиатура для моих задач
def my_tasks_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_back = types.KeyboardButton('Назад')
    btn_main = types.KeyboardButton('На главную')
    markup.row(btn_back, btn_main)
    return markup


# Функция для загрузки уникальных статусов заказов
def get_unique_statuses():
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        statuses = set()

        for row in range(2, sheet.max_row + 1):
            status = sheet.cell(row=row, column=14).value
            if status:
                statuses.add(status)

        return list(statuses)
    except Exception as e:
        print(f"Ошибка загрузки статусов: {e}")
        return []


# Функция для загрузки заказов по статусу
def get_orders_by_status(status):
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        orders = {}

        for row in range(2, sheet.max_row + 1):
            order_status = sheet.cell(row=row, column=14).value
            if status == "Все заказы" or str(order_status) == status:
                order_code = sheet.cell(row=row, column=3).value
                task_number = sheet.cell(row=row, column=2).value

                if order_code not in orders:
                    orders[order_code] = {
                        'order_code': order_code,
                        'task_number': task_number,
                        'datetime': sheet.cell(row=row, column=13).value,
                        'status': order_status,
                        'fio': sheet.cell(row=row, column=4).value,
                        'phone': sheet.cell(row=row, column=5).value,
                        'address': sheet.cell(row=row, column=6).value,
                        'chat_id': sheet.cell(row=row, column=7).value,
                        'telegram_contact': sheet.cell(row=row, column=8).value,
                        'comment': sheet.cell(row=row, column=15).value,
                        'items': []
                    }

                # Получаем данные о работниках для каждой позиции
                seamstress_code = sheet.cell(row=row, column=17).value
                seamstress_fio = sheet.cell(row=row, column=18).value
                seamstress_readiness = sheet.cell(row=row, column=19).value

                carpenter_code = sheet.cell(row=row, column=20).value
                carpenter_fio = sheet.cell(row=row, column=21).value
                carpenter_readiness = sheet.cell(row=row, column=22).value

                upholsterer_code = sheet.cell(row=row, column=23).value
                upholsterer_fio = sheet.cell(row=row, column=24).value
                upholsterer_readiness = sheet.cell(row=row, column=25).value

                packer_code = sheet.cell(row=row, column=26).value
                packer_fio = sheet.cell(row=row, column=27).value
                packer_readiness = sheet.cell(row=row, column=28).value

                orders[order_code]['items'].append({
                    'code': sheet.cell(row=row, column=9).value,
                    'material': sheet.cell(row=row, column=10).value,
                    'color': sheet.cell(row=row, column=11).value,
                    'price': sheet.cell(row=row, column=12).value,
                    'seamstress_code': seamstress_code,
                    'seamstress_fio': seamstress_fio,
                    'seamstress_readiness': seamstress_readiness,
                    'carpenter_code': carpenter_code,
                    'carpenter_fio': carpenter_fio,
                    'carpenter_readiness': carpenter_readiness,
                    'upholsterer_code': upholsterer_code,
                    'upholsterer_fio': upholsterer_fio,
                    'upholsterer_readiness': upholsterer_readiness,
                    'packer_code': packer_code,
                    'packer_fio': packer_fio,
                    'packer_readiness': packer_readiness
                })
        return orders
    except Exception as e:
        print(f"Ошибка загрузки заказов: {e}")
        return {}


# Функция для получения задач пользователя
def get_my_tasks(chat_id):
    """Получает все задачи пользователя из Orders_log.xlsx"""
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        tasks = []

        for row in range(2, sheet.max_row + 1):
            # Проверяем все роли в строке
            seamstress_code = sheet.cell(row=row, column=17).value
            carpenter_code = sheet.cell(row=row, column=20).value
            upholsterer_code = sheet.cell(row=row, column=23).value
            packer_code = sheet.cell(row=row, column=26).value

            # Определяем роль пользователя в этой задаче
            user_role = None
            readiness_value = None
            readiness_column = None

            if seamstress_code and str(seamstress_code) == str(chat_id):
                user_role = "Швея"
                readiness_value = sheet.cell(row=row, column=19).value
                readiness_column = 19
            elif carpenter_code and str(carpenter_code) == str(chat_id):
                user_role = "Столяр"
                readiness_value = sheet.cell(row=row, column=22).value
                readiness_column = 22
            elif upholsterer_code and str(upholsterer_code) == str(chat_id):
                user_role = "Обтяжчик"
                readiness_value = sheet.cell(row=row, column=25).value
                readiness_column = 25
            elif packer_code and str(packer_code) == str(chat_id):
                user_role = "Упаковщик"
                readiness_value = sheet.cell(row=row, column=28).value
                readiness_column = 28

            if user_role:
                task = {
                    'row': row,
                    'order_code': sheet.cell(row=row, column=3).value,
                    'task_number': sheet.cell(row=row, column=2).value,
                    'item_code': sheet.cell(row=row, column=9).value,
                    'material': sheet.cell(row=row, column=10).value,
                    'color': sheet.cell(row=row, column=11).value,
                    'price': sheet.cell(row=row, column=12).value,
                    'datetime': sheet.cell(row=row, column=13).value,
                    'status': sheet.cell(row=row, column=14).value,
                    'user_role': user_role,
                    'readiness': readiness_value if readiness_value else "нет",
                    'readiness_column': readiness_column,
                    'fio': sheet.cell(row=row, column=4).value,
                    'phone': sheet.cell(row=row, column=5).value,
                    'address': sheet.cell(row=row, column=6).value
                }
                tasks.append(task)

        return tasks
    except Exception as e:
        print(f"Ошибка загрузки задач пользователя: {e}")
        return []


# Функция для обновления готовности задачи
def update_task_readiness(row, readiness_column, readiness):
    """Обновляет готовность задачи в файле Orders_log.xlsx"""
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active

        sheet.cell(row=row, column=readiness_column, value=readiness)
        wb.save(orders_log)
        return True
    except Exception as e:
        print(f"Ошибка обновления готовности задачи: {e}")
        return False


# Функция для форматирования информации о работнике в позиции
def format_worker_display_info(fio, code, readiness):
    """Форматирует информацию о работнике в формате 'ФИО (готовность)'"""
    if not fio or fio == "нет":
        return "не назначен(а)"

    readiness_display = readiness if readiness and readiness != "нет" else "не указана"

    if not code or code == "нет":
        return f"⚠️{fio} ({readiness_display})"
    else:
        return f"{fio} ({readiness_display})"


# Функция для обновления позиции заказа
def update_order_item(order_code, item_code, field, value):
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        updated = False

        # Определение индекса колонки по названию поля
        column_index = {
            'seamstress_fio': 18,
            'carpenter_fio': 21,
            'upholsterer_fio': 24,
            'packer_fio': 27
        }.get(field)

        if not column_index:
            return False

        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=3).value) == order_code and str(
                    sheet.cell(row=row, column=9).value) == item_code:
                sheet.cell(row=row, column=column_index, value=value)
                updated = True

        if updated:
            wb.save(orders_log)
            return True
        return False
    except Exception as e:
        print(f"Ошибка обновления позиции заказа: {e}")
        return False


# Функция для обновления статуса заказа
def update_order_status(order_code, status):
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        updated = False

        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=3).value) == order_code:
                sheet.cell(row=row, column=14, value=status)
                updated = True

        if updated:
            wb.save(orders_log)
            return True
        return False
    except Exception as e:
        print(f"Ошибка обновления статуса заказа: {e}")
        return False


# Функция для удаления заказа
def delete_order(order_code):
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        rows_to_delete = []

        # Находим строки для удаления
        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=3).value) == order_code:
                rows_to_delete.append(row)

        # Удаляем строки в обратном порядке
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        wb.save(orders_log)
        return True
    except Exception as e:
        print(f"Ошибка удаления заказа: {e}")
        return False


# Функция для получения уникальных ролей работников
def get_unique_roles():
    roles = set()
    for worker_data in ADMIN_DATA.values():
        if worker_data.get('role'):
            roles.add(worker_data['role'])
    return list(roles)


# Функция для получения работников по роли
def get_workers_by_role(role):
    workers = {}
    for chat_id, worker_data in ADMIN_DATA.items():
        if role == "Все работники" or worker_data.get('role') == role:
            workers[chat_id] = worker_data
    return workers


# Функция для сохранения работников в файл
def save_workers():
    try:
        with open(ADMIN_FILE, 'w', encoding='utf-8') as f:
            for chat_id, data in ADMIN_DATA.items():
                f.write(f"{chat_id} - {data['username']} - {data['link']} - {data['fio']} - {data['role']}\n")
        return True
    except Exception as e:
        print(f"Ошибка сохранения работников: {e}")
        return False


# Функция для обновления данных работника
def update_worker(worker_id, field, value):
    if worker_id not in ADMIN_DATA:
        return False

    if field == 'fio':
        old_fio = ADMIN_DATA[worker_id]['fio']
        ADMIN_DATA[worker_id]['fio'] = value

        # Обновляем ФИО во всех заказах
        update_fio_in_orders(old_fio, value)
    elif field == 'role':
        ADMIN_DATA[worker_id]['role'] = value

    result = save_workers()
    # Перезагружаем данные аутентификации для обновления в памяти
    load_auth_data()
    return result


# Функция для обновления ФИО во всех заказах
def update_fio_in_orders(old_fio, new_fio):
    """Обновляет ФИО работника во всех заказах где он упоминается"""
    try:
        wb = openpyxl.load_workbook(orders_log)
        sheet = wb.active
        updated = False

        # Поля где может быть указано ФИО работника
        fio_columns = [18, 21, 24, 27]  # Швея, Столяр, Обтяжчик, Упаковщик

        for row in range(2, sheet.max_row + 1):
            for col in fio_columns:
                if sheet.cell(row=row, column=col).value == old_fio:
                    sheet.cell(row=row, column=col, value=new_fio)
                    updated = True

        if updated:
            wb.save(orders_log)
            print(f"✅ ФИО обновлено с '{old_fio}' на '{new_fio}' в {updated} позициях")
        return updated
    except Exception as e:
        print(f"Ошибка обновления ФИО в заказах: {e}")
        return False


# Функция для удаления работника
def delete_worker(worker_id):
    if worker_id not in ADMIN_DATA:
        return False

    # Удаляем работника из данных
    del ADMIN_DATA[worker_id]
    ADMIN_IDS.discard(worker_id)

    result = save_workers()
    # Перезагружаем данные аутентификации для обновления в памяти
    load_auth_data()
    return result


# Функция для получения работников по роли для назначения (ИЗМЕНЕНА ДЛЯ НОВОГО ПОРЯДКА)
def get_workers_for_role(role):
    """Получает работников для назначения на роль в порядке: Специалист -> Без роли -> Администратор"""
    # Сопоставление английских названий ролей с русскими
    role_mapping = {
        'seamstress': 'Швея',
        'carpenter': 'Столяр',
        'upholsterer': 'Обтяжчик',
        'packer': 'Упаковщик'
    }

    russian_role = role_mapping.get(role, role)

    workers_specialists = []  # Специалисты с нужной ролью
    workers_no_role = []  # Работники без роли
    workers_admins = []  # Администраторы

    for chat_id, worker_data in ADMIN_DATA.items():
        if worker_data.get('role') == russian_role:
            workers_specialists.append((chat_id, worker_data))
        elif worker_data.get('role') == 'пусто':
            workers_no_role.append((chat_id, worker_data))
        elif worker_data.get('role') == 'Администратор':
            workers_admins.append((chat_id, worker_data))

    # Объединяем в правильном порядке
    workers = workers_specialists + workers_no_role + workers_admins
    return workers


# Клавиатура со статусами заказов
def statuses_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # Добавляем кнопку "Все заказы" в начало
    markup.add(types.KeyboardButton('Все заказы'))

    statuses = get_unique_statuses()
    for status in statuses:
        markup.add(types.KeyboardButton(status))

    markup.row(types.KeyboardButton('Назад'), types.KeyboardButton('На главную'))
    return markup


# Клавиатура с ролями работников
def workers_roles_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # Добавляем кнопку "Все работники" в начало
    markup.add(types.KeyboardButton('Все работники'))

    roles = get_unique_roles()
    for role in roles:
        markup.add(types.KeyboardButton(role))

    markup.row(types.KeyboardButton('Назад'), types.KeyboardButton('На главную'))
    return markup


# Клавиатура для пагинации заказов
def orders_pagination_keyboard(page, total_pages):
    markup = types.InlineKeyboardMarkup()
    row_buttons = []

    if page > 0:
        row_buttons.append(types.InlineKeyboardButton('⬅️ Предыдущая', callback_data=f'prev_{page}'))

    if page < total_pages - 1:
        row_buttons.append(types.InlineKeyboardButton('Следующая ➡️', callback_data=f'next_{page}'))

    if row_buttons:
        markup.row(*row_buttons)

    markup.row(
        types.InlineKeyboardButton('Назад к статусам', callback_data='back_to_statuses'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура для пагинации работников
def workers_pagination_keyboard(page, total_pages):
    markup = types.InlineKeyboardMarkup()
    row_buttons = []

    if page > 0:
        row_buttons.append(types.InlineKeyboardButton('⬅️ Предыдущая', callback_data=f'workers_prev_{page}'))

    if page < total_pages - 1:
        row_buttons.append(types.InlineKeyboardButton('Следующая ➡️', callback_data=f'workers_next_{page}'))

    if row_buttons:
        markup.row(*row_buttons)

    markup.row(
        types.InlineKeyboardButton('Назад к ролям', callback_data='back_to_workers_roles'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура действий с заказом
def order_actions_keyboard(order_code):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton('✏️ Изменить статус', callback_data=f'change_status_{order_code}'))
    markup.add(types.InlineKeyboardButton('🗑️ Удалить заказ', callback_data=f'delete_{order_code}'))
    return markup


# Клавиатура действий с позицией заказа (ДОБАВЛЕНА КНОПКА "ПОДРОБНОСТИ")
def item_actions_keyboard(order_code, item_code, item_index, material_code=None):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton('✏️ Редактировать позицию',
                                          callback_data=f'edit_item_{order_code}_{item_code}_{item_index}'))
    # Добавляем кнопку "Подробности" если есть коды товара и материала
    if item_code and material_code:
        markup.add(types.InlineKeyboardButton('📋 Подробности',
                                              callback_data=f'details_{item_code}_{material_code}'))
    return markup


# Клавиатура действий с работником
def worker_actions_keyboard(worker_id, current_user_id):
    markup = types.InlineKeyboardMarkup()

    # Кнопки доступны для всех работников
    markup.add(types.InlineKeyboardButton('✏️ Изменить имя', callback_data=f'worker_edit_name_{worker_id}'))

    # Кнопка "Изменить роль" и "Удалить" недоступны если это текущий пользователь
    if worker_id != current_user_id:
        markup.add(types.InlineKeyboardButton('🔄 Изменить роль', callback_data=f'worker_edit_role_{worker_id}'))
        markup.add(types.InlineKeyboardButton('🗑️ Удалить', callback_data=f'worker_delete_{worker_id}'))

    markup.add(types.InlineKeyboardButton('📨 Написать', callback_data=f'worker_message_{worker_id}'))

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data=f'back_to_workers_list'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура для редактирования позиции заказа
def edit_item_keyboard(order_code, item_code, item_index):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton('Назначить швею',
                                          callback_data=f'assign_item_seamstress_{order_code}_{item_code}_{item_index}'))
    markup.add(types.InlineKeyboardButton('Назначить столяра',
                                          callback_data=f'assign_item_carpenter_{order_code}_{item_code}_{item_index}'))
    markup.add(types.InlineKeyboardButton('Назначить обтяжчика',
                                          callback_data=f'assign_item_upholsterer_{order_code}_{item_code}_{item_index}'))
    markup.add(types.InlineKeyboardButton('Назначить упаковщика',
                                          callback_data=f'assign_item_packer_{order_code}_{item_code}_{item_index}'))
    markup.row(
        types.InlineKeyboardButton('Назад', callback_data=f'back_to_order_{order_code}'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура со статусами для изменения
def status_change_keyboard(order_code):
    markup = types.InlineKeyboardMarkup()

    for status in FIXED_STATUSES:
        markup.add(types.InlineKeyboardButton(status, callback_data=f'newstatus_{order_code}_{status}'))

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data=f'back_to_order_{order_code}'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура с ролями для изменения
def role_change_keyboard(worker_id):
    markup = types.InlineKeyboardMarkup()

    roles = ["Администратор", "Менеджер", "Швея", "Столяр", "Обтяжчик", "Упаковщик", "пусто"]

    for role in roles:
        markup.add(types.InlineKeyboardButton(role, callback_data=f'newrole_{worker_id}_{role}'))

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data=f'back_to_worker_{worker_id}'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура для выбора работника для назначения
def workers_for_role_keyboard(workers, page, total_pages, role, order_code, item_code, item_index):
    markup = types.InlineKeyboardMarkup()

    # Добавляем кнопки для каждого работника
    for worker_id, worker_data in workers:
        markup.add(types.InlineKeyboardButton(
            f"{worker_data['fio']} ({worker_data['role']})",
            callback_data=f'select_worker_{worker_id}_{role}_{order_code}_{item_code}_{item_index}'
        ))

    # Кнопки пагинации
    row_buttons = []
    if page > 0:
        row_buttons.append(types.InlineKeyboardButton('⬅️ Предыдущая',
                                                      callback_data=f'workers_role_prev_{page}_{role}_{order_code}_{item_code}_{item_index}'))

    if page < total_pages - 1:
        row_buttons.append(types.InlineKeyboardButton('Следующая ➡️',
                                                      callback_data=f'workers_role_next_{page}_{role}_{order_code}_{item_code}_{item_index}'))

    if row_buttons:
        markup.row(*row_buttons)

    # Кнопка для ввода ФИО вручную
    markup.add(types.InlineKeyboardButton(
        '✏️ Ввести ФИО вручную',
        callback_data=f'manual_input_{role}_{order_code}_{item_code}_{item_index}'
    ))

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data=f'back_to_edit_item_{order_code}_{item_code}_{item_index}'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )

    return markup


# Клавиатура для пагинации моих задач
def my_tasks_pagination_keyboard(page, total_pages):
    markup = types.InlineKeyboardMarkup()
    row_buttons = []

    if page > 0:
        row_buttons.append(types.InlineKeyboardButton('⬅️ Предыдущая', callback_data=f'my_tasks_prev_{page}'))

    if page < total_pages - 1:
        row_buttons.append(types.InlineKeyboardButton('Следующая ➡️', callback_data=f'my_tasks_next_{page}'))

    if row_buttons:
        markup.row(*row_buttons)

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data='back_to_my_tasks'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Клавиатура для изменения готовности задачи (ДОБАВЛЕНА КНОПКА "ПОДРОБНОСТИ")
def task_readiness_keyboard(task_index, item_code=None, material_code=None):
    markup = types.InlineKeyboardMarkup()

    for status in READINESS_STATUSES:
        markup.add(types.InlineKeyboardButton(
            status,
            callback_data=f'change_readiness_{task_index}_{status}'
        ))

    # Добавляем кнопку "Подробности" если есть коды товара и материала
    if item_code and material_code:
        markup.add(types.InlineKeyboardButton('📋 Подробности',
                                              callback_data=f'details_{item_code}_{material_code}'))

    markup.row(
        types.InlineKeyboardButton('Назад', callback_data='back_to_my_tasks_list'),
        types.InlineKeyboardButton('На главную', callback_data='to_main_menu')
    )
    return markup


# Форматирование основной информации о заказе (без позиций и без работников)
def format_order_main_info(order):
    total_price = sum(item['price'] for item in order['items']) if order['items'] else 0

    order_info = (
        f"📦 <b>Заказ:</b> <code>{order['order_code']}</code>\n"
        f"📋 <b>Номер задачи:</b> {order['task_number']}\n"
        f"📅 <b>Дата создания:</b> {order['datetime']}\n"
        f"🔄 <b>Статус:</b> {order['status']}\n\n"
        f"👤 <b>Клиент:</b> {order['fio']}\n"
        f"📱 <b>Телефон:</b> {order['phone']}\n"
        f"🏠 <b>Адрес:</b> {order['address']}\n"
        f"💬 <b>Комментарий:</b> {order['comment'] or 'отсутствует'}\n\n"
        f"💰 <b>Итого:</b> {total_price} руб\n\n"
        f"<b>Состав заказа:</b>"
    )
    return order_info


# Форматирование информации о позиции заказа (с работниками) - ДОБАВЛЕНА КНОПКА "ПОДРОБНОСТИ"
def format_order_item(order_code, item, index):
    # Форматируем информацию о работниках для позиции
    seamstress_display = format_worker_display_info(
        item['seamstress_fio'],
        item['seamstress_code'],
        item['seamstress_readiness']
    )
    carpenter_display = format_worker_display_info(
        item['carpenter_fio'],
        item['carpenter_code'],
        item['carpenter_readiness']
    )
    upholsterer_display = format_worker_display_info(
        item['upholsterer_fio'],
        item['upholsterer_code'],
        item['upholsterer_readiness']
    )
    packer_display = format_worker_display_info(
        item['packer_fio'],
        item['packer_code'],
        item['packer_readiness']
    )

    item_info = (
        f"📦 <b>Заказ:</b> <code>{order_code}</code>\n"
        f"📋 <b>Позиция {index}:</b>\n"
        f"   <b>Код:</b> {item['code']}\n"
        f"   <b>Материал:</b> {item['material']}\n"
        f"   <b>Цвет:</b> {item['color']}\n"
        f"   <b>Цена:</b> {item['price']} руб\n"
        f"   👷 <b>Швея:</b> {seamstress_display}\n"
        f"   🔨 <b>Столяр:</b> {carpenter_display}\n"
        f"   🧵 <b>Обтяжчик:</b> {upholsterer_display}\n"
        f"   📦 <b>Упаковщик:</b> {packer_display}"
    )
    return item_info


# Форматирование информации о работнике
def format_worker_info(worker_id, worker_data):
    worker_info = (
        f"👤 <b>ФИО:</b> {worker_data['fio']}\n"
        f"📋 <b>Роль:</b> {worker_data['role']}\n"
        f"🔗 <b>Username:</b> {worker_data['username']}\n"
        f"🌐 <b>Ссылка:</b> {worker_data['link']}\n"
        f"🆔 <b>ID:</b> <code>{worker_id}</code>"
    )
    return worker_info


# Форматирование информации о профиле пользователя
def format_profile_info(worker_data):
    profile_info = (
        f"👤 <b>Ваш профиль</b>\n\n"
        f"📝 <b>ФИО:</b> {worker_data['fio']}\n"
        f"🎯 <b>Должность:</b> {worker_data['role']}\n"
        f"🔗 <b>Username:</b> {worker_data['username']}\n"
        f"🌐 <b>Ссылка:</b> {worker_data['link']}\n"
        f"🆔 <b>Ваш ID:</b> <code>{worker_data.get('chat_id', 'N/A')}</code>"
    )
    return profile_info


# Форматирование информации о задаче (ДОБАВЛЕНА КНОПКА "ПОДРОБНОСТИ")
def format_task_info(task, index):
    task_info = (
        f"📋 <b>Задача {index}:</b>\n"
        f"🆔 <b>Код заказа:</b> <code>{task['order_code']}</code>\n"
        f"🔢 <b>Номер задачи:</b> {task['task_number']}\n"
        f"📦 <b>Код позиции:</b> {task['item_code']}\n"
        f"🧵 <b>Материал:</b> {task['material']}\n"
        f"🎨 <b>Цвет:</b> {task['color']}\n"
        f"💰 <b>Цена:</b> {task['price']} руб\n"
        f"👤 <b>Ваша роль:</b> {task['user_role']}\n"
        f"📊 <b>Готовность:</b> {task['readiness']}\n"
        f"📅 <b>Дата создания:</b> {task['datetime']}\n"
        f"🔄 <b>Статус заказа:</b> {task['status']}\n"
        f"👨 <b>Клиент:</b> {task['fio']}\n"
        f"📱 <b>Телефон:</b> {task['phone']}\n"
        f"🏠 <b>Адрес:</b> {task['address']}"
    )
    return task_info


# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    chat_id = message.chat.id

    # Перезагружаем данные аутентификации для обновления в памяти
    load_auth_data()

    if not is_authenticated(chat_id):
        bot.send_message(chat_id, "🔒 Для доступа введите пароль:")
        return

    user_data[chat_id] = {
        'state': 'main_menu',
        'current_status': None,
        'current_page': 0,
        'orders': [],
        'current_order': None,
        'current_role': None,
        'delete_order_code': None,
        'workers_role': None,
        'workers_page': 0,
        'workers_list': [],
        'current_worker_id': None,
        'delete_worker_id': None,
        'current_item_code': None,
        'current_item_index': None,
        'current_workers_for_role': [],
        'current_workers_page': 0,
        'my_tasks_page': 0,
        'my_tasks': [],
        'current_task': None,
        'current_task_role': None
    }
    delete_previous_messages(chat_id, message.message_id)
    bot.send_message(
        chat_id,
        "👋 Добро пожаловать в систему управления заказами!",
        reply_markup=main_keyboard(chat_id),
        parse_mode='HTML'
    )


# Обработчик команды /password
@bot.message_handler(commands=['password'])
def password_command(message):
    chat_id = message.chat.id

    if not is_authenticated(chat_id):
        bot.send_message(chat_id, "🔒 Для доступа введите пароль:")
        return

    if not is_admin(chat_id):
        bot.send_message(chat_id, "❌ У вас недостаточно прав для выполнения этой команды.")
        return

    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'password_management'
    delete_previous_messages(chat_id, message.message_id)

    password_status = "установлен" if os.path.exists(PASSWORD_FILE) else "не установлен"

    bot.send_message(
        chat_id,
        f"🔐 <b>Управление паролем</b>\n\n"
        f"Текущий статус: {password_status}\n\n"
        f"Выберите действие:",
        reply_markup=password_management_keyboard(),
        parse_mode='HTML'
    )


# Обработка пароля - перехватывает ВСЕ сообщения неавторизованных пользователей
@bot.message_handler(func=lambda message: not is_authenticated(message.chat.id))
def auth_handler(message):
    chat_id = message.chat.id
    if handle_authentication(message):
        # После успешной аутентификации инициализируем сессию
        user_data[chat_id] = {
            'state': 'main_menu',
            'current_status': None,
            'current_page': 0,
            'orders': [],
            'current_order': None,
            'current_role': None,
            'delete_order_code': None,
            'workers_role': None,
            'workers_page': 0,
            'workers_list': [],
            'current_worker_id': None,
            'delete_worker_id': None,
            'current_item_code': None,
            'current_item_index': None,
            'current_workers_for_role': [],
            'current_workers_page': 0,
            'my_tasks_page': 0,
            'my_tasks': [],
            'current_task': None,
            'current_task_role': None
        }
        bot.send_message(
            chat_id,
            "👋 Добро пожаловать в систему управления заказами!",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )
    else:
        bot.send_message(chat_id, "🔒 Для доступа введите пароль:")


# Обработчик кнопки "Мой профиль"
@bot.message_handler(func=lambda message: message.text == 'Мой профиль' and is_authenticated(message.chat.id))
def show_profile(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'viewing_profile'
    delete_previous_messages(chat_id, message.message_id)

    if chat_id in ADMIN_DATA:
        worker_data = ADMIN_DATA[chat_id].copy()
        worker_data['chat_id'] = chat_id
        profile_info = format_profile_info(worker_data)

        bot.send_message(
            chat_id,
            profile_info,
            reply_markup=profile_keyboard(),
            parse_mode='HTML'
        )
    else:
        bot.send_message(
            chat_id,
            "❌ Ошибка: данные вашего профиля не найдены",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )


# Обработчик кнопки "Мои задачи"
@bot.message_handler(func=lambda message: message.text == 'Мои задачи' and is_authenticated(message.chat.id))
def show_my_tasks(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'viewing_my_tasks'
    user_data[chat_id]['my_tasks_page'] = 0
    delete_previous_messages(chat_id, message.message_id)

    # Загружаем задачи пользователя
    tasks = get_my_tasks(chat_id)
    user_data[chat_id]['my_tasks'] = tasks

    if not tasks:
        bot.send_message(
            chat_id,
            "📭 У вас нет назначенных задач.",
            reply_markup=my_tasks_keyboard(),
            parse_mode='HTML'
        )
        return

    # Показываем первую страницу задач
    show_my_tasks_page(chat_id, 0)


# Функция для отображения страницы моих задач
def show_my_tasks_page(chat_id, page):
    ensure_user_data(chat_id)
    tasks = user_data[chat_id]['my_tasks']

    total_tasks = len(tasks)
    total_pages = max(1, (total_tasks + 4) // 5)  # 5 задач на страницу

    # Корректировка номера страницы
    if page < 0:
        page = 0
    elif page >= total_pages and total_pages > 0:
        page = total_pages - 1

    user_data[chat_id]['my_tasks_page'] = page
    start_idx = page * 5
    end_idx = min(start_idx + 5, total_tasks)
    page_tasks = tasks[start_idx:end_idx]

    # Отправляем информацию о странице
    bot.send_message(
        chat_id,
        f"📋 <b>Мои задачи</b>\n"
        f"📄 Страница {page + 1}/{total_pages}\n"
        f"✅ Задачи: {start_idx + 1}-{end_idx} из {total_tasks}",
        parse_mode='HTML',
        reply_markup=my_tasks_keyboard()
    )

    # Отправляем каждую задачу на странице
    for idx, task in enumerate(page_tasks, start_idx + 1):
        task_info = format_task_info(task, idx)

        # Создаем клавиатуру для изменения готовности с кнопкой "Подробности"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton(
            '✏️ Изменить готовность',
            callback_data=f'change_task_readiness_{idx - 1}'
        ))

        # Добавляем кнопку "Подробности" с передачей кодов товара и материала
        if task['item_code'] and task['material']:
            # Извлекаем код материала из поля material (формат: "Код: XXXXX")
            material_code = None
            if 'Код:' in task['material']:
                material_code = task['material'].split('Код:')[-1].strip()
            else:
                material_code = task['material']  # Если код материала хранится напрямую

            markup.add(types.InlineKeyboardButton(
                '📋 Подробности',
                callback_data=f'details_{task["item_code"]}_{material_code}'
            ))

        bot.send_message(
            chat_id,
            task_info,
            reply_markup=markup,
            parse_mode='HTML'
        )

    # Отправляем клавиатуру пагинации
    if total_pages > 1:
        pagination_keyboard = my_tasks_pagination_keyboard(page, total_pages)
        bot.send_message(
            chat_id,
            "⬅️➡️ Переключение страниц:",
            reply_markup=pagination_keyboard,
            parse_mode='HTML'
        )


# Обработчик кнопки "Заказы" (только для администраторов)
@bot.message_handler(
    func=lambda message: message.text == 'Заказы' and is_authenticated(message.chat.id) and is_admin(message.chat.id))
def show_statuses(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id]['state'] = 'viewing_statuses'
    delete_previous_messages(chat_id, message.message_id)
    bot.send_message(
        chat_id,
        "📊 Выберите статус заказов:",
        reply_markup=statuses_keyboard(),
        parse_mode='HTML'
    )


# Обработчик кнопки "Работники" (только для администраторов)
@bot.message_handler(
    func=lambda message: message.text == 'Работники' and is_authenticated(message.chat.id) and is_admin(
        message.chat.id))
def show_workers_roles(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id]['state'] = 'viewing_workers_roles'
    delete_previous_messages(chat_id, message.message_id)
    bot.send_message(
        chat_id,
        "👥 Выберите роль работников:",
        reply_markup=workers_roles_keyboard(),
        parse_mode='HTML'
    )


# Обработчик общих кнопок Назад/На главную
@bot.message_handler(func=lambda message: message.text in ['Назад', 'На главную'] and is_authenticated(message.chat.id))
def handle_common_buttons(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    state = user_data[chat_id]['state']
    text = message.text

    if text == 'На главную':
        # Переход в главное меню
        user_data[chat_id] = {
            'state': 'main_menu',
            'current_status': None,
            'current_page': 0,
            'orders': [],
            'current_order': None,
            'current_role': None,
            'delete_order_code': None,
            'workers_role': None,
            'workers_page': 0,
            'workers_list': [],
            'current_worker_id': None,
            'delete_worker_id': None,
            'current_item_code': None,
            'current_item_index': None,
            'current_workers_for_role': [],
            'current_workers_page': 0,
            'my_tasks_page': 0,
            'my_tasks': [],
            'current_task': None,
            'current_task_role': None
        }
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )
        return

    # Обработка кнопки "Назад"
    if state == 'viewing_profile':
        # Назад из профиля -> главное меню
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )

    elif state == 'awaiting_new_fio':
        # Назад из ввода нового ФИО -> профиль
        user_data[chat_id]['state'] = 'viewing_profile'
        delete_previous_messages(chat_id, message.message_id)

        if chat_id in ADMIN_DATA:
            worker_data = ADMIN_DATA[chat_id].copy()
            worker_data['chat_id'] = chat_id
            profile_info = format_profile_info(worker_data)

            bot.send_message(
                chat_id,
                profile_info,
                reply_markup=profile_keyboard(),
                parse_mode='HTML'
            )
        else:
            bot.send_message(
                chat_id,
                "❌ Ошибка: данные вашего профиля не найдены",
                reply_markup=main_keyboard(chat_id),
                parse_mode='HTML'
            )

    elif state == 'viewing_statuses':
        # Назад из статусов -> главное меню
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )

    elif state == 'viewing_orders':
        # Назад из списка заказов -> к статусам
        user_data[chat_id]['state'] = 'viewing_statuses'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "📊 Выберите статус заказов:",
            reply_markup=statuses_keyboard(),
            parse_mode='HTML'
        )

    elif state == 'viewing_workers_roles':
        # Назад из ролей работников -> главное меню
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )

    elif state == 'viewing_workers':
        # Назад из списка работников -> к ролям
        user_data[chat_id]['state'] = 'viewing_workers_roles'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "👥 Выберите роль работников:",
            reply_markup=workers_roles_keyboard(),
            parse_mode='HTML'
        )

    elif state == 'editing_item':
        # Назад из редактирования позиции -> к списку заказов
        user_data[chat_id]['state'] = 'viewing_orders'
        delete_previous_messages(chat_id, message.message_id)
        show_orders_page(chat_id, user_data[chat_id]['current_page'])

    elif state == 'editing_worker':
        # Назад из редактирования работника -> к списку работников
        user_data[chat_id]['state'] = 'viewing_workers'
        delete_previous_messages(chat_id, message.message_id)
        show_workers_page(chat_id, user_data[chat_id]['workers_page'])

    elif state == 'changing_status':
        # Назад из выбора статуса -> к списку заказов
        user_data[chat_id]['state'] = 'viewing_orders'
        delete_previous_messages(chat_id, message.message_id)
        show_orders_page(chat_id, user_data[chat_id]['current_page'])

    elif state == 'changing_worker_role':
        # Назад из выбора роли -> к редактированию работника
        worker_id = user_data[chat_id]['current_worker_id']
        user_data[chat_id]['state'] = 'editing_worker'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
            reply_markup=worker_actions_keyboard(worker_id, chat_id),
            parse_mode='HTML'
        )

    elif state == 'awaiting_item_worker_input':
        # Назад из ввода работника для позиции -> к редактированию позиции
        order_code = user_data[chat_id]['current_order']
        item_code = user_data[chat_id]['current_item_code']
        item_index = user_data[chat_id]['current_item_index']
        user_data[chat_id]['state'] = 'editing_item'
        delete_previous_messages(chat_id, message.message_id)

        # Получаем актуальные данные о позиции
        status = user_data[chat_id]['current_status']
        orders = get_orders_by_status(status)
        if order_code in orders:
            order = orders[order_code]
            item = None
            for it in order['items']:
                if it['code'] == item_code:
                    item = it
                    break

            if item:
                item_info = format_order_item(order_code, item, item_index)
                bot.send_message(
                    chat_id,
                    f"✏️ Редактирование позиции:\n{item_info}",
                    reply_markup=edit_item_keyboard(order_code, item_code, item_index),
                    parse_mode='HTML'
                )

    elif state == 'awaiting_worker_name':
        # Назад из ввода имени работника -> к редактированию работника
        worker_id = user_data[chat_id]['current_worker_id']
        user_data[chat_id]['state'] = 'editing_worker'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
            reply_markup=worker_actions_keyboard(worker_id, chat_id),
            parse_mode='HTML'
        )

    elif state == 'confirming_delete':
        # Назад из подтверждения удаления -> к списку заказов
        user_data[chat_id]['state'] = 'viewing_orders'
        delete_previous_messages(chat_id, message.message_id)
        show_orders_page(chat_id, user_data[chat_id]['current_page'])

    elif state == 'confirming_worker_delete':
        # Назад из подтверждения удаления работника -> к редактированию работника
        worker_id = user_data[chat_id]['current_worker_id']
        user_data[chat_id]['state'] = 'editing_worker'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
            reply_markup=worker_actions_keyboard(worker_id, chat_id),
            parse_mode='HTML'
        )

    elif state == 'password_management':
        # Назад из управления паролем -> главное меню
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )

    elif state == 'awaiting_new_password':
        # Назад из ввода нового пароля -> управление паролем
        user_data[chat_id]['state'] = 'password_management'
        delete_previous_messages(chat_id, message.message_id)
        password_status = "установлен" if os.path.exists(PASSWORD_FILE) else "не установлен"
        bot.send_message(
            chat_id,
            f"🔐 <b>Управление паролем</b>\n\n"
            f"Текущий статус: {password_status}\n\n"
            f"Выберите действие:",
            reply_markup=password_management_keyboard(),
            parse_mode='HTML'
        )

    elif state == 'confirming_password_delete':
        # Назад из подтверждения удаления пароля -> управление паролем
        user_data[chat_id]['state'] = 'password_management'
        delete_previous_messages(chat_id, message.message_id)
        password_status = "установлен" if os.path.exists(PASSWORD_FILE) else "не установлен"
        bot.send_message(
            chat_id,
            f"🔐 <b>Управление паролем</b>\n\n"
            f"Текущий статус: {password_status}\n\n"
            f"Выберите действие:",
            reply_markup=password_management_keyboard(),
            parse_mode='HTML'
        )

    elif state == 'selecting_worker_for_role':
        # Назад из выбора работника -> к редактированию позиции
        order_code = user_data[chat_id]['current_order']
        item_code = user_data[chat_id]['current_item_code']
        item_index = user_data[chat_id]['current_item_index']
        user_data[chat_id]['state'] = 'editing_item'
        delete_previous_messages(chat_id, message.message_id)

        # Получаем актуальные данные о позиции
        status = user_data[chat_id]['current_status']
        orders = get_orders_by_status(status)
        if order_code in orders:
            order = orders[order_code]
            item = None
            for it in order['items']:
                if it['code'] == item_code:
                    item = it
                    break

            if item:
                item_info = format_order_item(order_code, item, item_index)
                bot.send_message(
                    chat_id,
                    f"✏️ Редактирование позиции:\n{item_info}",
                    reply_markup=edit_item_keyboard(order_code, item_code, item_index),
                    parse_mode='HTML'
                )

    elif state == 'viewing_my_tasks':
        # Назад из моих задач -> главное меню
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "Главное меню:",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )

    elif state == 'changing_task_readiness':
        # Назад из изменения готовности -> к списку задач
        user_data[chat_id]['state'] = 'viewing_my_tasks'
        delete_previous_messages(chat_id, message.message_id)
        show_my_tasks_page(chat_id, user_data[chat_id]['my_tasks_page'])


# Обработчик кнопки "Изменить ФИО" в профиле
@bot.message_handler(func=lambda message: message.text == '✏️ Изменить ФИО' and is_authenticated(message.chat.id))
def change_fio_handler(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'awaiting_new_fio'
    delete_previous_messages(chat_id, message.message_id)

    current_fio = ADMIN_DATA[chat_id]['fio'] if chat_id in ADMIN_DATA else "не указано"

    bot.send_message(
        chat_id,
        f"✏️ <b>Изменение ФИО</b>\n\n"
        f"Текущее ФИО: <b>{current_fio}</b>\n\n"
        f"Введите новое ФИО:",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )


# Обработчик ввода нового ФИО
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get('state') == 'awaiting_new_fio' and is_authenticated(
        message.chat.id))
def handle_new_fio_input(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    new_fio = message.text.strip()

    if not new_fio:
        bot.send_message(
            chat_id,
            "❌ ФИО не может быть пустым. Пожалуйста, введите ваше ФИО:",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        return

    if chat_id not in ADMIN_DATA:
        bot.send_message(
            chat_id,
            "❌ Ошибка: ваш профиль не найден",
            parse_mode='HTML',
            reply_markup=main_keyboard(chat_id)
        )
        return

    # Сохраняем старое ФИО для обновления в заказах
    old_fio = ADMIN_DATA[chat_id]['fio']

    # Обновляем ФИО в профиле
    if update_worker(chat_id, 'fio', new_fio):
        bot.send_message(
            chat_id,
            f"✅ ФИО успешно изменено!\n\n"
            f"<b>Было:</b> {old_fio}\n"
            f"<b>Стало:</b> {new_fio}\n\n"
            f"✅ ФИО также обновлено во всех заказах, где вы были назначены.",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

        # Возвращаемся к просмотру профиля
        user_data[chat_id]['state'] = 'viewing_profile'

        worker_data = ADMIN_DATA[chat_id].copy()
        worker_data['chat_id'] = chat_id
        profile_info = format_profile_info(worker_data)

        bot.send_message(
            chat_id,
            profile_info,
            reply_markup=profile_keyboard(),
            parse_mode='HTML'
        )
    else:
        bot.send_message(
            chat_id,
            "❌ Ошибка при изменении ФИО. Попробуйте еще раз.",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Обработчик выбора статуса
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get('state') == 'viewing_statuses' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def show_orders_by_status(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    status = message.text

    # Сохраняем выбранный статус и сбрасываем страницу
    user_data[chat_id]['current_status'] = status
    user_data[chat_id]['current_page'] = 0
    user_data[chat_id]['state'] = 'viewing_orders'

    # Загружаем заказы по статусу
    orders = get_orders_by_status(status)
    order_codes = list(orders.keys())
    user_data[chat_id]['orders'] = order_codes

    # Отображаем первую страницу заказов
    show_orders_page(chat_id, 0)


# Обработчик выбора роли работников
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get(
        'state') == 'viewing_workers_roles' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def show_workers_by_role(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    role = message.text

    # Сохраняем выбранную роль и сбрасываем страницу
    user_data[chat_id]['workers_role'] = role
    user_data[chat_id]['workers_page'] = 0
    user_data[chat_id]['state'] = 'viewing_workers'

    # Отображаем первую страницу работников
    show_workers_page(chat_id, 0)


# Обработчик действий в меню управления паролем
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get('state') == 'password_management' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def handle_password_management(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    text = message.text

    if text == 'Изменить пароль':
        user_data[chat_id]['state'] = 'awaiting_new_password'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "🔐 Введите новый пароль:",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

    elif text == 'Удалить пароль':
        if os.path.exists(PASSWORD_FILE):
            user_data[chat_id]['state'] = 'confirming_password_delete'
            delete_previous_messages(chat_id, message.message_id)
            bot.send_message(
                chat_id,
                "⚠️ <b>ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ</b>\n\n"
                "Вы собираетесь удалить файл пароля.\n"
                "После этого аутентификация будет отключена.\n\n"
                "Для подтверждения введите: <b>УДАЛИТЬ</b>\n"
                "Это действие нельзя отменить!",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
        else:
            bot.send_message(
                chat_id,
                "❌ Файл пароля не существует!",
                parse_mode='HTML',
                reply_markup=password_management_keyboard()
            )


# Обработчик ввода нового пароля
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get(
        'state') == 'awaiting_new_password' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def handle_new_password_input(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    new_password = message.text.strip()

    try:
        with open(PASSWORD_FILE, 'w', encoding='utf-8') as f:
            f.write(new_password)

        # Обновляем глобальную переменную
        global PASSWORD
        PASSWORD = new_password

        bot.send_message(
            chat_id,
            f"✅ Пароль успешно {'установлен' if not os.path.exists(PASSWORD_FILE) else 'обновлен'}!",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

        # Возвращаемся к управлению паролем
        user_data[chat_id]['state'] = 'password_management'
        password_status = "установлен" if os.path.exists(PASSWORD_FILE) else "не установлен"
        bot.send_message(
            chat_id,
            f"🔐 <b>Управление паролем</b>\n\n"
            f"Текущий статус: {password_status}\n\n"
            f"Выберите действие:",
            reply_markup=password_management_keyboard(),
            parse_mode='HTML'
        )

    except Exception as e:
        bot.send_message(
            chat_id,
            f"❌ Ошибка при сохранении пароля: {str(e)}",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Обработчик подтверждения удаления пароля
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get(
        'state') == 'confirming_password_delete' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def handle_password_delete_confirmation(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    text = message.text.strip()

    if text == "УДАЛИТЬ":
        try:
            if os.path.exists(PASSWORD_FILE):
                os.remove(PASSWORD_FILE)
                global PASSWORD
                PASSWORD = None

                bot.send_message(
                    chat_id,
                    "✅ Файл пароля успешно удален! Аутентификация отключена.",
                    parse_mode='HTML',
                    reply_markup=common_keyboard()
                )
            else:
                bot.send_message(
                    chat_id,
                    "❌ Файл пароля не существует!",
                    parse_mode='HTML',
                    reply_markup=common_keyboard()
                )
        except Exception as e:
            bot.send_message(
                chat_id,
                f"❌ Ошибка при удалении файла пароля: {str(e)}",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )

        # Возвращаемся к управлению паролем
        user_data[chat_id]['state'] = 'password_management'
        password_status = "установлен" if os.path.exists(PASSWORD_FILE) else "не установлен"
        bot.send_message(
            chat_id,
            f"🔐 <b>Управление паролем</b>\n\n"
            f"Текущий статус: {password_status}\n\n"
            f"Выберите действие:",
            reply_markup=password_management_keyboard(),
            parse_mode='HTML'
        )
    else:
        bot.send_message(
            chat_id,
            "❌ Неверное подтверждение!\n"
            "Для удаления пароля введите: <b>УДАЛИТЬ</b>\n"
            "Используйте заглавные буквы как показано.",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Функция для отображения страницы заказов
def show_orders_page(chat_id, page):
    ensure_user_data(chat_id)  # Гарантируем наличие записи

    # Загружаем актуальные заказы по текущему статусу
    status = user_data[chat_id]['current_status']
    orders_dict = get_orders_by_status(status)

    # Обновляем список заказов
    order_codes = list(orders_dict.keys())
    user_data[chat_id]['orders'] = order_codes

    total_orders = len(order_codes)
    total_pages = max(1, (total_orders + 9) // 10)  # Округление вверх, минимум 1 страница

    # Корректировка номера страницы
    if page < 0:
        page = 0
    elif page >= total_pages and total_pages > 0:
        page = total_pages - 1

    user_data[chat_id]['current_page'] = page
    start_idx = page * 10
    end_idx = min(start_idx + 10, total_orders)
    page_orders = order_codes[start_idx:end_idx]

    # Отправляем информацию о странице
    if total_orders == 0:
        bot.send_message(
            chat_id,
            f"📭 Заказы со статусом '{status}' отсутствуют",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        return

    bot.send_message(
        chat_id,
        f"📄 Страница {page + 1}/{total_pages}\n"
        f"📊 Статус: {status}\n"
        f"📦 Заказы: {start_idx + 1}-{end_idx} из {total_orders}",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )

    # Отправляем каждый заказ на странице
    for order_code in page_orders:
        if order_code not in orders_dict:
            continue  # Пропускаем несуществующие заказы

        order = orders_dict[order_code]

        # Отправляем основную информацию о заказе (без работников)
        order_info = format_order_main_info(order)

        # Создаем кнопку "Написать"
        write_button = types.InlineKeyboardButton(
            '📨 Написать',
            url=order['telegram_contact']
        )

        # Создаем клавиатуру действий для заказа
        actions_keyboard = order_actions_keyboard(order_code)
        actions_keyboard.add(write_button)

        bot.send_message(
            chat_id,
            order_info,
            reply_markup=actions_keyboard,
            parse_mode='HTML'
        )

        # Отправляем каждую позицию отдельным сообщением с работниками и кнопкой редактирования
        if order['items']:
            for idx, item in enumerate(order['items'], 1):
                item_info = format_order_item(order_code, item, idx)

                # Извлекаем код материала из поля material (формат: "Код: XXXXX")
                material_code = None
                if 'Код:' in item['material']:
                    material_code = item['material'].split('Код:')[-1].strip()
                else:
                    material_code = item['material']  # Если код материала хранится напрямую

                item_keyboard = item_actions_keyboard(order_code, item['code'], idx, material_code)
                bot.send_message(
                    chat_id,
                    item_info,
                    reply_markup=item_keyboard,
                    parse_mode='HTML'
                )

    # Отправляем клавиатуру пагинации
    if total_pages > 1:
        pagination_keyboard = orders_pagination_keyboard(page, total_pages)
        bot.send_message(
            chat_id,
            "⬅️➡️ Переключение страниц:",
            reply_markup=pagination_keyboard,
            parse_mode='HTML'
        )


# Функция для отображения страницы работников
def show_workers_page(chat_id, page):
    ensure_user_data(chat_id)  # Гарантируем наличие записи

    # Загружаем актуальных работников по текущей роли
    role = user_data[chat_id]['workers_role']
    workers_dict = get_workers_by_role(role)

    # Обновляем список работников
    worker_ids = list(workers_dict.keys())
    user_data[chat_id]['workers_list'] = worker_ids

    total_workers = len(worker_ids)
    total_pages = max(1, (total_workers + 9) // 10)  # Округление вверх, минимум 1 страница

    # Корректировка номера страницы
    if page < 0:
        page = 0
    elif page >= total_pages and total_pages > 0:
        page = total_pages - 1

    user_data[chat_id]['workers_page'] = page
    start_idx = page * 10
    end_idx = min(start_idx + 10, total_workers)
    page_workers = worker_ids[start_idx:end_idx]

    # Отправляем информацию о странице
    if total_workers == 0:
        bot.send_message(
            chat_id,
            f"👥 Работники с ролью '{role}' отсутствуют",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        return

    bot.send_message(
        chat_id,
        f"📄 Страница {page + 1}/{total_pages}\n"
        f"📋 Роль: {role}\n"
        f"👥 Работники: {start_idx + 1}-{end_idx} из {total_workers}",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )

    # Отправляем каждого работника на странице
    for worker_id in page_workers:
        if worker_id not in workers_dict:
            continue  # Пропускаем несуществующих работников

        worker = workers_dict[worker_id]

        # Форматируем информацию о работнике
        worker_info = format_worker_info(worker_id, worker)

        # Создаем клавиатуру действий
        actions_keyboard = worker_actions_keyboard(worker_id, chat_id)

        bot.send_message(
            chat_id,
            worker_info,
            reply_markup=actions_keyboard,
            parse_mode='HTML'
        )

    # Отправляем клавиатуру пагинации
    if total_pages > 1:
        pagination_keyboard = workers_pagination_keyboard(page, total_pages)
        bot.send_message(
            chat_id,
            "⬅️➡️ Переключение страниц:",
            reply_markup=pagination_keyboard,
            parse_mode='HTML'
        )


# Функция для отображения работников для назначения на роль
def show_workers_for_role(chat_id, role, order_code, item_code, item_index, page=0):
    ensure_user_data(chat_id)

    # Получаем работников для роли (в новом порядке: Специалист -> Без роли -> Администратор)
    workers = get_workers_for_role(role)
    user_data[chat_id]['current_workers_for_role'] = workers

    total_workers = len(workers)
    total_pages = max(1, (total_workers + 9) // 10)  # Округление вверх, минимум 1 страница

    # Корректировка номера страницы
    if page < 0:
        page = 0
    elif page >= total_pages and total_pages > 0:
        page = total_pages - 1

    user_data[chat_id]['current_workers_page'] = page
    start_idx = page * 10
    end_idx = min(start_idx + 10, total_workers)
    page_workers = workers[start_idx:end_idx]

    # Роль для отображения пользователю
    role_names = {
        'seamstress': 'швею',
        'carpenter': 'столяра',
        'upholsterer': 'обтяжчика',
        'packer': 'упаковщика'
    }
    russian_role = role_names.get(role, role)

    # Отправляем информацию о странице
    if total_workers == 0:
        # Создаем инлайн-клавиатуру с кнопкой "Ввести вручную" и кнопками навигации
        markup = workers_for_role_keyboard([], 0, 0, role, order_code, item_code, item_index)
        bot.send_message(
            chat_id,
            f"👥 Нет доступных работников для роли '{russian_role}'. Вы можете ввести ФИО вручную:",
            parse_mode='HTML',
            reply_markup=markup
        )
        return

    bot.send_message(
        chat_id,
        f"👥 Выберите {russian_role} из списка (Страница {page + 1}/{total_pages}):",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )

    # Отправляем каждого работника на странице с кнопкой "Назначить"
    for worker_id, worker_data in page_workers:
        worker_info = format_worker_info(worker_id, worker_data)

        # Создаем клавиатуру с кнопкой "Назначить"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton(
            '✅ Назначить',
            callback_data=f'select_worker_{worker_id}_{role}_{order_code}_{item_code}_{item_index}'
        ))

        bot.send_message(
            chat_id,
            worker_info,
            reply_markup=markup,
            parse_mode='HTML'
        )

    # Отправляем клавиатуру пагинации и дополнительных действий
    markup = workers_for_role_keyboard(page_workers, page, total_pages, role, order_code, item_code, item_index)
    bot.send_message(
        chat_id,
        "⬅️➡️ Переключение страниц или ввод ФИО вручную:",
        reply_markup=markup,
        parse_mode='HTML'
    )


# Обработчик колбэков пагинации
@bot.callback_query_handler(
    func=lambda call: (call.data.startswith('prev_') or call.data.startswith('next_')) and is_authenticated(
        call.message.chat.id) and is_admin(call.message.chat.id))
def pagination_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    data = call.data.split('_')
    action = data[0]
    current_page = int(data[1])

    if action == 'prev':
        new_page = current_page - 1
    else:  # next
        new_page = current_page + 1

    delete_previous_messages(chat_id, call.message.message_id)
    show_orders_page(chat_id, new_page)
    bot.answer_callback_query(call.id)


# Обработчик колбэков пагинации работников
@bot.callback_query_handler(
    func=lambda call: (call.data.startswith('workers_prev_') or call.data.startswith(
        'workers_next_')) and is_authenticated(
        call.message.chat.id) and is_admin(call.message.chat.id))
def workers_pagination_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    data = call.data.split('_')
    action = data[1]
    current_page = int(data[2])

    if action == 'prev':
        new_page = current_page - 1
    else:  # next
        new_page = current_page + 1

    delete_previous_messages(chat_id, call.message.message_id)
    show_workers_page(chat_id, new_page)
    bot.answer_callback_query(call.id)


# Обработчик колбэков пагинации работников для роли
@bot.callback_query_handler(
    func=lambda call: (call.data.startswith('workers_role_prev_') or call.data.startswith(
        'workers_role_next_')) and is_authenticated(
        call.message.chat.id) and is_admin(call.message.chat.id))
def workers_role_pagination_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')
    action = data[2]
    current_page = int(data[3])
    role = data[4]
    order_code = data[5]
    item_code = data[6]
    item_index = int(data[7])

    if action == 'prev':
        new_page = current_page - 1
    else:  # next
        new_page = current_page + 1

    delete_previous_messages(chat_id, call.message.message_id)
    show_workers_for_role(chat_id, role, order_code, item_code, item_index, new_page)
    bot.answer_callback_query(call.id)


# Обработчик колбэков пагинации моих задач
@bot.callback_query_handler(
    func=lambda call: (call.data.startswith('my_tasks_prev_') or call.data.startswith(
        'my_tasks_next_')) and is_authenticated(call.message.chat.id))
def my_tasks_pagination_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')
    action = data[2]
    current_page = int(data[3])

    if action == 'prev':
        new_page = current_page - 1
    else:  # next
        new_page = current_page + 1

    delete_previous_messages(chat_id, call.message.message_id)
    show_my_tasks_page(chat_id, new_page)
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад к статусам"
@bot.callback_query_handler(
    func=lambda call: call.data == 'back_to_statuses' and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def back_to_statuses_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id]['state'] = 'viewing_statuses'
    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        "📊 Выберите статус заказов:",
        reply_markup=statuses_keyboard(),
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад к ролям"
@bot.callback_query_handler(
    func=lambda call: call.data == 'back_to_workers_roles' and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def back_to_workers_roles_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id]['state'] = 'viewing_workers_roles'
    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        "👥 Выберите роль работников:",
        reply_markup=workers_roles_keyboard(),
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад к списку работников"
@bot.callback_query_handler(
    func=lambda call: call.data == 'back_to_workers_list' and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def back_to_workers_list_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id]['state'] = 'viewing_workers'
    delete_previous_messages(chat_id, call.message.message_id)
    show_workers_page(chat_id, user_data[chat_id]['workers_page'])
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад к моим задачам"
@bot.callback_query_handler(
    func=lambda call: call.data == 'back_to_my_tasks' and is_authenticated(call.message.chat.id))
def back_to_my_tasks_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'viewing_my_tasks'
    delete_previous_messages(chat_id, call.message.message_id)
    show_my_tasks_page(chat_id, user_data[chat_id]['my_tasks_page'])
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад к списку задач"
@bot.callback_query_handler(
    func=lambda call: call.data == 'back_to_my_tasks_list' and is_authenticated(call.message.chat.id))
def back_to_my_tasks_list_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    user_data[chat_id]['state'] = 'viewing_my_tasks'
    delete_previous_messages(chat_id, call.message.message_id)
    show_my_tasks_page(chat_id, user_data[chat_id]['my_tasks_page'])
    bot.answer_callback_query(call.id)


# Обработчик кнопки "На главную"
@bot.callback_query_handler(func=lambda call: call.data == 'to_main_menu' and is_authenticated(call.message.chat.id))
def to_main_menu_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    user_data[chat_id] = {
        'state': 'main_menu',
        'current_status': None,
        'current_page': 0,
        'orders': [],
        'current_order': None,
        'current_role': None,
        'delete_order_code': None,
        'workers_role': None,
        'workers_page': 0,
        'workers_list': [],
        'current_worker_id': None,
        'delete_worker_id': None,
        'current_item_code': None,
        'current_item_index': None,
        'current_workers_for_role': [],
        'current_workers_page': 0,
        'my_tasks_page': 0,
        'my_tasks': [],
        'current_task': None,
        'current_task_role': None
    }
    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        "Главное меню:",
        reply_markup=main_keyboard(chat_id),
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Редактировать позицию"
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('edit_item_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def edit_item_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    data = call.data.split('_')
    order_code = data[2]
    item_code = data[3]
    item_index = int(data[4])

    user_data[chat_id]['current_order'] = order_code
    user_data[chat_id]['current_item_code'] = item_code
    user_data[chat_id]['current_item_index'] = item_index
    user_data[chat_id]['state'] = 'editing_item'

    delete_previous_messages(chat_id, call.message.message_id)

    # Получаем полную информацию о заказе и позиции
    status = user_data[chat_id]['current_status']
    orders = get_orders_by_status(status)

    # Проверяем существование заказа и позиции
    if order_code not in orders:
        bot.send_message(
            chat_id,
            f"❌ Заказ <code>{order_code}</code> не найден!",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        show_orders_page(chat_id, user_data[chat_id]['current_page'])
        return

    order = orders[order_code]
    item = None
    for it in order['items']:
        if it['code'] == item_code:
            item = it
            break

    if item is None:
        bot.send_message(
            chat_id,
            f"❌ Позиция <code>{item_code}</code> не найдена!",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        show_orders_page(chat_id, user_data[chat_id]['current_page'])
        return

    # Форматируем информацию о позиции
    item_info = format_order_item(order_code, item, item_index)

    bot.send_message(
        chat_id,
        f"✏️ Редактирование позиции:\n{item_info}",
        reply_markup=edit_item_keyboard(order_code, item_code, item_index),
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Удалить заказ"
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('delete_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def delete_order_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    order_code = call.data.split('_')[1]

    # Сохраняем код заказа для удаления
    user_data[chat_id]['delete_order_code'] = order_code
    user_data[chat_id]['state'] = 'confirming_delete'

    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        f"⚠️ <b>ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ</b>\n\n"
        f"Вы собираетесь удалить заказ: <code>{order_code}</code>\n"
        "Для подтверждения введите: <b>УДАЛИТЬ</b>\n"
        "Это действие нельзя отменить!",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )
    bot.answer_callback_query(call.id)


# Обработчик действий с работником
@bot.callback_query_handler(
    func=lambda call: (call.data.startswith('worker_') or call.data.startswith(
        'back_to_worker_') or call.data.startswith('newrole_')) and is_authenticated(
        call.message.chat.id) and is_admin(call.message.chat.id))
def worker_action_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи

    if call.data.startswith('worker_edit_name_'):
        # Изменение имени работника
        worker_id = int(call.data.split('_')[3])
        user_data[chat_id]['current_worker_id'] = worker_id
        user_data[chat_id]['state'] = 'awaiting_worker_name'

        delete_previous_messages(chat_id, call.message.message_id)
        bot.send_message(
            chat_id,
            f"✏️ Введите новое ФИО для работника:\n"
            f"Текущее ФИО: {ADMIN_DATA[worker_id]['fio']}",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

    elif call.data.startswith('worker_edit_role_'):
        # Изменение роли работника
        worker_id = int(call.data.split('_')[3])
        user_data[chat_id]['current_worker_id'] = worker_id
        user_data[chat_id]['state'] = 'changing_worker_role'

        delete_previous_messages(chat_id, call.message.message_id)
        bot.send_message(
            chat_id,
            f"🔄 Выберите новую роль для работника:\n"
            f"Текущая роль: {ADMIN_DATA[worker_id]['role']}",
            reply_markup=role_change_keyboard(worker_id),
            parse_mode='HTML'
        )

    elif call.data.startswith('worker_delete_'):
        # Удаление работника
        worker_id = int(call.data.split('_')[2])

        # Проверяем, не пытается ли пользователь удалить себя
        if worker_id == chat_id:
            bot.answer_callback_query(call.id, "❌ Вы не можете удалить себя!")
            return

        user_data[chat_id]['current_worker_id'] = worker_id
        user_data[chat_id]['delete_worker_id'] = worker_id
        user_data[chat_id]['state'] = 'confirming_worker_delete'

        delete_previous_messages(chat_id, call.message.message_id)
        bot.send_message(
            chat_id,
            f"⚠️ <b>ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ</b>\n\n"
            f"Вы собираетесь удалить работника:\n"
            f"{format_worker_info(worker_id, ADMIN_DATA[worker_id])}\n\n"
            "Для подтверждения введите: <b>УДАЛИТЬ</b>\n"
            "Это действие нельзя отменить!",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

    elif call.data.startswith('worker_message_'):
        # Написать работнику
        worker_id = int(call.data.split('_')[2])
        worker_data = ADMIN_DATA[worker_id]

        # Создаем кнопку для написания сообщения
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton('📨 Написать', url=worker_data['link']))
        markup.add(types.InlineKeyboardButton('Назад', callback_data='back_to_workers_list'))

        bot.send_message(
            chat_id,
            f"✉️ Написать работнику:\n{format_worker_info(worker_id, worker_data)}",
            reply_markup=markup,
            parse_mode='HTML'
        )

    elif call.data.startswith('back_to_worker_'):
        # Назад к редактированию работника
        worker_id = int(call.data.split('_')[3])
        user_data[chat_id]['current_worker_id'] = worker_id
        user_data[chat_id]['state'] = 'editing_worker'

        delete_previous_messages(chat_id, call.message.message_id)
        bot.send_message(
            chat_id,
            f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
            reply_markup=worker_actions_keyboard(worker_id, chat_id),
            parse_mode='HTML'
        )

    elif call.data.startswith('newrole_'):
        # Установка новой роли
        worker_id = int(call.data.split('_')[1])
        new_role = call.data.split('_')[2]

        # Проверяем, не пытается ли пользователь изменить свою роль
        if worker_id == chat_id:
            bot.answer_callback_query(call.id, "❌ Вы не можете изменить свою роль!")
            return

        if update_worker(worker_id, 'role', new_role):
            bot.answer_callback_query(call.id, "✅ Роль успешно обновлена")

            # Возвращаемся к списку работников
            user_data[chat_id]['state'] = 'viewing_workers'
            delete_previous_messages(chat_id, call.message.message_id)
            show_workers_page(chat_id, user_data[chat_id]['workers_page'])
        else:
            bot.answer_callback_query(call.id, "❌ Ошибка обновления роли")

    bot.answer_callback_query(call.id)


# Обработчик подтверждения удаления заказа
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get('state') == 'confirming_delete' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def confirm_delete_handler(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    text = message.text.strip()
    order_code = user_data[chat_id]['delete_order_code']

    if text == "УДАЛИТЬ":
        if delete_order(order_code):
            bot.send_message(
                chat_id,
                f"✅ Заказ <code>{order_code}</code> успешно удален!",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
            # Возвращаемся к списку заказов
            user_data[chat_id]['state'] = 'viewing_orders'
            show_orders_page(chat_id, user_data[chat_id]['current_page'])
        else:
            bot.send_message(
                chat_id,
                f"❌ Ошибка при удалении заказа <code>{order_code}</code>",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
            # Возвращаемся к списку заказов
            user_data[chat_id]['state'] = 'viewing_orders'
            show_orders_page(chat_id, user_data[chat_id]['current_page'])
    else:
        bot.send_message(
            chat_id,
            "❌ Неверное подтверждение!\n"
            "Для удаления заказа введите: <b>УДАЛИТЬ</b>\n"
            "Используйте заглавные буквы как показано.",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Обработчик подтверждения удаления работника
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get(
        'state') == 'confirming_worker_delete' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def confirm_worker_delete_handler(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    text = message.text.strip()
    worker_id = user_data[chat_id]['delete_worker_id']

    if text == "УДАЛИТЬ":
        # Проверяем, не пытается ли пользователь удалить себя
        if worker_id == chat_id:
            bot.send_message(
                chat_id,
                "❌ Вы не можете удалить себя!",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
            return

        if delete_worker(worker_id):
            bot.send_message(
                chat_id,
                f"✅ Работник успешно удален!",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
            # Возвращаемся к списку работников
            user_data[chat_id]['state'] = 'viewing_workers'
            show_workers_page(chat_id, user_data[chat_id]['workers_page'])
        else:
            bot.send_message(
                chat_id,
                f"❌ Ошибка при удалении работника",
                parse_mode='HTML',
                reply_markup=common_keyboard()
            )
            # Возвращаемся к редактированию
            worker_id = user_data[chat_id]['current_worker_id']
            user_data[chat_id]['state'] = 'editing_worker'
            bot.send_message(
                chat_id,
                f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
                reply_markup=worker_actions_keyboard(worker_id, chat_id),
                parse_mode='HTML'
            )
    else:
        bot.send_message(
            chat_id,
            "❌ Неверное подтверждение!\n"
            "Для удаления работника введите: <b>УДАЛИТЬ</b>\n"
            "Используйте заглавные буквы как показано.",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Обработчик ввода нового имени работника
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get('state') == 'awaiting_worker_name' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def handle_worker_name_input(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    new_fio = message.text.strip()
    worker_id = user_data[chat_id]['current_worker_id']

    if update_worker(worker_id, 'fio', new_fio):
        bot.send_message(
            chat_id,
            f"✅ ФИО работника успешно обновлено на: {new_fio}",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        # Возвращаемся к списку работников
        user_data[chat_id]['state'] = 'viewing_workers'
        show_workers_page(chat_id, user_data[chat_id]['workers_page'])
    else:
        bot.send_message(
            chat_id,
            f"❌ Ошибка при обновлении ФИО работника",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        # Возвращаемся к редактированию
        user_data[chat_id]['state'] = 'editing_worker'
        bot.send_message(
            chat_id,
            f"✏️ Редактирование работника:\n{format_worker_info(worker_id, ADMIN_DATA[worker_id])}",
            reply_markup=worker_actions_keyboard(worker_id, chat_id),
            parse_mode='HTML'
        )


# Обработчик кнопки "Назад" при просмотре заказа
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('back_to_order_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def back_to_order_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    order_code = call.data.split('_')[3]

    # Возвращаемся к просмотру заказов
    user_data[chat_id]['state'] = 'viewing_orders'
    delete_previous_messages(chat_id, call.message.message_id)
    show_orders_page(chat_id, user_data[chat_id]['current_page'])
    bot.answer_callback_query(call.id)


# Обработчик кнопки "Назад" при редактировании позиции
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('back_to_edit_item_') and is_authenticated(
        call.message.chat.id) and is_admin(
        call.message.chat.id))
def back_to_edit_item_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')
    order_code = data[4]
    item_code = data[5]
    item_index = int(data[6])

    user_data[chat_id]['current_order'] = order_code
    user_data[chat_id]['current_item_code'] = item_code
    user_data[chat_id]['current_item_index'] = item_index
    user_data[chat_id]['state'] = 'editing_item'

    delete_previous_messages(chat_id, call.message.message_id)

    # Получаем актуальные данные о позиции
    status = user_data[chat_id]['current_status']
    orders = get_orders_by_status(status)
    if order_code in orders:
        order = orders[order_code]
        item = None
        for it in order['items']:
            if it['code'] == item_code:
                item = it
                break

        if item:
            item_info = format_order_item(order_code, item, item_index)
            bot.send_message(
                chat_id,
                f"✏️ Редактирование позиции:\n{item_info}",
                reply_markup=edit_item_keyboard(order_code, item_code, item_index),
                parse_mode='HTML'
            )

    bot.answer_callback_query(call.id)


# Обработчик выбора действия для позиции
@bot.callback_query_handler(func=lambda call: (call.data.startswith('assign_item_')) and is_authenticated(
    call.message.chat.id) and is_admin(call.message.chat.id))
def edit_item_action_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    data = call.data.split('_')

    role = data[2]  # seamstress, carpenter, upholsterer, packer
    order_code = data[3]
    item_code = data[4]
    item_index = int(data[5])

    user_data[chat_id]['current_order'] = order_code
    user_data[chat_id]['current_item_code'] = item_code
    user_data[chat_id]['current_item_index'] = item_index
    user_data[chat_id]['current_role'] = role
    user_data[chat_id]['state'] = 'selecting_worker_for_role'

    delete_previous_messages(chat_id, call.message.message_id)
    show_workers_for_role(chat_id, role, order_code, item_code, item_index, 0)
    bot.answer_callback_query(call.id)


# Обработчик выбора работника для назначения
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('select_worker_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def select_worker_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')

    worker_id = int(data[2])
    role = data[3]
    order_code = data[4]
    item_code = data[5]
    item_index = int(data[6])

    # Получаем данные работника
    worker_data = ADMIN_DATA.get(worker_id)
    if not worker_data:
        bot.answer_callback_query(call.id, "❌ Работник не найден")
        return

    worker_fio = worker_data['fio']

    # Определяем поле для обновления в зависимости от роли
    field_mapping = {
        'seamstress': 'seamstress_fio',
        'carpenter': 'carpenter_fio',
        'upholsterer': 'upholsterer_fio',
        'packer': 'packer_fio'
    }

    field_to_update = field_mapping.get(role)

    if not field_to_update:
        bot.answer_callback_query(call.id, f"❌ Ошибка: неизвестная роль {role}")
        return

    # Обновляем позицию
    if update_order_item(order_code, item_code, field_to_update, worker_fio):
        # Роль для отображения пользователю
        role_names = {
            'seamstress': 'Швея',
            'carpenter': 'Столяр',
            'upholsterer': 'Обтяжчик',
            'packer': 'Упаковщик'
        }
        russian_role = role_names.get(role, role)

        bot.answer_callback_query(
            call.id,
            f"✅ {russian_role} назначен: {worker_fio}"
        )

        # Возвращаемся к редактированию позиции
        user_data[chat_id]['state'] = 'editing_item'
        delete_previous_messages(chat_id, call.message.message_id)

        # Получаем актуальные данные о позиции
        status = user_data[chat_id]['current_status']
        orders = get_orders_by_status(status)
        if order_code in orders:
            order = orders[order_code]
            item = None
            for it in order['items']:
                if it['code'] == item_code:
                    item = it
                    break

            if item:
                item_info = format_order_item(order_code, item, item_index)
                bot.send_message(
                    chat_id,
                    f"✏️ Редактирование позиции:\n{item_info}",
                    reply_markup=edit_item_keyboard(order_code, item_code, item_index),
                    parse_mode='HTML'
                )
    else:
        bot.answer_callback_query(call.id, "❌ Ошибка назначения работника")


# Обработчик ручного ввода ФИО
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('manual_input_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def manual_input_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')

    role = data[2]
    order_code = data[3]
    item_code = data[4]
    item_index = int(data[5])

    user_data[chat_id]['current_order'] = order_code
    user_data[chat_id]['current_item_code'] = item_code
    user_data[chat_id]['current_item_index'] = item_index
    user_data[chat_id]['current_role'] = role
    user_data[chat_id]['state'] = 'awaiting_item_worker_input'

    # Роль для отображения пользователю
    role_names = {
        'seamstress': 'швеи',
        'carpenter': 'столяра',
        'upholsterer': 'обтяжчика',
        'packer': 'упаковщика'
    }
    russian_role = role_names.get(role, role)

    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        f"✏️ Введите ФИО {russian_role} вручную:",
        parse_mode='HTML',
        reply_markup=common_keyboard()
    )
    bot.answer_callback_query(call.id)


# Обработчик ввода ФИО работника для позиции вручную
@bot.message_handler(
    func=lambda message: user_data.get(message.chat.id, {}).get(
        'state') == 'awaiting_item_worker_input' and is_authenticated(
        message.chat.id) and is_admin(message.chat.id))
def handle_item_worker_input(message):
    chat_id = message.chat.id
    ensure_user_data(chat_id)
    worker_fio = message.text.strip()

    if not worker_fio:
        bot.send_message(
            chat_id,
            "❌ ФИО не может быть пустым. Введите ФИО:",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        return

    order_code = user_data[chat_id]['current_order']
    item_code = user_data[chat_id]['current_item_code']
    role = user_data[chat_id]['current_role']

    # Определяем поле для обновления в зависимости от роли
    field_mapping = {
        'seamstress': 'seamstress_fio',
        'carpenter': 'carpenter_fio',
        'upholsterer': 'upholsterer_fio',
        'packer': 'packer_fio'
    }

    field_to_update = field_mapping.get(role)

    if not field_to_update:
        bot.send_message(
            chat_id,
            f"❌ Ошибка: неизвестная роль {role}",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )
        return

    # Обновляем позицию
    if update_order_item(order_code, item_code, field_to_update, worker_fio):
        # Роль для отображения пользователю
        role_names = {
            'seamstress': 'Швея',
            'carpenter': 'Столяр',
            'upholsterer': 'Обтяжчик',
            'packer': 'Упаковщик'
        }
        russian_role = role_names.get(role, role)

        bot.send_message(
            chat_id,
            f"✅ {russian_role} назначен: {worker_fio}",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )

        # Возвращаемся к редактированию позиции
        user_data[chat_id]['state'] = 'editing_item'
        delete_previous_messages(chat_id, message.message_id)

        # Получаем актуальные данные о позиции
        status = user_data[chat_id]['current_status']
        orders = get_orders_by_status(status)
        if order_code in orders:
            order = orders[order_code]
            item = None
            for it in order['items']:
                if it['code'] == item_code:
                    item = it
                    break

            if item:
                item_info = format_order_item(order_code, item, user_data[chat_id]['current_item_index'])
                bot.send_message(
                    chat_id,
                    f"✏️ Редактирование позиции:\n{item_info}",
                    reply_markup=edit_item_keyboard(order_code, item_code, user_data[chat_id]['current_item_index']),
                    parse_mode='HTML'
                )
    else:
        bot.send_message(
            chat_id,
            f"❌ Ошибка назначения",
            parse_mode='HTML',
            reply_markup=common_keyboard()
        )


# Обработчик кнопки "Изменить статус"
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('change_status_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def change_status_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    order_code = call.data.split('_')[2]

    user_data[chat_id]['current_order'] = order_code
    user_data[chat_id]['state'] = 'changing_status'

    delete_previous_messages(chat_id, call.message.message_id)
    bot.send_message(
        chat_id,
        f"🔄 Выберите новый статус для заказа <code>{order_code}</code>:",
        reply_markup=status_change_keyboard(order_code),
        parse_mode='HTML'
    )
    bot.answer_callback_query(call.id)


# Обработчик выбора нового статуса
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('newstatus_') and is_authenticated(call.message.chat.id) and is_admin(
        call.message.chat.id))
def new_status_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)  # Гарантируем наличие записи
    data = call.data.split('_')
    order_code = data[1]
    status = '_'.join(data[2:])  # Объединяем оставшиеся части для статусов с пробелами

    if update_order_status(order_code, status):
        bot.answer_callback_query(call.id, f"✅ Статус обновлен на: {status}")

        # Возвращаемся к списку заказов
        user_data[chat_id]['state'] = 'viewing_orders'
        delete_previous_messages(chat_id, call.message.message_id)
        show_orders_page(chat_id, user_data[chat_id]['current_page'])
    else:
        bot.answer_callback_query(call.id, "❌ Ошибка обновления статуса")


# Обработчик изменения готовности задачи
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('change_task_readiness_') and is_authenticated(call.message.chat.id))
def change_task_readiness_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')
    task_index = int(data[3])

    user_data[chat_id]['state'] = 'changing_task_readiness'
    user_data[chat_id]['current_task'] = task_index

    delete_previous_messages(chat_id, call.message.message_id)

    # Получаем задачу
    tasks = user_data[chat_id]['my_tasks']
    if task_index < 0 or task_index >= len(tasks):
        bot.send_message(
            chat_id,
            "❌ Задача не найдена!",
            parse_mode='HTML',
            reply_markup=my_tasks_keyboard()
        )
        show_my_tasks_page(chat_id, user_data[chat_id]['my_tasks_page'])
        return

    task = tasks[task_index]

    # Извлекаем код материала из поля material (формат: "Код: XXXXX")
    material_code = None
    if 'Код:' in task['material']:
        material_code = task['material'].split('Код:')[-1].strip()
    else:
        material_code = task['material']  # Если код материала хранится напрямую

    bot.send_message(
        chat_id,
        f"✏️ <b>Изменение готовности задачи</b>\n\n"
        f"📋 <b>Задача:</b> {task['order_code']} - {task['item_code']}\n"
        f"👤 <b>Ваша роль:</b> {task['user_role']}\n"
        f"📊 <b>Текущая готовность:</b> {task['readiness']}\n\n"
        f"Выберите новую готовность:",
        parse_mode='HTML',
        reply_markup=task_readiness_keyboard(task_index, task['item_code'], material_code)
    )
    bot.answer_callback_query(call.id)


# Обработчик выбора новой готовности задачи
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('change_readiness_') and is_authenticated(call.message.chat.id))
def change_readiness_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')
    task_index = int(data[2])
    readiness = data[3]

    # Получаем задачу
    tasks = user_data[chat_id]['my_tasks']
    if task_index < 0 or task_index >= len(tasks):
        bot.answer_callback_query(call.id, "❌ Задача не найдена!")
        return

    task = tasks[task_index]

    # Обновляем готовность
    if update_task_readiness(task['row'], task['readiness_column'], readiness):
        bot.answer_callback_query(call.id, f"✅ Готовность обновлена на: {readiness}")

        # Обновляем задачу в списке
        tasks[task_index]['readiness'] = readiness
        user_data[chat_id]['my_tasks'] = tasks

        # Возвращаемся к списку задач
        user_data[chat_id]['state'] = 'viewing_my_tasks'
        delete_previous_messages(chat_id, call.message.message_id)
        show_my_tasks_page(chat_id, user_data[chat_id]['my_tasks_page'])
    else:
        bot.answer_callback_query(call.id, "❌ Ошибка обновления готовности")


# НОВЫЙ ОБРАБОТЧИК ДЛЯ КНОПКИ "ПОДРОБНОСТИ"
@bot.callback_query_handler(
    func=lambda call: call.data.startswith('details_') and is_authenticated(call.message.chat.id))
def details_handler(call):
    chat_id = call.message.chat.id
    ensure_user_data(chat_id)
    data = call.data.split('_')

    if len(data) < 3:
        bot.answer_callback_query(call.id, "❌ Ошибка: неверный формат данных")
        return

    product_code = data[1]
    material_code = data[2]

    # Показываем карточку товара
    show_product_card(chat_id, product_code)

    # Показываем карточку материала
    show_material_card(chat_id, material_code)

    bot.answer_callback_query(call.id, "📋 Показаны карточки товара и материала")


# Обработчик команды /cancel
@bot.message_handler(commands=['cancel'])
def cancel_handler(message):
    chat_id = message.chat.id

    if not is_authenticated(chat_id):
        bot.send_message(chat_id, "🔒 Для доступа введите пароль:")
        return

    ensure_user_data(chat_id)
    state = user_data[chat_id]['state']

    if state in ['awaiting_new_fio', 'awaiting_worker_name', 'awaiting_item_worker_input',
                 'awaiting_new_password', 'confirming_delete', 'confirming_worker_delete',
                 'confirming_password_delete', 'changing_task_readiness']:
        user_data[chat_id]['state'] = 'main_menu'
        delete_previous_messages(chat_id, message.message_id)
        bot.send_message(
            chat_id,
            "❌ Операция отменена",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )
    elif state in ['changing_status', 'changing_worker_role', 'selecting_worker_for_role']:
        user_data[chat_id]['state'] = 'viewing_orders'
        delete_previous_messages(chat_id, message.message_id)
        show_orders_page(chat_id, user_data[chat_id]['current_page'])
    else:
        bot.send_message(
            chat_id,
            "❌ Нечего отменять",
            parse_mode='HTML'
        )


# Обработчик неизвестных сообщений
@bot.message_handler(func=lambda message: True)
def unknown_message(message):
    chat_id = message.chat.id

    if not is_authenticated(chat_id):
        bot.send_message(chat_id, "🔒 Для доступа введите пароль:")
        return

    ensure_user_data(chat_id)
    state = user_data[chat_id]['state']

    if state == 'main_menu':
        bot.send_message(
            chat_id,
            "❓ Неизвестная команда. Используйте кнопки меню.",
            reply_markup=main_keyboard(chat_id),
            parse_mode='HTML'
        )
    else:
        bot.send_message(
            chat_id,
            "❓ Неизвестная команда. Используйте кнопки меню.",
            parse_mode='HTML'
        )


# Запуск бота
if __name__ == '__main__':
    print("Бот запущен...")
    try:
        bot.polling(none_stop=True, interval=0, timeout=20)
    except Exception as e:
        print(f"Ошибка запуска бота: {e}")
        traceback.print_exc()