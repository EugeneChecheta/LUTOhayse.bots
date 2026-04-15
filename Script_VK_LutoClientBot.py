# -*- coding: utf-8 -*-
import vk_api
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.utils import get_random_id

import openpyxl
from openpyxl import Workbook, load_workbook
import os
import random
import string
from datetime import datetime
import traceback
from typing import List, Dict, Optional, Any, Tuple

# ================== КОНСТАНТЫ ==================
TOKEN = 'токен'
GROUP_ID = 0

MANAGER_IDS = [295730168]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, 'DataBase')
PHOTO_DIR = os.path.join(BASE_DIR, 'Photo')
PRODUCTS_PHOTO_DIR = os.path.join(PHOTO_DIR, 'Products')
MATERIALS_PHOTO_DIR = os.path.join(PHOTO_DIR, 'Materials')
CONSTRUCTION_PHOTO_DIR = os.path.join(PHOTO_DIR, 'Construction')

PRODUCTS_FILE = os.path.join(DB_DIR, 'Products_log.xlsx')
MATERIALS_FILE = os.path.join(DB_DIR, 'Materials_log.xlsx')
MODULES_FILE = os.path.join(DB_DIR, 'Moduls_log.xlsx')
SOFAS_FILE = os.path.join(DB_DIR, 'Sofas_log.xlsx')
ORDERS_FILE = os.path.join(DB_DIR, 'Orders_log.xlsx')
CART_FILE = os.path.join(DB_DIR, 'Cart_log.xlsx')

user_states = {}

# Типы диванов для конструирования (отображаемое имя -> префикс папки)
SOFA_TYPES = [
    ('Бауэн', 'CA'),
    ('Тайль', 'CP'),
    ('Вельтраум', 'CT')
]
# Маппинг префикса папки на букву модуля (для поиска в Moduls_log)
PREFIX_TO_MODULE_LETTER = {
    'CA': 'A',
    'CP': 'P',
    'CT': 'T'
}
# Обратный маппинг
MODULE_LETTER_TO_PREFIX = {v: k for k, v in PREFIX_TO_MODULE_LETTER.items()}

# Доступные типы материалов (колонки в Moduls_log)
MATERIAL_COLUMNS = ['Велюр', 'Рогожка', 'Букле', 'Эко-кожа']


# ================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==================
def ensure_dirs():
    for d in [DB_DIR, PRODUCTS_PHOTO_DIR, MATERIALS_PHOTO_DIR, CONSTRUCTION_PHOTO_DIR]:
        os.makedirs(d, exist_ok=True)


def generate_task_number() -> str:
    return str(random.randint(100000, 999999))


def generate_order_code() -> str:
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choices(chars, k=10))


def get_current_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_current_timestamp_display() -> str:
    return datetime.now().strftime("%d.%m.%Y %H:%M:%S")


def get_next_task_number() -> int:
    if not os.path.exists(ORDERS_FILE):
        return 1000000
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    max_num = 1000000 - 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            try:
                val = int(row[1])
                if val > max_num:
                    max_num = val
            except:
                pass
    return max_num + 1


def is_order_code_unique(code: str) -> bool:
    if not os.path.exists(ORDERS_FILE):
        return True
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == code:
            return False
    return True


def generate_unique_order_code() -> str:
    while True:
        code = generate_order_code()
        if is_order_code_unique(code):
            return code


# ================== РАБОТА С EXCEL (КОРЗИНА) ==================
def create_cart_table():
    if not os.path.exists(CART_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cart"
        headers = ["Мессенджер", "ID записи", "ID чата", "Тип позиции", "Код позиции",
                   "Название позиции", "Код материала", "Тип материала", "Цвет материала",
                   "Цена", "Дата добавления"]
        ws.append(headers)
        wb.save(CART_FILE)


def save_to_cart_full(messenger: str, chat_id: int, item_type: str, item_code: str,
                      item_name: str, material_code: str, material_type: str,
                      material_color: str, price: float):
    create_cart_table()
    wb = load_workbook(CART_FILE)
    ws = wb.active

    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            try:
                max_id = max(max_id, int(row[1]))
            except:
                pass
    new_id = max_id + 1

    date_added = get_current_timestamp()
    ws.append([messenger, new_id, chat_id, item_type, item_code, item_name,
               material_code, material_type, material_color, price, date_added])
    wb.save(CART_FILE)


def get_user_cart(user_id: int) -> List[Dict]:
    create_cart_table()
    if not os.path.exists(CART_FILE):
        return []
    wb = load_workbook(CART_FILE)
    ws = wb.active
    items = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == user_id:
            items.append(dict(zip(headers, row)))
    return items


def remove_from_cart(record_id: int):
    if not os.path.exists(CART_FILE):
        return
    wb = load_workbook(CART_FILE)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] == record_id:
            ws.delete_rows(i)
            break
    wb.save(CART_FILE)


def clear_user_cart(user_id: int):
    if not os.path.exists(CART_FILE):
        return
    wb = load_workbook(CART_FILE)
    ws = wb.active
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] == user_id:
            rows_to_delete.append(i)
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)
    wb.save(CART_FILE)


def update_cart_item(record_id: int, new_material_code: str, new_material_type: str,
                     new_material_color: str, new_price: float):
    if not os.path.exists(CART_FILE):
        return
    wb = load_workbook(CART_FILE)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] == record_id:
            ws.cell(row=i, column=7, value=new_material_code)
            ws.cell(row=i, column=8, value=new_material_type)
            ws.cell(row=i, column=9, value=new_material_color)
            ws.cell(row=i, column=10, value=new_price)
            ws.cell(row=i, column=11, value=get_current_timestamp())
            break
    wb.save(CART_FILE)


# ================== ЗАГРУЗКА ДАННЫХ ИЗ БАЗ ==================
def load_modules() -> List[Dict]:
    if not os.path.exists(MODULES_FILE):
        return []
    wb = load_workbook(MODULES_FILE)
    ws = wb.active
    modules = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            modules.append(dict(zip(headers, row)))
    return modules


def find_module_by_code(code: str) -> Optional[Dict]:
    modules = load_modules()
    for m in modules:
        if m.get('Код модуля') == code:
            return m
    return None


def get_module_details(module_code: str) -> Optional[Dict]:
    return find_module_by_code(module_code)


def get_module_photos(module_code: str) -> List[str]:
    folder = os.path.join(PRODUCTS_PHOTO_DIR, f"S{module_code}")
    if not os.path.exists(folder):
        return []
    photos = []
    for f in os.listdir(folder):
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            photos.append(os.path.join(folder, f))
    return photos


def load_products() -> List[Dict]:
    if not os.path.exists(PRODUCTS_FILE):
        return []
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    products = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            products.append(dict(zip(headers, row)))
    return products


def find_product_by_code(code: str) -> Optional[Dict]:
    products = load_products()
    for p in products:
        if p.get('Код продукта') == code:
            return p
    return None


def get_product_photos(product_code: str) -> List[str]:
    folder = os.path.join(PRODUCTS_PHOTO_DIR, product_code)
    if not os.path.exists(folder):
        return []
    photos = []
    for f in os.listdir(folder):
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            photos.append(os.path.join(folder, f))
    return photos


def load_materials() -> List[Dict]:
    if not os.path.exists(MATERIALS_FILE):
        return []
    wb = load_workbook(MATERIALS_FILE)
    ws = wb.active
    materials = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            materials.append(dict(zip(headers, row)))
    return materials


def get_material_types() -> List[str]:
    materials = load_materials()
    types = set(m.get('Тип материала') for m in materials if m.get('Тип материала'))
    return list(types)


def find_material_by_code(code: str) -> Optional[Dict]:
    materials = load_materials()
    for m in materials:
        if m.get('Код материала') == code:
            return m
    return None


def get_materials_by_type(material_type: str) -> List[Dict]:
    materials = load_materials()
    return [m for m in materials if m.get('Тип материала') == material_type]


def get_material_photos(material_code: str) -> List[str]:
    folder = os.path.join(MATERIALS_PHOTO_DIR, material_code)
    if not os.path.exists(folder):
        return []
    photos = []
    for f in os.listdir(folder):
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            photos.append(os.path.join(folder, f))
    return photos


def is_material_available_for_product(product_code: str, material_code: str) -> bool:
    product = find_product_by_code(product_code)
    if not product:
        return False
    material = find_material_by_code(material_code)
    if not material:
        return False
    material_type = material.get('Тип материала')
    if not material_type:
        return False
    price_val = product.get(material_type)
    if price_val and str(price_val).strip():
        try:
            price = float(str(price_val).replace(' ', '').replace(',', '.'))
            return price > 0
        except ValueError:
            return False
    return False


def get_available_material_types_for_product(product_code: str) -> List[str]:
    product = find_product_by_code(product_code)
    if not product:
        return []
    available = []
    for col in MATERIAL_COLUMNS:
        price_val = product.get(col)
        if price_val and str(price_val).strip():
            try:
                price = float(str(price_val).replace(' ', '').replace(',', '.'))
                if price > 0:
                    available.append(col)
            except ValueError:
                pass
    return available


def get_available_sofa_types() -> List[str]:
    if not os.path.exists(CONSTRUCTION_PHOTO_DIR):
        return []
    types = []
    for item in os.listdir(CONSTRUCTION_PHOTO_DIR):
        if os.path.isdir(os.path.join(CONSTRUCTION_PHOTO_DIR, item)):
            types.append(item)
    return types


def get_base_modules(sofa_type: str) -> List[Dict]:
    base_codes = []
    if os.path.exists(CONSTRUCTION_PHOTO_DIR):
        for folder in os.listdir(CONSTRUCTION_PHOTO_DIR):
            if folder.startswith(f"C{sofa_type}") and len(folder) >= 3:
                code = folder[1:]
                base_codes.append(code)
    modules = load_modules()
    return [m for m in modules if m.get('code') in base_codes]


def get_available_extensions(current_modules: List[str]) -> List[Dict]:
    all_modules = load_modules()
    existing_codes = set(current_modules)
    extensions = [m for m in all_modules if m.get('code') not in existing_codes]
    return extensions


def calculate_sofa_prices(module_codes: List[str]) -> Dict[str, float]:
    materials = load_materials()
    prices = {}
    for mat in materials:
        mat_type = mat.get('type')
        if mat_type:
            prices[mat_type] = len(module_codes) * 1000
    return prices


def generate_sofa_code(module_codes: List[str]) -> str:
    return "-".join(sorted(module_codes))


def parse_sofa_code(sofa_code: str) -> List[str]:
    return sofa_code.split("-")


def save_sofa(user_id: int, sofa_code: str, module_codes: List[str], material_prices: Dict[str, float]):
    if not os.path.exists(SOFAS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sofas"
        headers = ["user_id", "sofa_code", "modules", "material_prices", "date_saved"]
        ws.append(headers)
    else:
        wb = load_workbook(SOFAS_FILE)
        ws = wb.active
    modules_str = ",".join(module_codes)
    prices_str = ";".join([f"{k}:{v}" for k, v in material_prices.items()])
    ws.append([user_id, sofa_code, modules_str, prices_str, get_current_timestamp()])
    wb.save(SOFAS_FILE)


def get_user_sofas(user_id: int) -> List[Dict]:
    if not os.path.exists(SOFAS_FILE):
        return []
    wb = load_workbook(SOFAS_FILE)
    ws = wb.active
    sofas = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            record = dict(zip(headers, row))
            if record.get('modules'):
                record['modules'] = record['modules'].split(',')
            if record.get('material_prices'):
                prices = {}
                for item in record['material_prices'].split(';'):
                    if ':' in item:
                        k, v = item.split(':', 1)
                        prices[k] = float(v)
                record['material_prices'] = prices
            sofas.append(record)
    return sofas


def get_sofa_schema(sofa_code: str) -> Optional[str]:
    folder = os.path.join(CONSTRUCTION_PHOTO_DIR, sofa_code)
    schema_path = os.path.join(folder, 'cells.png')
    if os.path.exists(schema_path):
        return schema_path
    return None


# ================== НОВЫЕ ФУНКЦИИ ДЛЯ КОНСТРУИРОВАНИЯ ДИВАНОВ ==================
def split_sofa_code(sofa_code: str) -> Tuple[str, List[str]]:
    """Разбивает код дивана на префикс (CA/CP/CT) и список двузначных суффиксов."""
    if len(sofa_code) < 4 or sofa_code[0] != 'C':
        return '', []
    prefix = sofa_code[:2]  # CA, CP, CT
    if prefix not in PREFIX_TO_MODULE_LETTER:
        return '', []
    rest = sofa_code[2:]
    if len(rest) % 2 != 0:
        return '', []
    suffixes = [rest[i:i+2] for i in range(0, len(rest), 2)]
    return prefix, suffixes


def get_module_codes_from_sofa_code(sofa_code: str) -> List[str]:
    """Возвращает список кодов модулей (например, ['A02','A01']) для данного дивана."""
    prefix, suffixes = split_sofa_code(sofa_code)
    if not prefix:
        return []
    letter = PREFIX_TO_MODULE_LETTER.get(prefix)
    if not letter:
        return []
    return [f"{letter}{suffix}" for suffix in suffixes]


def get_sofa_type_name_from_code(sofa_code: str) -> str:
    """Возвращает отображаемое имя типа дивана по коду."""
    prefix = sofa_code[:2] if len(sofa_code) >= 2 else ''
    for name, p in SOFA_TYPES:
        if p == prefix:
            return name
    return ''


def get_sofa_total_price(sofa_code: str, material_type: str) -> Optional[float]:
    """Возвращает общую стоимость дивана в указанном материале (сумма модулей)."""
    module_codes = get_module_codes_from_sofa_code(sofa_code)
    if not module_codes:
        return None
    total = 0.0
    for mod_code in module_codes:
        module = find_module_by_code(mod_code)
        if not module:
            return None
        price_val = module.get(material_type)
        if not price_val or str(price_val).strip() == '':
            return None
        try:
            price = float(str(price_val).replace(' ', '').replace(',', '.'))
            if price <= 0:
                return None
            total += price
        except ValueError:
            return None
    return total


def get_available_material_types_for_sofa(sofa_code: str) -> List[str]:
    """Возвращает список типов материалов, доступных для всех модулей дивана."""
    module_codes = get_module_codes_from_sofa_code(sofa_code)
    if not module_codes:
        return []
    available = []
    for mat_type in MATERIAL_COLUMNS:
        ok = True
        for mod_code in module_codes:
            module = find_module_by_code(mod_code)
            if not module:
                ok = False
                break
            price_val = module.get(mat_type)
            if not price_val or str(price_val).strip() == '':
                ok = False
                break
            try:
                price = float(str(price_val).replace(' ', '').replace(',', '.'))
                if price <= 0:
                    ok = False
                    break
            except ValueError:
                ok = False
                break
        if ok:
            available.append(mat_type)
    return available


def get_sofa_name(sofa_code: str) -> str:
    """Генерирует название для сконструированного дивана."""
    return f"Сконструированный диван {sofa_code}"


def get_sofa_photo(sofa_code: str) -> Optional[str]:
    """Возвращает путь к схеме дивана (cells.png)."""
    folder = os.path.join(CONSTRUCTION_PHOTO_DIR, sofa_code)
    schema_path = os.path.join(folder, 'cells.png')
    if os.path.exists(schema_path):
        return schema_path
    return None


def display_sofa_card(vk, user_id, sofa_code: str, keyboard=None):
    """Отправляет пользователю карточку дивана: схема, список модулей, цены."""
    module_codes = get_module_codes_from_sofa_code(sofa_code)
    if not module_codes:
        send_message(vk, user_id, "❌ Неверный код дивана.")
        return

    # Получаем названия модулей
    module_names = []
    for mod_code in module_codes:
        mod = find_module_by_code(mod_code)
        if mod:
            module_names.append(mod.get('Название модуля', mod_code))
        else:
            module_names.append(mod_code)

    # Формируем сообщение
    type_name = get_sofa_type_name_from_code(sofa_code)
    msg = f"🛋️ **Сконструированный диван**\n"
    msg += f"📋 Код: {sofa_code}\n"
    if type_name:
        msg += f"📌 Тип: {type_name}\n"
    msg += f"🧩 Модули: {', '.join(module_names)}\n\n"
    msg += f"💰 **Стоимость в разных материалах:**\n"

    # Цены по материалам
    available_materials = get_available_material_types_for_sofa(sofa_code)
    if not available_materials:
        msg += "❌ Нет доступных материалов для этого набора модулей.\n"
    else:
        for mat_type in available_materials:
            price = get_sofa_total_price(sofa_code, mat_type)
            if price is not None:
                msg += f"• {mat_type}: {price:,.0f} руб.\n".replace(',', ' ')
            else:
                msg += f"• {mat_type}: цена не определена\n"

    # Отправляем фото схемы
    photo_path = get_sofa_photo(sofa_code)
    attachment = None
    if photo_path:
        attachments = upload_photos(vk, [photo_path])
        if attachments:
            attachment = ','.join(attachments)
    send_message(vk, user_id, msg, keyboard=keyboard, attachment=attachment)
    if not photo_path:
        send_message(vk, user_id, "⚠️ Схема для этого дивана не найдена.", keyboard=keyboard)


def sofa_code_exists(sofa_code: str) -> bool:
    """Проверяет, существует ли папка с таким кодом в Photo/Construction."""
    folder = os.path.join(CONSTRUCTION_PHOTO_DIR, sofa_code)
    return os.path.isdir(folder)


def get_base_modules_for_type(type_name: str) -> List[Dict]:
    """Возвращает список базовых модулей (с длиной кода 4) для выбранного типа дивана."""
    prefix = None
    for name, p in SOFA_TYPES:
        if name == type_name:
            prefix = p
            break
    if not prefix:
        return []

    # Ищем все папки с префиксом и длиной ровно 4 (базовые модули)
    base_folders = []
    if os.path.exists(CONSTRUCTION_PHOTO_DIR):
        for folder in os.listdir(CONSTRUCTION_PHOTO_DIR):
            if folder.startswith(prefix) and len(folder) == 4:
                base_folders.append(folder)

    # Извлекаем суффиксы и ищем модули в Moduls_log
    letter = PREFIX_TO_MODULE_LETTER[prefix]
    modules = []
    for folder in base_folders:
        suffix = folder[2:]  # например, "02"
        mod_code = f"{letter}{suffix}"
        module = find_module_by_code(mod_code)
        if module:
            modules.append(module)
    return modules


def get_extension_modules(current_sofa_code: str) -> List[Dict]:
    """Возвращает список модулей, которые можно добавить к текущему дивану."""
    prefix, suffixes = split_sofa_code(current_sofa_code)
    if not prefix:
        return []
    letter = PREFIX_TO_MODULE_LETTER[prefix]

    # Ищем все папки, которые начинаются с current_sofa_code и длиннее на 2 символа
    extensions = set()
    if os.path.exists(CONSTRUCTION_PHOTO_DIR):
        for folder in os.listdir(CONSTRUCTION_PHOTO_DIR):
            if folder.startswith(current_sofa_code) and len(folder) == len(current_sofa_code) + 2:
                next_suffix = folder[len(current_sofa_code):]  # два символа
                extensions.add(next_suffix)

    # Преобразуем в модули
    result = []
    for suffix in extensions:
        mod_code = f"{letter}{suffix}"
        module = find_module_by_code(mod_code)
        if module:
            result.append(module)
    return result


def start_sofa_construction(vk, user_id):
    """Начинает процесс конструирования: выбор типа дивана."""
    # Показываем кнопки с типами
    keyboard = VkKeyboard(one_time=False)
    for type_name, _ in SOFA_TYPES:
        keyboard.add_button(type_name, color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
    keyboard.add_button("Ввести код дивана-конструкта", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    user_states[user_id] = {
        'state': 'selecting_sofa_type',
        'data': {}
    }
    send_message(vk, user_id, "Выберите тип дивана для конструирования:", keyboard.get_keyboard())


def handle_selecting_sofa_type(vk, user_id, text, data):
    if text == "Назад":
        handle_start(vk, user_id)
        return
    if text == "Ввести код дивана-конструкта":
        user_states[user_id] = {
            'state': 'entering_sofa_code',
            'data': {}
        }
        send_message(vk, user_id, "Введите код дивана (начинается с C, например CA02 или CA0201):",
                     create_back_keyboard())
        return

    # Проверяем, что выбранный тип существует
    type_name = text
    prefix = None
    for name, p in SOFA_TYPES:
        if name == type_name:
            prefix = p
            break
    if not prefix:
        send_message(vk, user_id, "Пожалуйста, выберите тип из списка.")
        return

    # Получаем базовые модули для этого типа
    modules = get_base_modules_for_type(type_name)
    if not modules:
        send_message(vk, user_id, f"❌ Для типа '{type_name}' не найдено ни одного базового модуля.")
        handle_start(vk, user_id)
        return

    # Сохраняем данные и показываем список модулей
    user_states[user_id] = {
        'state': 'selecting_sofa_module',
        'data': {
            'type_name': type_name,
            'prefix': prefix,
            'modules': modules
        }
    }
    keyboard = VkKeyboard(one_time=False)
    for mod in modules:
        mod_name = mod.get('Название модуля', mod.get('Код модуля', '???'))
        keyboard.add_button(mod_name, color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    send_message(vk, user_id, f"Выберите начальный модуль для дивана типа '{type_name}':",
                 keyboard.get_keyboard())


def handle_selecting_sofa_module(vk, user_id, text, data):
    if text == "Назад":
        start_sofa_construction(vk, user_id)
        return

    # Ищем выбранный модуль по названию
    selected_module = None
    for mod in data['modules']:
        mod_name = mod.get('Название модуля', mod.get('Код модуля', ''))
        if mod_name == text:
            selected_module = mod
            break
    if not selected_module:
        send_message(vk, user_id, "Пожалуйста, выберите модуль из списка.")
        return

    # Формируем код дивана: prefix + суффикс из кода модуля (без буквы)
    mod_code = selected_module.get('Код модуля', '')
    if not mod_code:
        send_message(vk, user_id, "Ошибка: у модуля нет кода.")
        return
    # Из кода модуля (например, A02) извлекаем числовой суффикс
    suffix = mod_code[1:]  # отрезаем первую букву
    sofa_code = data['prefix'] + suffix

    # Переходим в режим конструирования
    user_states[user_id] = {
        'state': 'constructing_sofa',
        'data': {
            'sofa_code': sofa_code,
            'type_name': data['type_name'],
            'prefix': data['prefix']
        }
    }
    show_construction_state(vk, user_id)


def show_construction_state(vk, user_id):
    """Отображает текущее состояние сконструированного дивана и доступные действия."""
    state_data = user_states[user_id]['data']
    sofa_code = state_data['sofa_code']
    type_name = state_data['type_name']
    prefix = state_data['prefix']

    # Проверяем, существует ли папка с таким кодом (должна, но на всякий случай)
    if not sofa_code_exists(sofa_code):
        send_message(vk, user_id, f"❌ Ошибка: схема для дивана {sofa_code} не найдена.")
        handle_start(vk, user_id)
        return

    # Отправляем карточку дивана
    display_sofa_card(vk, user_id, sofa_code)

    # Получаем возможные расширения
    extensions = get_extension_modules(sofa_code)

    # Формируем клавиатуру
    keyboard = VkKeyboard(one_time=False)
    if extensions:
        for ext in extensions:
            ext_name = ext.get('Название модуля', ext.get('Код модуля', '???'))
            keyboard.add_button(ext_name, color=VkKeyboardColor.PRIMARY)
            keyboard.add_line()
    else:
        send_message(vk, user_id, "⚠️ Нет доступных модулей для добавления.")

    # Кнопки управления
    if len(sofa_code) > 4:  # если есть хотя бы один добавленный модуль сверх базы
        keyboard.add_button("Удалить последний модуль", color=VkKeyboardColor.SECONDARY)
        keyboard.add_line()
    keyboard.add_button("Добавить в корзину", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Ввести код дивана-конструкта", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад (выбор типа)", color=VkKeyboardColor.SECONDARY)
    keyboard.add_line()
    keyboard.add_button("Главное меню", color=VkKeyboardColor.NEGATIVE)

    send_message(vk, user_id, "Выберите действие:", keyboard.get_keyboard())


def handle_constructing_sofa(vk, user_id, text, data):
    sofa_code = data['sofa_code']
    prefix = data['prefix']
    type_name = data['type_name']

    # Проверка на выбор модуля расширения
    extensions = get_extension_modules(sofa_code)
    for ext in extensions:
        ext_name = ext.get('Название модуля', ext.get('Код модуля', ''))
        if text == ext_name:
            # Добавляем модуль
            mod_code = ext.get('Код модуля', '')
            suffix = mod_code[1:]  # отрезаем букву
            new_sofa_code = sofa_code + suffix
            data['sofa_code'] = new_sofa_code
            user_states[user_id]['data'] = data
            show_construction_state(vk, user_id)
            return

    if text == "Удалить последний модуль":
        if len(sofa_code) > 4:
            new_sofa_code = sofa_code[:-2]
            data['sofa_code'] = new_sofa_code
            user_states[user_id]['data'] = data
            show_construction_state(vk, user_id)
        else:
            send_message(vk, user_id, "Нельзя удалить единственный модуль. Выберите другой начальный модуль.")
        return

    elif text == "Добавить в корзину":
        # Переходим к выбору материала для этого дивана
        start_sofa_add_to_cart(vk, user_id, sofa_code)
        return

    elif text == "Ввести код дивана-конструкта":
        user_states[user_id] = {
            'state': 'entering_sofa_code',
            'data': {}
        }
        send_message(vk, user_id, "Введите код дивана (начинается с C, например CA02 или CA0201):",
                     create_back_keyboard())
        return

    elif text == "Назад (выбор типа)":
        start_sofa_construction(vk, user_id)
        return

    elif text == "Главное меню":
        handle_start(vk, user_id)
        return

    else:
        send_message(vk, user_id, "Используйте кнопки меню.")
        show_construction_state(vk, user_id)


def start_sofa_add_to_cart(vk, user_id, sofa_code):
    """Начинает процесс добавления дивана в корзину: выбор материала."""
    available_types = get_available_material_types_for_sofa(sofa_code)
    if not available_types:
        send_message(vk, user_id, "❌ Для этого дивана нет доступных материалов.")
        show_construction_state(vk, user_id)
        return

    # Сохраняем данные и переходим к выбору материала
    user_states[user_id] = {
        'state': 'choosing_material_method',
        'data': {
            'is_sofa': True,
            'sofa_code': sofa_code,
            'sofa_name': get_sofa_name(sofa_code),
            'available_types': available_types,
            'prev_state': 'constructing_sofa',
            'prev_data': user_states[user_id]['data']
        }
    }
    send_message(vk, user_id, "Выберите способ выбора материала для дивана:", create_material_method_keyboard())


def handle_entering_sofa_code(vk, user_id, text):
    if text == "Назад":
        start_sofa_construction(vk, user_id)
        return

    sofa_code = text.upper().strip()
    if not sofa_code.startswith('C') or len(sofa_code) < 4:
        send_message(vk, user_id, "❌ Неверный формат. Код должен начинаться с C и содержать не менее 4 символов (например, CA02).")
        return

    if not sofa_code_exists(sofa_code):
        send_message(vk, user_id, f"❌ Диван с кодом {sofa_code} не найден. Проверьте правильность кода.")
        return

    # Проверяем, что все модули существуют в Moduls_log
    module_codes = get_module_codes_from_sofa_code(sofa_code)
    valid = True
    for mod_code in module_codes:
        if not find_module_by_code(mod_code):
            valid = False
            break
    if not valid:
        send_message(vk, user_id, "❌ Диван содержит неизвестные модули. Обратитесь к менеджеру.")
        return

    # Переходим в режим конструирования с этим кодом
    prefix = sofa_code[:2]
    type_name = get_sofa_type_name_from_code(sofa_code)
    if not type_name:
        send_message(vk, user_id, "❌ Неизвестный тип дивана.")
        return

    user_states[user_id] = {
        'state': 'constructing_sofa',
        'data': {
            'sofa_code': sofa_code,
            'type_name': type_name,
            'prefix': prefix
        }
    }
    show_construction_state(vk, user_id)


# ================== РАБОТА С ЗАКАЗАМИ ==================
def create_orders_table():
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        headers = [
            "Порядковый номер позиции", "Номер задачи", "Код заказа", "ФИО Заказчика",
            "Телефон", "Адрес", "Код чата", "Ссылка", "Код позиции", "Материал",
            "Цвет материала", "Стоимость", "Дата создания заказа", "Статус заказа",
            "Комментарий", "", "Код Швеи", "ФИО Швеи", "Готовность пошива",
            "Код Столяра", "ФИО Столяра", "Готовность Каркаса", "Код Обтяжчика",
            "ФИО Обтяжчика", "Готовность продукта", "Код упаковщика", "ФИО Упаковщика",
            "Готовность позиции"
        ]
        ws.append(headers)
        wb.save(ORDERS_FILE)


def save_order_from_cart(user_id: int, order_data: Dict):
    create_orders_table()
    cart_items = get_user_cart(user_id)
    if not cart_items:
        return False

    order_code = generate_unique_order_code()
    task_number = get_next_task_number()
    timestamp = get_current_timestamp_display()
    user_link = f"https://vk.com/id{user_id}"

    wb = load_workbook(ORDERS_FILE)
    ws = wb.active

    position = 1
    for item in cart_items:
        row = [
            position,
            task_number,
            order_code,
            order_data['name'][:50],
            order_data['phone'][:20],
            order_data['address'][:100],
            user_id,
            user_link,
            item['Код позиции'],
            item['Код материала'],
            item['Цвет материала'],
            item['Цена'],
            timestamp,
            "Ожидает подтверждения",
            order_data['comment'][:500],
            "",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет",
            "нет"
        ]
        ws.append(row)
        position += 1

    wb.save(ORDERS_FILE)
    clear_user_cart(user_id)
    return order_code


# ================== НОВЫЕ ФУНКЦИИ ДЛЯ РАБОТЫ С ЗАКАЗАМИ ==================
def get_products_dict() -> Dict[str, str]:
    products = load_products()
    return {p.get('Код продукта', ''): p.get('Название продукта', '') for p in products}


def get_materials_dict() -> Dict[str, Dict]:
    materials = load_materials()
    return {m.get('Код материала', ''): {'Тип материала': m.get('Тип материала', ''), 'Цвет материала': m.get('Цвет материала', '')} for m in materials}


def get_user_orders_grouped(user_id: int) -> List[Dict]:
    if not os.path.exists(ORDERS_FILE):
        return []
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    orders_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6] == user_id:  # Код чата
            order_code = row[2]
            if order_code not in orders_dict:
                orders_dict[order_code] = {
                    'order_code': order_code,
                    'task_number': row[1],
                    'date': row[12],
                    'status': row[13],
                    'items': []
                }
            item = {
                'position_number': row[0],
                'product_code': row[8],
                'material_code': row[9],
                'material_color': row[10],
                'price': row[11],
                'comment': row[14]
            }
            orders_dict[order_code]['items'].append(item)
    return list(orders_dict.values())


def get_order_by_code(order_code: str) -> Optional[Dict]:
    if not os.path.exists(ORDERS_FILE):
        return None
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == order_code:
            return {
                'order_code': order_code,
                'task_number': row[1],
                'date': row[12],
                'status': row[13],
                'comment': row[14]
            }
    return None


def get_order_items(order_code: str) -> List[Dict]:
    if not os.path.exists(ORDERS_FILE):
        return []
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == order_code:
            items.append({
                'position_number': row[0],
                'product_code': row[8],
                'material_code': row[9],
                'material_color': row[10],
                'price': row[11],
                'comment': row[14]
            })
    return items


# ================== ФУНКЦИИ ОТПРАВКИ ==================
def send_message(vk, user_id, text, keyboard=None, attachment=None):
    params = {
        'user_id': user_id,
        'message': text,
        'random_id': get_random_id(),
    }
    if keyboard is not None:
        params['keyboard'] = keyboard
    if attachment:
        params['attachment'] = attachment
    vk.messages.send(**params)


def upload_photos(vk, photo_paths: List[str]) -> List[str]:
    attachments = []
    upload = vk_api.VkUpload(vk)
    for path in photo_paths[:10]:
        try:
            photo = upload.photo_messages(path)[0]
            attachments.append(f"photo{photo['owner_id']}_{photo['id']}")
        except Exception as e:
            print(f"Ошибка загрузки фото {path}: {e}")
    return attachments


def send_media_group(vk, user_id, photos):
    if not photos:
        return
    attachments = upload_photos(vk, photos)
    if attachments:
        vk.messages.send(
            user_id=user_id,
            random_id=get_random_id(),
            attachment=','.join(attachments)
        )


def send_sofa_schema(vk, user_id, sofa_code):
    schema_path = get_sofa_schema(sofa_code)
    if schema_path:
        send_media_group(vk, user_id, [schema_path])
    else:
        send_message(vk, user_id, "Схема для этого дивана не найдена.")


def send_material_photos(vk, user_id, material_code):
    photos = get_material_photos(material_code)
    if photos:
        send_media_group(vk, user_id, photos)
    else:
        send_message(vk, user_id, "Фотографии материала не найдены.")


def go_to_site(vk, user_id):
    send_message(vk, user_id, "Перейдите на наш сайт: https://lutohouse.ru/")


def send_manager_contact(vk, user_id):
    if MANAGER_IDS:
        manager_id = random.choice(MANAGER_IDS)
        send_message(vk, user_id, f"Свяжитесь с менеджером: https://vk.com/id{manager_id}")
    else:
        send_message(vk, user_id, "К сожалению, менеджеры недоступны.")


# ================== КЛАВИАТУРЫ ==================
def create_main_keyboard():
    keyboard = VkKeyboard(one_time=False)
    buttons = [
        "Сконструировать диван",
        "Корзина",
        "Добавить позицию к заказу",
        "Оформить заказ",
        "Мои заказы",
        "Проверить статус заказа",
        "Перейти на сайт",
        "Связаться с менеджером",
        "Перезапустить бота"
    ]
    for i, btn in enumerate(buttons):
        if btn == "Перейти на сайт":
            keyboard.add_openlink_button(btn, link="https://lutohouse.ru/")
        else:
            keyboard.add_button(btn, color=VkKeyboardColor.PRIMARY)
        if i % 2 == 1 and i != len(buttons) - 1:
            keyboard.add_line()
    return keyboard.get_keyboard()


def create_back_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_adding_item_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Ввести код продукта", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Выбрать из списка", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Ввести код дивана-конструкта", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_product_card_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("В корзину", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Предыдущий", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Следующий", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    keyboard.add_button("Главное меню", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()


def create_product_code_card_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("В корзину", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Ввести код заново", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_material_method_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Ввести код материала", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Выбрать из списка", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_enter_material_code_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_material_code_card_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Выбрать этот материал", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Ввести код заново", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    keyboard.add_line()
    keyboard.add_button("Главное меню", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()


def create_material_type_keyboard(types: List[str]):
    keyboard = VkKeyboard(one_time=False)
    for t in types:
        keyboard.add_button(t, color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_material_card_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Выбрать этот материал", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Предыдущий", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Следующий", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    keyboard.add_button("Главное меню", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()


def create_confirm_add_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Добавить в корзину", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Отмена", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_post_add_keyboard():
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Оформить заказ", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Перейти в корзину", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("В Главное меню", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()


def create_cart_options_keyboard(has_items: bool):
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("Добавить позицию к заказу", color=VkKeyboardColor.PRIMARY)
    if has_items:
        keyboard.add_line()
        keyboard.add_button("Отправить заказ на обработку", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Главное меню", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()


def create_order_collection_keyboard(cancel_text="Отменить"):
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button(cancel_text, color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()


def create_skip_comment_keyboard():
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Пропустить",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "order_skip_comment"}
    )
    return keyboard.get_keyboard()


def create_order_confirmation_keyboard():
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Подтвердить",
        color=VkKeyboardColor.POSITIVE,
        payload={"type": "order_confirm"}
    )
    keyboard.add_line()
    keyboard.add_callback_button(
        label="Ввести данные заново",
        color=VkKeyboardColor.NEGATIVE,
        payload={"type": "order_redo"}
    )
    return keyboard.get_keyboard()


def create_order_success_keyboard(order_code: str):
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Копировать код",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "order_copy_code", "code": order_code}
    )
    return keyboard.get_keyboard()


def create_order_inline_keyboard(order_code: str):
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Подробнее",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "order_detail", "order_code": order_code}
    )
    keyboard.add_line()
    keyboard.add_callback_button(
        label="Связаться с менеджером",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "order_contact_manager", "order_code": order_code}
    )
    return keyboard.get_keyboard()


def create_order_item_inline_keyboard(product_code: str, material_code: str):
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Подробнее",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "order_item_detail", "product_code": product_code, "material_code": material_code}
    )
    return keyboard.get_keyboard()


def create_order_detail_navigation_keyboard():
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Назад",
        color=VkKeyboardColor.SECONDARY,
        payload={"type": "order_detail_back"}
    )
    keyboard.add_line()
    keyboard.add_callback_button(
        label="Главное меню",
        color=VkKeyboardColor.NEGATIVE,
        payload={"type": "order_detail_main_menu"}
    )
    return keyboard.get_keyboard()


def create_cart_item_inline_keyboard(record_id: int) -> str:
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Подробнее",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "cart_detail", "record_id": record_id}
    )
    keyboard.add_line()
    keyboard.add_callback_button(
        label="Изменить материал",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "cart_change_material", "record_id": record_id}
    )
    keyboard.add_line()
    keyboard.add_callback_button(
        label="Удалить",
        color=VkKeyboardColor.NEGATIVE,
        payload={"type": "cart_remove", "record_id": record_id}
    )
    return keyboard.get_keyboard()


def create_back_to_cart_inline_keyboard() -> str:
    keyboard = VkKeyboard(inline=True)
    keyboard.add_callback_button(
        label="Вернуться в корзину",
        color=VkKeyboardColor.PRIMARY,
        payload={"type": "cart_back"}
    )
    return keyboard.get_keyboard()


# ================== РАБОТА С ТИПАМИ ПРОДУКТОВ ==================
def get_unique_product_types() -> List[str]:
    products = load_products()
    types = set()
    for p in products:
        t = p.get('Тип продукта')
        if t and isinstance(t, str):
            types.add(t.strip())
    return sorted(types)


def get_products_by_type(product_type: str) -> List[Dict]:
    products = load_products()
    return [p for p in products if p.get('Тип продукта') == product_type]


def get_product_price_for_material(product: Dict, material_type: str) -> Optional[float]:
    price_val = product.get(material_type)
    if price_val and str(price_val).strip():
        try:
            price_str = str(price_val).replace(' ', '').replace(',', '.')
            return float(price_str)
        except ValueError:
            return None
    return None


def display_product_card(vk, user_id, product: Dict, keyboard=None):
    msg = f"🔹 Код: {product.get('Код продукта', '')}\n"
    msg += f"📌 Название: {product.get('Название продукта', '')}\n"
    msg += f"🔗 Ссылка: {product.get('Ссылка', '')}\n"

    material_prices = []
    for mat_name in MATERIAL_COLUMNS:
        price_val = product.get(mat_name)
        if price_val and str(price_val).strip():
            try:
                price_str = str(price_val).replace(' ', '').replace(',', '.')
                price = float(price_str)
                if price > 0:
                    material_prices.append(f"{mat_name}: {price:,.0f} руб.".replace(',', ' '))
            except ValueError:
                pass

    if material_prices:
        msg += "\n💰 Стоимость в разных материалах:\n" + "\n".join(material_prices)

    product_code = product.get('Код продукта', '')
    photos = get_product_photos(product_code)
    attachments = upload_photos(vk, photos) if photos else []

    send_message(vk, user_id, msg, keyboard=keyboard, attachment=','.join(attachments) if attachments else None)
    if not photos:
        send_message(vk, user_id, "Фотографии для этого продукта не найдены.", keyboard=keyboard)


def display_material_card(vk, user_id, material: Dict, keyboard=None):
    msg = f"🔹 Код: {material.get('Код материала', '')}\n"
    msg += f"📌 Тип: {material.get('Тип материала', '')}\n"
    msg += f"🎨 Цвет: {material.get('Цвет материала', '')}\n"
    msg += f"🔗 Ссылка: {material.get('Ссылка', '')}"
    photos = get_material_photos(material.get('Код материала', ''))
    attachments = upload_photos(vk, photos) if photos else []
    send_message(vk, user_id, msg, keyboard=keyboard, attachment=','.join(attachments) if attachments else None)
    if not photos:
        send_message(vk, user_id, "Фотографии для этого материала не найдены.", keyboard=keyboard)


# ================== КОРЗИНА ==================
def show_cart(vk, user_id):
    cart_items = get_user_cart(user_id)
    if not cart_items:
        send_message(vk, user_id, "Ваша корзина пуста.")
        send_message(vk, user_id, "Выберите действие:", create_cart_options_keyboard(has_items=False))
        return

    for item in cart_items:
        if item['Тип позиции'] == 'sofa':
            # Показываем диван-конструкт
            text = f"🛋️ Код дивана: {item['Код позиции']}\n"
            text += f"📌 Название: {item['Название позиции']}\n"
            text += f"🧵 Код материала: {item['Код материала']}\n"
            text += f"📂 Тип материала: {item['Тип материала']}\n"
            text += f"🎨 Цвет: {item['Цвет материала']}\n"
            text += f"💰 Цена: {item['Цена']:,.0f} руб.\n".replace(',', ' ')
        else:
            text = f"📦 Код товара: {item['Код позиции']}\n"
            text += f"📌 Название: {item['Название позиции']}\n"
            text += f"🧵 Код материала: {item['Код материала']}\n"
            text += f"📂 Тип материала: {item['Тип материала']}\n"
            text += f"🎨 Цвет: {item['Цвет материала']}\n"
            text += f"💰 Цена: {item['Цена']:,.0f} руб.\n".replace(',', ' ')
        send_message(vk, user_id, text, keyboard=create_cart_item_inline_keyboard(item['ID записи']))

    send_message(vk, user_id, "Выберите действие:", create_cart_options_keyboard(has_items=True))


def show_cart_detail(vk, user_id, record_id: int):
    cart_items = get_user_cart(user_id)
    item = next((i for i in cart_items if i['ID записи'] == record_id), None)
    if not item:
        send_message(vk, user_id, "Позиция не найдена.")
        return

    if item['Тип позиции'] == 'sofa':
        sofa_code = item['Код позиции']
        display_sofa_card(vk, user_id, sofa_code, keyboard=create_back_to_cart_inline_keyboard())
    else:
        product = find_product_by_code(item['Код позиции'])
        if product:
            display_product_card(vk, user_id, product, keyboard=create_back_to_cart_inline_keyboard())
        else:
            send_message(vk, user_id, "Продукт не найден.", keyboard=create_back_to_cart_inline_keyboard())

    material = find_material_by_code(item['Код материала'])
    if material:
        display_material_card(vk, user_id, material, keyboard=create_back_to_cart_inline_keyboard())
    else:
        send_message(vk, user_id, "Материал не найден.", keyboard=create_back_to_cart_inline_keyboard())


def start_material_change(vk, user_id, record_id: int):
    cart_items = get_user_cart(user_id)
    item = next((i for i in cart_items if i['ID записи'] == record_id), None)
    if not item:
        send_message(vk, user_id, "Позиция не найдена.")
        return

    if item['Тип позиции'] == 'sofa':
        sofa_code = item['Код позиции']
        available_types = get_available_material_types_for_sofa(sofa_code)
        if not available_types:
            send_message(vk, user_id, "❌ Для этого дивана нет доступных материалов для смены.")
            show_cart(vk, user_id)
            return
        user_states[user_id] = {
            'state': 'changing_material_for_cart',
            'data': {
                'update_cart_item_id': record_id,
                'is_sofa': True,
                'sofa_code': sofa_code,
                'sofa_name': item['Название позиции'],
                'available_types': available_types,
                'current_material_code': item['Код материала'],
                'current_material_type': item['Тип материала'],
                'current_material_color': item['Цвет материала'],
                'current_price': item['Цена'],
                'prev_state': user_states.get(user_id, {}).get('state', 'main_menu'),
                'prev_data': user_states.get(user_id, {}).get('data', {})
            }
        }
    else:
        product = find_product_by_code(item['Код позиции'])
        if not product:
            send_message(vk, user_id, "Продукт не найден.")
            return
        user_states[user_id] = {
            'state': 'changing_material_for_cart',
            'data': {
                'update_cart_item_id': record_id,
                'is_sofa': False,
                'product': product,
                'product_code': item['Код позиции'],
                'product_name': item['Название позиции'],
                'current_material_code': item['Код материала'],
                'current_material_type': item['Тип материала'],
                'current_material_color': item['Цвет материала'],
                'current_price': item['Цена'],
                'prev_state': user_states.get(user_id, {}).get('state', 'main_menu'),
                'prev_data': user_states.get(user_id, {}).get('data', {})
            }
        }

    send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())


# ================== МОИ ЗАКАЗЫ ==================
def show_user_orders(vk, user_id):
    orders = get_user_orders_grouped(user_id)
    if not orders:
        send_message(vk, user_id, "У вас пока нет заказов.")
        return

    products_dict = get_products_dict()
    for order in orders:
        msg = f"📋 **Заказ №{order['order_code']}**\n"
        msg += f"📅 Дата: {order['date']}\n"
        msg += f"⚙️ Статус: {order['status']}\n\n"
        msg += "**Позиции:**\n"
        total = 0
        for item in order['items']:
            product_code = item['product_code']
            if product_code.startswith('C'):  # диван-конструкт
                product_name = get_sofa_name(product_code)
            else:
                product_name = products_dict.get(product_code, 'Неизвестно')
            material = find_material_by_code(item['material_code'])
            material_type = material.get('Тип материала', '') if material else ''
            material_color = item['material_color']
            price = item['price']
            total += price
            msg += f"• {product_name} ({product_code})\n"
            msg += f"  Материал: {material_type} - {material_color}\n"
            msg += f"  Цена: {price:,.0f} руб.\n\n".replace(',', ' ')
        msg += f"**Итого: {total:,.0f} руб.**".replace(',', ' ')
        send_message(vk, user_id, msg, keyboard=create_order_inline_keyboard(order['order_code']))


def show_order_items_list(vk, user_id, order_code: str):
    items = get_order_items(order_code)
    if not items:
        send_message(vk, user_id, "Заказ не найден.")
        return

    products_dict = get_products_dict()
    for item in items:
        product_code = item['product_code']
        if product_code.startswith('C'):
            product_name = get_sofa_name(product_code)
        else:
            product_name = products_dict.get(product_code, 'Неизвестно')
        material = find_material_by_code(item['material_code'])
        material_type = material.get('Тип материала', '') if material else ''
        material_color = item['material_color']
        price = item['price']

        text = f"📦 Код позиции: {product_code}\n"
        text += f"📌 Название: {product_name}\n"
        text += f"🧵 Код материала: {item['material_code']}\n"
        text += f"📂 Тип материала: {material_type}\n"
        text += f"🎨 Цвет: {material_color}\n"
        text += f"💰 Цена: {price:,.0f} руб.\n".replace(',', ' ')
        send_message(vk, user_id, text, keyboard=create_order_item_inline_keyboard(product_code, item['material_code']))

    send_message(vk, user_id, "Выберите действие:", keyboard=create_order_detail_navigation_keyboard())


def show_order_item_detail(vk, user_id, product_code: str, material_code: str):
    if product_code.startswith('C'):
        display_sofa_card(vk, user_id, product_code, keyboard=create_back_to_cart_inline_keyboard())
    else:
        product = find_product_by_code(product_code)
        if product:
            display_product_card(vk, user_id, product, keyboard=create_back_to_cart_inline_keyboard())
        else:
            send_message(vk, user_id, "Продукт не найден.", keyboard=create_back_to_cart_inline_keyboard())

    material = find_material_by_code(material_code)
    if material:
        display_material_card(vk, user_id, material, keyboard=create_back_to_cart_inline_keyboard())
    else:
        send_message(vk, user_id, "Материал не найден.", keyboard=create_back_to_cart_inline_keyboard())


# ================== ПРОВЕРКА СТАТУСА ЗАКАЗА ==================
def start_order_status_check(vk, user_id):
    user_states[user_id] = {
        'state': 'entering_order_code',
        'data': {}
    }
    send_message(vk, user_id, "Введите код заказа для проверки статуса:", create_back_keyboard())


def handle_entering_order_code(vk, user_id, text):
    if text == "Назад":
        handle_start(vk, user_id)
        return

    order = get_order_by_code(text)
    if order:
        msg = f"📋 **Заказ №{order['order_code']}**\n"
        msg += f"📅 Дата: {order['date']}\n"
        msg += f"⚙️ Статус: {order['status']}\n"
        if order['comment'] != 'Пусто':
            msg += f"💬 Комментарий: {order['comment']}\n"
        send_message(vk, user_id, msg, keyboard=create_main_keyboard())
        user_states.pop(user_id, None)
    else:
        send_message(vk, user_id, "❌ Заказ с таким кодом не найден. Попробуйте снова или нажмите 'Назад'.")


# ================== НАЧАЛО ОФОРМЛЕНИЯ ЗАКАЗА ==================
def start_order_collection(vk, user_id):
    cart_items = get_user_cart(user_id)
    if not cart_items:
        send_message(vk, user_id, "Ваша корзина пуста. Добавьте товары перед оформлением.")
        send_message(vk, user_id, "Главное меню:", create_main_keyboard())
        return

    user_states[user_id] = {
        'state': 'collecting_order_name',
        'data': {}
    }
    send_message(vk, user_id, "Введите ваше ФИО (полностью, не более 50 символов):",
                 create_order_collection_keyboard())


# ================== ОБРАБОТЧИКИ ==================
def handle_start(vk, user_id):
    user_states[user_id] = {'state': 'main_menu', 'data': {}}
    welcome_text = """
🛋️ Добро пожаловать в конструктор диванов и магазин!

Здесь вы можете:
• Сконструировать диван из модулей
• Сохранить свои диваны
• Выбрать готовые продукты
• Оформить заказ на конструкторские диваны и готовые продукты
• Отслеживать статус заказов
• Связаться с менеджером
    """.strip()
    send_message(vk, user_id, welcome_text, create_main_keyboard())


def handle_contact_manager(vk, user_id):
    send_message(vk, user_id, "✅ Запрос на связь с менеджером отправлен.")
    if MANAGER_IDS:
        user_info = vk.users.get(user_ids=user_id)[0]
        name = f"{user_info['first_name']} {user_info['last_name']}"
        link = f"https://vk.com/id{user_id}"
        msg = f"Пользователь {name} ({link}) хочет связаться с менеджером."
        vk.messages.send(
            peer_id=MANAGER_IDS[0],
            message=msg,
            random_id=get_random_id()
        )
    handle_start(vk, user_id)


def show_adding_item_menu(vk, user_id, from_state=None, from_data=None):
    user_states[user_id] = {
        'state': 'adding_item_menu',
        'data': {
            'prev_state': from_state,
            'prev_data': from_data
        }
    }
    send_message(vk, user_id, "Выберите способ добавления позиции:", create_adding_item_keyboard())


def handle_adding_item_menu(vk, user_id, text):
    state_data = user_states[user_id].get('data', {})
    if text == "Ввести код продукта":
        user_states[user_id] = {'state': 'entering_product_code', 'data': state_data}
        send_message(vk, user_id, "Введите код продукта:", create_back_keyboard())
    elif text == "Выбрать из списка":
        types = get_unique_product_types()
        if not types:
            send_message(vk, user_id, "Нет доступных типов продуктов.")
            show_adding_item_menu(vk, user_id)
            return
        keyboard = VkKeyboard(one_time=False)
        for t in types:
            keyboard.add_button(t, color=VkKeyboardColor.PRIMARY)
            keyboard.add_line()
        keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
        user_states[user_id] = {
            'state': 'selecting_product_type',
            'data': {
                'types': types,
                'prev_state': state_data.get('prev_state'),
                'prev_data': state_data.get('prev_data')
            }
        }
        send_message(vk, user_id, "Выберите тип продукта:", keyboard.get_keyboard())
    elif text == "Ввести код дивана-конструкта":
        user_states[user_id] = {
            'state': 'entering_sofa_code',
            'data': state_data
        }
        send_message(vk, user_id, "Введите код дивана (начинается с C, например CA02 или CA0201):",
                     create_back_keyboard())
    elif text == "Назад":
        prev_state = state_data.get('prev_state')
        prev_data = state_data.get('prev_data', {})
        if prev_state == 'order_menu':
            show_cart(vk, user_id)
        else:
            handle_start(vk, user_id)
    else:
        send_message(vk, user_id, "Используйте кнопки меню.", create_adding_item_keyboard())


def handle_selecting_product_type(vk, user_id, text, data):
    types = data.get('types', [])
    if text in types:
        products = get_products_by_type(text)
        if not products:
            send_message(vk, user_id, "Продукты данного типа не найдены.")
            show_adding_item_menu(vk, user_id)
            return
        user_states[user_id] = {
            'state': 'browsing_products',
            'data': {
                'product_list': products,
                'current_index': 0,
                'product_type': text,
                'prev_state': data.get('prev_state'),
                'prev_data': data.get('prev_data')
            }
        }
        send_message(vk, user_id, "Загрузка карточки товара. Подождите...")
        display_product_card(vk, user_id, products[0], create_product_card_keyboard())
    elif text == "Назад":
        show_adding_item_menu(vk, user_id)
    else:
        send_message(vk, user_id, "Пожалуйста, выберите тип из списка.")


def handle_browsing_products(vk, user_id, text, data):
    product_list = data.get('product_list', [])
    current_index = data.get('current_index', 0)

    if not product_list:
        send_message(vk, user_id, "Список продуктов пуст.")
        show_adding_item_menu(vk, user_id)
        return

    if text == "В корзину":
        product = product_list[current_index]
        user_states[user_id] = {
            'state': 'choosing_material_method',
            'data': {
                'is_sofa': False,
                'product_code': product.get('Код продукта'),
                'product': product,
                'prev_state': 'browsing_products',
                'prev_data': data
            }
        }
        send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
        return

    if text == "Следующий":
        current_index = (current_index + 1) % len(product_list)
    elif text == "Предыдущий":
        current_index = (current_index - 1) % len(product_list)
    elif text == "Назад":
        types = get_unique_product_types()
        keyboard = VkKeyboard(one_time=False)
        for t in types:
            keyboard.add_button(t, color=VkKeyboardColor.PRIMARY)
            keyboard.add_line()
        keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
        user_states[user_id] = {
            'state': 'selecting_product_type',
            'data': {
                'types': types,
                'prev_state': data.get('prev_state'),
                'prev_data': data.get('prev_data')
            }
        }
        send_message(vk, user_id, "Выберите тип продукта:", keyboard.get_keyboard())
        return
    elif text == "Главное меню":
        handle_start(vk, user_id)
        return
    else:
        send_message(vk, user_id, "Используйте кнопки навигации.", create_product_card_keyboard())
        return

    data['current_index'] = current_index
    user_states[user_id]['data'] = data
    send_message(vk, user_id, "Загрузка карточки товара. Подождите...")
    display_product_card(vk, user_id, product_list[current_index], create_product_card_keyboard())


def handle_entering_product_code(vk, user_id, text, data):
    if text == "Назад":
        show_adding_item_menu(vk, user_id)
        return

    # Проверяем, не является ли введённый код кодом дивана-конструкта
    if text.upper().startswith('C') and len(text) >= 4:
        # Перенаправляем на обработку кода дивана
        handle_entering_sofa_code(vk, user_id, text)
        return

    product = find_product_by_code(text)
    if product:
        send_message(vk, user_id, "Загрузка карточки товара. Подождите...")
        data['product_code'] = text
        user_states[user_id] = {'state': 'viewing_product_by_code', 'data': data}
        display_product_card(vk, user_id, product, create_product_code_card_keyboard())
    else:
        send_message(vk, user_id, "❌ Продукт с таким кодом не найден. Попробуйте снова:")


def handle_viewing_product_by_code(vk, user_id, text, data):
    product_code = data.get('product_code')
    product = find_product_by_code(product_code) if product_code else None
    if not product:
        send_message(vk, user_id, "Ошибка: товар не найден. Вернитесь в главное меню.")
        handle_start(vk, user_id)
        return

    if text == "В корзину":
        user_states[user_id] = {
            'state': 'choosing_material_method',
            'data': {
                'is_sofa': False,
                'product_code': product_code,
                'product': product,
                'prev_state': 'viewing_product_by_code',
                'prev_data': data
            }
        }
        send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
        return
    elif text == "Ввести код заново":
        user_states[user_id] = {'state': 'entering_product_code', 'data': data}
        send_message(vk, user_id, "Введите код продукта:", create_back_keyboard())
        return
    elif text == "Назад":
        show_adding_item_menu(vk, user_id)
        return
    else:
        send_message(vk, user_id, "Пожалуйста, используйте кнопки под карточкой товара.")
        display_product_card(vk, user_id, product, create_product_code_card_keyboard())


def handle_choosing_material_method(vk, user_id, text, data):
    if text == "Ввести код материала":
        user_states[user_id] = {
            'state': 'entering_material_code',
            'data': data
        }
        send_message(vk, user_id, "Введите код материала:", create_enter_material_code_keyboard())
    elif text == "Выбрать из списка":
        if data.get('is_sofa'):
            sofa_code = data['sofa_code']
            available_types = get_available_material_types_for_sofa(sofa_code)
            if not available_types:
                send_message(vk, user_id, "❌ Для данного дивана нет доступных материалов.")
                send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
                return
            user_states[user_id] = {
                'state': 'selecting_material_type',
                'data': {
                    **data,
                    'available_types': available_types
                }
            }
            keyboard = create_material_type_keyboard(available_types)
            send_message(vk, user_id, "Выберите тип материала:", keyboard)
        else:
            product_code = data.get('product_code')
            available_types = get_available_material_types_for_product(product_code)
            if not available_types:
                send_message(vk, user_id, "❌ Для данного товара нет доступных материалов.")
                send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
                return
            user_states[user_id] = {
                'state': 'selecting_material_type',
                'data': {
                    **data,
                    'available_types': available_types
                }
            }
            keyboard = create_material_type_keyboard(available_types)
            send_message(vk, user_id, "Выберите тип материала:", keyboard)
    elif text == "Назад":
        prev_state = data.get('prev_state')
        prev_data = data.get('prev_data', {})
        if prev_state == 'browsing_products':
            user_states[user_id] = {'state': prev_state, 'data': prev_data}
            product_list = prev_data.get('product_list', [])
            current_index = prev_data.get('current_index', 0)
            if product_list:
                send_message(vk, user_id, "Загрузка карточки товара. Подождите...")
                display_product_card(vk, user_id, product_list[current_index], create_product_card_keyboard())
            else:
                handle_start(vk, user_id)
        elif prev_state == 'viewing_product_by_code':
            product_code = prev_data.get('product_code')
            product = find_product_by_code(product_code)
            if product:
                user_states[user_id] = {'state': prev_state, 'data': prev_data}
                send_message(vk, user_id, "Загрузка карточки товара. Подождите...")
                display_product_card(vk, user_id, product, create_product_code_card_keyboard())
            else:
                handle_start(vk, user_id)
        elif prev_state == 'constructing_sofa':
            # Возврат к конструированию дивана
            user_states[user_id] = {'state': prev_state, 'data': prev_data}
            show_construction_state(vk, user_id)
        else:
            handle_start(vk, user_id)
    else:
        send_message(vk, user_id, "Используйте кнопки меню.", create_material_method_keyboard())


def handle_entering_material_code(vk, user_id, text, data):
    if text == "Назад":
        user_states[user_id] = {
            'state': 'choosing_material_method',
            'data': data
        }
        send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
        return

    material = find_material_by_code(text)
    if not material:
        send_message(vk, user_id, "❌ Материал с таким кодом не найден. Попробуйте снова:")
        return

    material_type = material.get('Тип материала')
    if data.get('is_sofa'):
        sofa_code = data['sofa_code']
        if material_type not in get_available_material_types_for_sofa(sofa_code):
            send_message(vk, user_id, f"❌ Материал типа '{material_type}' недоступен для этого дивана. Попробуйте другой код:")
            return
        price = get_sofa_total_price(sofa_code, material_type)
        if price is None:
            send_message(vk, user_id, "❌ Не удалось определить цену для данного материала.")
            return
    else:
        product_code = data.get('product_code')
        if not is_material_available_for_product(product_code, text):
            send_message(vk, user_id, "❌ Данный тип материала недоступен для выбранного товара. Попробуйте другой код:")
            return
        price = get_product_price_for_material(data['product'], material_type)
        if price is None:
            send_message(vk, user_id, "❌ Не удалось определить цену для данного материала.")
            return

    user_states[user_id] = {
        'state': 'viewing_material_by_code',
        'data': {
            **data,
            'material_code': text,
            'material': material,
            'price': price
        }
    }
    send_message(vk, user_id, "Загрузка карточки материала. Подождите...")
    display_material_card(vk, user_id, material, create_material_code_card_keyboard())


def handle_viewing_material_by_code(vk, user_id, text, data):
    material = data.get('material')
    if not material:
        handle_start(vk, user_id)
        return

    if text == "Выбрать этот материал":
        price = data['price']
        material_type = material.get('Тип материала')
        material_color = material.get('Цвет материала')

        update_item_id = data.get('update_cart_item_id')
        if update_item_id:
            update_cart_item(update_item_id, material.get('Код материала'), material_type,
                             material_color, price)
            send_message(vk, user_id, "✅ Материал успешно изменён!")
            show_cart(vk, user_id)
            return

        # Добавление в корзину
        if data.get('is_sofa'):
            sofa_code = data['sofa_code']
            sofa_name = get_sofa_name(sofa_code)
            save_to_cart_full("VK", user_id, "sofa", sofa_code, sofa_name,
                              material.get('Код материала'), material_type, material_color, price)
        else:
            product = data['product']
            save_to_cart_full("VK", user_id, "product", product.get('Код продукта'),
                              product.get('Название продукта'),
                              material.get('Код материала'), material_type, material_color, price)

        send_message(vk, user_id, "✅ Позиция добавлена в корзину!", create_post_add_keyboard())
        user_states[user_id] = {'state': 'post_add_menu', 'data': {}}
        return

    elif text == "Ввести код заново":
        user_states[user_id] = {
            'state': 'entering_material_code',
            'data': {k: v for k, v in data.items() if k not in ['material', 'price']}
        }
        send_message(vk, user_id, "Введите код материала:", create_enter_material_code_keyboard())
        return
    elif text == "Назад":
        user_states[user_id] = {
            'state': 'choosing_material_method',
            'data': {k: v for k, v in data.items() if k not in ['material', 'price']}
        }
        send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
        return
    elif text == "Главное меню":
        handle_start(vk, user_id)
        return
    else:
        display_material_card(vk, user_id, material, create_material_code_card_keyboard())


def handle_selecting_material_type(vk, user_id, text, data):
    available_types = data.get('available_types', [])
    if text in available_types:
        materials = get_materials_by_type(text)
        if not materials:
            send_message(vk, user_id, "❌ Материалы данного типа не найдены.")
            keyboard = create_material_type_keyboard(available_types)
            send_message(vk, user_id, "Выберите тип материала:", keyboard)
            return
        user_states[user_id] = {
            'state': 'browsing_materials',
            'data': {
                **data,
                'material_list': materials,
                'current_index': 0,
                'material_type': text
            }
        }
        send_message(vk, user_id, "Загрузка карточки материала. Подождите...")
        display_material_card(vk, user_id, materials[0], create_material_card_keyboard())
    elif text == "Назад":
        user_states[user_id] = {
            'state': 'choosing_material_method',
            'data': {k: v for k, v in data.items() if k not in ['available_types']}
        }
        send_message(vk, user_id, "Выберите способ выбора материала:", create_material_method_keyboard())
    else:
        send_message(vk, user_id, "Пожалуйста, выберите тип из списка.")


def handle_browsing_materials(vk, user_id, text, data):
    material_list = data.get('material_list', [])
    current_index = data.get('current_index', 0)

    if not material_list:
        send_message(vk, user_id, "Список материалов пуст.")
        available_types = data.get('available_types', [])
        keyboard = create_material_type_keyboard(available_types)
        user_states[user_id] = {
            'state': 'selecting_material_type',
            'data': {k: v for k, v in data.items() if k not in ['material_list', 'current_index']}
        }
        send_message(vk, user_id, "Выберите тип материала:", keyboard)
        return

    if text == "Выбрать этот материал":
        material = material_list[current_index]
        material_type = material.get('Тип материала')
        material_color = material.get('Цвет материала')

        if data.get('is_sofa'):
            sofa_code = data['sofa_code']
            price = get_sofa_total_price(sofa_code, material_type)
            if price is None:
                send_message(vk, user_id, "❌ Не удалось определить цену для данного материала.")
                display_material_card(vk, user_id, material, create_material_card_keyboard())
                return
        else:
            product = data['product']
            price = get_product_price_for_material(product, material_type)
            if price is None:
                send_message(vk, user_id, "❌ Не удалось определить цену для данного материала.")
                display_material_card(vk, user_id, material, create_material_card_keyboard())
                return

        update_item_id = data.get('update_cart_item_id')
        if update_item_id:
            update_cart_item(update_item_id, material.get('Код материала'), material_type,
                             material_color, price)
            send_message(vk, user_id, "✅ Материал успешно изменён!")
            show_cart(vk, user_id)
            return

        # Добавление в корзину
        if data.get('is_sofa'):
            sofa_code = data['sofa_code']
            sofa_name = get_sofa_name(sofa_code)
            save_to_cart_full("VK", user_id, "sofa", sofa_code, sofa_name,
                              material.get('Код материала'), material_type, material_color, price)
        else:
            product = data['product']
            save_to_cart_full("VK", user_id, "product", product.get('Код продукта'),
                              product.get('Название продукта'),
                              material.get('Код материала'), material_type, material_color, price)

        send_message(vk, user_id, "✅ Позиция добавлена в корзину!", create_post_add_keyboard())
        user_states[user_id] = {'state': 'post_add_menu', 'data': {}}
        return

    elif text == "Следующий":
        current_index = (current_index + 1) % len(material_list)
    elif text == "Предыдущий":
        current_index = (current_index - 1) % len(material_list)
    elif text == "Назад":
        available_types = data.get('available_types', [])
        keyboard = create_material_type_keyboard(available_types)
        user_states[user_id] = {
            'state': 'selecting_material_type',
            'data': {k: v for k, v in data.items() if k not in ['material_list', 'current_index']}
        }
        send_message(vk, user_id, "Выберите тип материала:", keyboard)
        return
    elif text == "Главное меню":
        handle_start(vk, user_id)
        return
    else:
        send_message(vk, user_id, "Используйте кнопки навигации.", create_material_card_keyboard())
        return

    data['current_index'] = current_index
    user_states[user_id]['data'] = data
    send_message(vk, user_id, "Загрузка карточки материала. Подождите...")
    display_material_card(vk, user_id, material_list[current_index], create_material_card_keyboard())


def handle_confirm_add_to_cart(vk, user_id, text, data):
    # Эта функция больше не используется, но оставлена для совместимости
    if text == "Добавить в корзину":
        if data.get('is_sofa'):
            sofa_code = data['sofa_code']
            sofa_name = get_sofa_name(sofa_code)
            save_to_cart_full("VK", user_id, "sofa", sofa_code, sofa_name,
                              data['material_code'], data['material_type'], data['material_color'], data['price'])
        else:
            save_to_cart_full("VK", user_id, "product", data['product_code'],
                              data['product_name'],
                              data['material_code'], data['material_type'], data['material_color'], data['price'])
        send_message(vk, user_id, "✅ Позиция добавлена в корзину!", create_post_add_keyboard())
        user_states[user_id] = {'state': 'post_add_menu', 'data': {}}
    elif text == "Отмена":
        handle_start(vk, user_id)
    else:
        send_message(vk, user_id, "Пожалуйста, выберите действие с помощью кнопок.")
        send_message(vk, user_id, "Хотите добавить эту позицию в корзину?", create_confirm_add_keyboard())


def handle_post_add_menu(vk, user_id, text):
    if text == "В Главное меню":
        handle_start(vk, user_id)
    elif text == "Оформить заказ":
        show_cart(vk, user_id)
    elif text == "Перейти в корзину":
        show_cart(vk, user_id)
    else:
        send_message(vk, user_id, "Пожалуйста, используйте кнопки меню.")
        send_message(vk, user_id, "Выберите действие:", create_post_add_keyboard())


def handle_main_menu(vk, user_id, text):
    if text == "Сконструировать диван":
        start_sofa_construction(vk, user_id)
    elif text == "Корзина" or text == "Оформить заказ":
        show_cart(vk, user_id)
    elif text == "Добавить позицию к заказу":
        show_adding_item_menu(vk, user_id, from_state='main_menu')
    elif text == "Отправить заказ на обработку":
        start_order_collection(vk, user_id)
    elif text == "Мои заказы":
        show_user_orders(vk, user_id)
    elif text == "Проверить статус заказа":
        start_order_status_check(vk, user_id)
    else:
        send_message(vk, user_id, "Используйте кнопки меню.", create_main_keyboard())


# ================== ОБРАБОТЧИКИ СБОРА ДАННЫХ ЗАКАЗА ==================
def handle_collecting_order_name(vk, user_id, text):
    if text == "Отменить":
        handle_start(vk, user_id)
        return
    if len(text) > 50:
        send_message(vk, user_id, "ФИО не должно превышать 50 символов. Введите снова:")
        return
    user_states[user_id]['data']['name'] = text
    user_states[user_id]['state'] = 'collecting_order_phone'
    send_message(vk, user_id, "Введите номер телефона (не более 20 символов):",
                 create_order_collection_keyboard())


def handle_collecting_order_phone(vk, user_id, text):
    if text == "Отменить":
        handle_start(vk, user_id)
        return
    if len(text) > 20:
        send_message(vk, user_id, "Номер телефона не должен превышать 20 символов. Введите снова:")
        return
    user_states[user_id]['data']['phone'] = text
    user_states[user_id]['state'] = 'collecting_order_address'
    send_message(vk, user_id, "Введите адрес доставки (не более 100 символов):",
                 create_order_collection_keyboard())


def handle_collecting_order_address(vk, user_id, text):
    if text == "Отменить":
        handle_start(vk, user_id)
        return
    if len(text) > 100:
        send_message(vk, user_id, "Адрес не должен превышать 100 символов. Введите снова:")
        return
    user_states[user_id]['data']['address'] = text
    user_states[user_id]['state'] = 'collecting_order_comment'
    send_message(vk, user_id, "Введите комментарий к заказу (не более 500 символов) или нажмите 'Пропустить':",
                 keyboard=create_order_collection_keyboard())
    send_message(vk, user_id, "Для пропуска комментария нажмите кнопку ниже:",
                 keyboard=create_skip_comment_keyboard())


def handle_collecting_order_comment(vk, user_id, text):
    if text == "Отменить":
        handle_start(vk, user_id)
        return
    if len(text) > 500:
        send_message(vk, user_id, "Комментарий не должен превышать 500 символов. Введите снова:")
        return
    user_states[user_id]['data']['comment'] = text
    show_order_confirmation(vk, user_id)


def handle_order_skip_comment(vk, user_id):
    if user_id not in user_states or user_states[user_id].get('state') != 'collecting_order_comment':
        return
    user_states[user_id]['data']['comment'] = "Пусто"
    show_order_confirmation(vk, user_id)
    try:
        vk.messages.sendMessageEventAnswer(
            event_id=None,
            user_id=user_id,
            peer_id=user_id
        )
    except:
        pass


def show_order_confirmation(vk, user_id):
    data = user_states[user_id]['data']
    text = "Проверьте введённые данные:\n"
    text += f"ФИО: {data['name']}\n"
    text += f"Телефон: {data['phone']}\n"
    text += f"Адрес: {data['address']}\n"
    text += f"Комментарий: {data['comment']}\n"
    text += "\nЕсли всё верно, нажмите 'Подтвердить'. Для изменения данных нажмите 'Ввести данные заново'."

    user_states[user_id]['state'] = 'confirm_order_details'
    send_message(vk, user_id, text, keyboard=create_order_confirmation_keyboard())


def handle_order_redo(vk, user_id):
    user_states[user_id] = {
        'state': 'collecting_order_name',
        'data': {}
    }
    send_message(vk, user_id, "Введите ваше ФИО (полностью, не более 50 символов):",
                 create_order_collection_keyboard())
    try:
        vk.messages.sendMessageEventAnswer(
            event_id=None,
            user_id=user_id,
            peer_id=user_id
        )
    except:
        pass


def handle_order_confirm(vk, user_id):
    data = user_states[user_id].get('data', {})
    if not data:
        handle_start(vk, user_id)
        return

    send_message(vk, user_id, "Отправляем заказ на обработку. Подождите...")
    order_code = save_order_from_cart(user_id, data)
    if order_code:
        user_states.pop(user_id, None)
        msg = f"✅ Заказ успешно оформлен!\n📋 Код заказа: {order_code}\n📞 С вами свяжутся для уточнения деталей."
        send_message(vk, user_id, msg, keyboard=create_order_success_keyboard(order_code))
        send_message(vk, user_id, "Главное меню:", create_main_keyboard())
    else:
        send_message(vk, user_id, "❌ Произошла ошибка при оформлении заказа. Попробуйте позже.")
        handle_start(vk, user_id)

    try:
        vk.messages.sendMessageEventAnswer(
            event_id=None,
            user_id=user_id,
            peer_id=user_id
        )
    except:
        pass


def handle_order_copy_code(vk, user_id, order_code):
    msg = f"Код заказа: {order_code}\nСкопируйте его, нажав и удерживая на тексте."
    send_message(vk, user_id, msg)
    send_message(vk, user_id, "Главное меню:", create_main_keyboard())
    try:
        vk.messages.sendMessageEventAnswer(
            event_id=None,
            user_id=user_id,
            peer_id=user_id
        )
    except:
        pass


# ================== ОБРАБОТЧИК CALLBACK ==================
def handle_callback(vk, event):
    user_id = event.user_id
    payload = event.payload
    if not payload:
        return

    action = payload.get('type')
    if action == 'cart_detail':
        record_id = payload.get('record_id')
        if record_id is not None:
            show_cart_detail(vk, user_id, record_id)
    elif action == 'cart_change_material':
        record_id = payload.get('record_id')
        if record_id is not None:
            start_material_change(vk, user_id, record_id)
    elif action == 'cart_remove':
        record_id = payload.get('record_id')
        if record_id is not None:
            remove_from_cart(record_id)
            send_message(vk, user_id, "✅ Позиция удалена из корзины.")
            show_cart(vk, user_id)
    elif action == 'cart_back':
        show_cart(vk, user_id)
    elif action == 'order_skip_comment':
        handle_order_skip_comment(vk, user_id)
    elif action == 'order_redo':
        handle_order_redo(vk, user_id)
    elif action == 'order_confirm':
        handle_order_confirm(vk, user_id)
    elif action == 'order_copy_code':
        order_code = payload.get('code')
        if order_code:
            handle_order_copy_code(vk, user_id, order_code)
    elif action == 'order_detail':
        order_code = payload.get('order_code')
        if order_code:
            show_order_items_list(vk, user_id, order_code)
    elif action == 'order_contact_manager':
        send_manager_contact(vk, user_id)
    elif action == 'order_item_detail':
        product_code = payload.get('product_code')
        material_code = payload.get('material_code')
        if product_code and material_code:
            show_order_item_detail(vk, user_id, product_code, material_code)
    elif action == 'order_detail_back':
        show_user_orders(vk, user_id)
    elif action == 'order_detail_main_menu':
        handle_start(vk, user_id)
    else:
        return

    try:
        vk.messages.sendMessageEventAnswer(
            event_id=event.event_id,
            user_id=user_id,
            peer_id=event.peer_id
        )
    except Exception as e:
        print(f"Ошибка при ответе на callback: {e}")


# ================== ГЛАВНЫЙ ОБРАБОТЧИК СООБЩЕНИЙ ==================
def handle_message(vk, message):
    user_id = message['from_id']
    text = message.get('text', '').strip()
    text_lower = text.lower()

    if text_lower in ('начать', 'start', '/start', 'перезапустить бота'):
        handle_start(vk, user_id)
        return

    if text_lower == 'перейти на сайт':
        go_to_site(vk, user_id)
        send_message(vk, user_id, "Главное меню:", create_main_keyboard())
        return

    if text_lower == 'связаться с менеджером':
        handle_contact_manager(vk, user_id)
        return

    if text == "Добавить позицию к заказу":
        show_adding_item_menu(vk, user_id, from_state=user_states.get(user_id, {}).get('state', 'main_menu'))
        return

    if text == "Главное меню":
        handle_start(vk, user_id)
        return

    state_info = user_states.get(user_id, {'state': 'main_menu', 'data': {}})
    state = state_info['state']
    data = state_info['data']

    if state == 'main_menu':
        handle_main_menu(vk, user_id, text)
    elif state == 'adding_item_menu':
        handle_adding_item_menu(vk, user_id, text)
    elif state == 'entering_product_code':
        handle_entering_product_code(vk, user_id, text, data)
    elif state == 'selecting_product_type':
        handle_selecting_product_type(vk, user_id, text, data)
    elif state == 'browsing_products':
        handle_browsing_products(vk, user_id, text, data)
    elif state == 'viewing_product_by_code':
        handle_viewing_product_by_code(vk, user_id, text, data)
    elif state == 'choosing_material_method':
        handle_choosing_material_method(vk, user_id, text, data)
    elif state == 'entering_material_code':
        handle_entering_material_code(vk, user_id, text, data)
    elif state == 'viewing_material_by_code':
        handle_viewing_material_by_code(vk, user_id, text, data)
    elif state == 'selecting_material_type':
        handle_selecting_material_type(vk, user_id, text, data)
    elif state == 'browsing_materials':
        handle_browsing_materials(vk, user_id, text, data)
    elif state == 'post_add_menu':
        handle_post_add_menu(vk, user_id, text)
    elif state == 'selecting_sofa_type':
        handle_selecting_sofa_type(vk, user_id, text, data)
    elif state == 'selecting_sofa_module':
        handle_selecting_sofa_module(vk, user_id, text, data)
    elif state == 'constructing_sofa':
        handle_constructing_sofa(vk, user_id, text, data)
    elif state == 'entering_sofa_code':
        handle_entering_sofa_code(vk, user_id, text)
    elif state == 'changing_material_for_cart':
        handle_choosing_material_method(vk, user_id, text, data)
    elif state == 'collecting_order_name':
        handle_collecting_order_name(vk, user_id, text)
    elif state == 'collecting_order_phone':
        handle_collecting_order_phone(vk, user_id, text)
    elif state == 'collecting_order_address':
        handle_collecting_order_address(vk, user_id, text)
    elif state == 'collecting_order_comment':
        handle_collecting_order_comment(vk, user_id, text)
    elif state == 'confirm_order_details':
        send_message(vk, user_id, "Пожалуйста, используйте кнопки для подтверждения или изменения данных.")
    elif state == 'entering_order_code':
        handle_entering_order_code(vk, user_id, text)
    else:
        handle_start(vk, user_id)


# ================== ОСНОВНОЙ ЦИКЛ ==================
def main():
    ensure_dirs()
    create_cart_table()
    create_orders_table()

    vk_session = vk_api.VkApi(token=TOKEN)
    vk = vk_session.get_api()
    longpoll = VkBotLongPoll(vk_session, GROUP_ID)

    print("Бот запущен и ожидает команды...")

    for event in longpoll.listen():
        try:
            if event.type == VkBotEventType.MESSAGE_NEW and event.object.message:
                handle_message(vk, event.object.message)
            elif event.type == VkBotEventType.MESSAGE_EVENT:
                handle_callback(vk, event.object)
        except Exception as e:
            print("Ошибка:")
            traceback.print_exc()
            if event.type == VkBotEventType.MESSAGE_NEW and event.object.message:
                send_message(vk, event.object.message['from_id'], "Произошла внутренняя ошибка. Попробуйте позже.")


if __name__ == '__main__':
    main()